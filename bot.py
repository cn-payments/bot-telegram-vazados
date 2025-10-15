import json
import logging
import os
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, Bot
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, ContextTypes, MessageHandler, filters, JobQueue
import qrcode
from PIL import Image
from urllib.parse import urlparse
import io
import mercadopago
import asyncio
import threading
import time
from threading import Thread
from database import Database
import requests
import hashlib
import hmac
import queue
import tempfile
import shutil
from flask import Flask, jsonify
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import sqlite3

# Importações para processamento de vídeo
# Importar funções de processamento de vídeo
try:
    from video_processor import process_video_for_telegram, validate_video_for_telegram
    VIDEO_PROCESSOR_AVAILABLE = True
except ImportError:
    VIDEO_PROCESSOR_AVAILABLE = False
    logging.warning("Módulo de processamento de vídeo não disponível.")

# Fila global para eventos de entrega de acesso VIP
access_delivery_queue = queue.Queue()

# Configuração de logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Variável global para a instância do bot
_bot_instance = None
_bot_context = None
_application_instance = None

def get_bot_instance():
    """Retorna a instância global do bot"""
    global _bot_instance
    return _bot_instance

def set_bot_instance(bot):
    """Define a instância global do bot"""
    global _bot_instance
    _bot_instance = bot

def get_bot_context():
    """Retorna o contexto global do bot"""
    global _bot_context
    return _bot_context

def set_bot_context(context):
    """Define o contexto global do bot"""
    global _bot_context
    _bot_context = context

def get_application_instance():
    """Retorna a instância da aplicação"""
    global _application_instance
    return _application_instance

def set_application_instance(application):
    """Define a instância da aplicação"""
    global _application_instance
    _application_instance = application

def create_bot_context():
    """Cria um contexto do bot que pode ser usado por outras threads"""
    global _bot_context, _bot_instance, _application_instance
    
    if _application_instance and _bot_instance:
        # Criar um contexto que pode ser usado por outras threads
        from telegram.ext import ContextTypes
        context = ContextTypes.DEFAULT_TYPE(_application_instance)
        # Não definir context.bot ou context.application diretamente pois são propriedades sem setter
        _bot_context = context
        return context
    return None

def is_admin(user_id):
    """Verifica se um user_id está na tabela de admins"""
    db = Database()
    try:
        # Conecta ao banco
        conn = db.connect()
        if conn is None:
            logger.error("[ADMIN CHECK] Falha ao conectar ao banco de dados!")
            return False

        # Log detalhado
        logger.info(f"[ADMIN CHECK] Verificando admin_id={user_id} (type={type(user_id)})")
        
        # Executa a query - garantindo que comparamos os mesmos tipos
        result = db.execute_fetch_one(
            "SELECT 1 FROM admins WHERE admin_id = %s", 
            (int(user_id),)  # Garante que comparamos como inteiro
        )
        
        logger.info(f"[ADMIN CHECK] Resultado: {result}")
        return result is not None
        
    except Exception as e:
        logger.error(f"[ADMIN CHECK] Erro ao verificar admin: {str(e)}")
        return False
    finally:
        db.close()

def get_all_admin_ids():
    db = Database()
    db.connect()
    admins = db.execute_fetch_all("SELECT admin_id FROM admins")
    db.close()
    return [a['admin_id'] for a in admins if a['admin_id']]

def add_admin(user_id, added_by, username=None):
    db = Database()
    conn = db.connect()
    if not conn:
        return False
    try:
        return db.execute_query(
            "INSERT IGNORE INTO admins (user, admin_id) VALUES (%s, %s)",
            (username or str(user_id), str(user_id)),
            commit=True
        )
    finally:
        db.close()

def remove_admin(user_id):
    db = Database()
    conn = db.connect()
    if not conn:
        return False
    try:
        return db.execute_query(
            "DELETE FROM admins WHERE admin_id = %s",
            (str(user_id),),
            commit=True
        )
    finally:
        db.close()

class SharedBotContext:
    """Classe para gerenciar contexto compartilhado entre threads"""
    
    def __init__(self):
        self.bot = None
        self.context = None
        self.application = None
        self._lock = threading.Lock()
    
    def set_context(self, bot, application, context):
        """Define o contexto compartilhado"""
        with self._lock:
            self.bot = bot
            self.application = application
            self.context = context
    
    def get_bot(self):
        """Retorna a instância do bot"""
        with self._lock:
            return self.bot
    
    def get_context(self):
        """Retorna o contexto"""
        with self._lock:
            return self.context
    
    def get_application(self):
        """Retorna a aplicação"""
        with self._lock:
            return self.application
    
    def is_available(self):
        """Verifica se o contexto está disponível"""
        with self._lock:
            return self.bot is not None and self.context is not None
    
    async def send_message(self, chat_id, text, **kwargs):
        """Envia mensagem usando o contexto compartilhado"""
        with self._lock:
            if self.bot:
                return await self.bot.send_message(chat_id=chat_id, text=text, **kwargs)
            else:
                raise RuntimeError("Bot não disponível")

# Instância global do contexto compartilhado
_shared_context = SharedBotContext()

def get_shared_context():
    """Retorna a instância global do contexto compartilhado"""
    return _shared_context

def set_shared_context(bot, application, context):
    """Define o contexto compartilhado global"""
    _shared_context.set_context(bot, application, context)

# =====================================================
# FUNÇÕES AUXILIARES PARA BANCO DE DADOS
# =====================================================

def get_user_subscriptions(user_id):
    """Obtém todas as assinaturas de um usuário"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        subscriptions = db.execute_fetch_all(
            """SELECT s.*, vp.name as plan_name, vp.price, vp.duration_days
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            WHERE s.user_id = %s
            ORDER BY s.created_at DESC""",
            (user_id,)
        )
        return subscriptions
    except Exception as e:
        logger.error(f"Erro ao obter assinaturas do usuário {user_id}: {e}")
        return []
    finally:
        db.close()

def get_active_subscription(user_id):
    """Obtém a assinatura ativa de um usuário"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return None
        subscription = db.execute_fetch_one(
            """SELECT s.*, vp.name as plan_name, vp.price, vp.duration_days
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            WHERE s.user_id = %s 
            AND s.is_active = TRUE
            AND (s.is_permanent = TRUE OR s.end_date > NOW())
            ORDER BY s.end_date DESC
            LIMIT 1""",
            (user_id,)
        )
        return subscription
    except Exception as e:
        logger.error(f"Erro ao obter assinatura ativa do usuário {user_id}: {e}")
        return None
    finally:
        db.close()

def check_payment_processed(payment_id):
    """Verifica se um pagamento já foi processado"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        result = db.execute_fetch_one(
            "SELECT id FROM subscriptions WHERE payment_id = %s",
            (payment_id,)
        )
        return result is not None
    except Exception as e:
        logger.error(f"Erro ao verificar pagamento {payment_id}: {e}")
        return False
    finally:
        db.close()

def get_all_active_subscriptions():
    """Obtém todas as assinaturas ativas"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        subscriptions = db.execute_fetch_all(
            """SELECT s.*, vp.name as plan_name, vp.price, vp.duration_days, u.username, u.first_name, u.last_name
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            JOIN users u ON s.user_id = u.id
            WHERE s.is_active = TRUE
            AND (s.is_permanent = TRUE OR s.end_date > NOW())
            ORDER BY s.end_date ASC"""
        )
        return subscriptions
    except Exception as e:
        logger.error(f"Erro ao obter assinaturas ativas: {e}")
        return []
    finally:
        db.close()

def get_subscriptions_for_export():
    """Obtém assinaturas com informações detalhadas para exportação"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        
        # Query mais detalhada para exportação
        subscriptions = db.execute_fetch_all(
            """SELECT 
                s.id as subscription_id,
                s.user_id,
                s.plan_id,
                s.payment_id,
                s.payment_method,
                s.payment_status,
                s.start_date,
                s.end_date,
                s.is_permanent,
                s.is_active,
                s.created_at,
                vp.name as plan_name,
                vp.price,
                vp.duration_days,
                u.username,
                u.first_name,
                u.last_name,
                u.joined_date,
                -- Calcular dias restantes
                CASE 
                    WHEN s.is_permanent = TRUE THEN 999999
                    WHEN s.end_date > NOW() THEN DATEDIFF(s.end_date, NOW())
                    ELSE 0
                END as days_remaining,
                -- Calcular dias já pagos
                CASE 
                    WHEN s.is_permanent = TRUE THEN 999999
                    WHEN s.start_date <= NOW() AND s.end_date > NOW() THEN DATEDIFF(NOW(), s.start_date)
                    WHEN s.end_date <= NOW() THEN DATEDIFF(s.end_date, s.start_date)
                    ELSE 0
                END as days_paid,
                -- Calcular total de dias do plano
                CASE 
                    WHEN s.is_permanent = TRUE THEN 999999
                    WHEN vp.duration_days > 0 THEN vp.duration_days
                    ELSE DATEDIFF(s.end_date, s.start_date)
                END as total_days,
                -- Status de expiração
                CASE 
                    WHEN s.is_permanent = TRUE THEN 'Permanente'
                    WHEN s.end_date <= NOW() THEN 'Expirada'
                    WHEN DATEDIFF(s.end_date, NOW()) <= 3 THEN 'Expirando em breve'
                    WHEN DATEDIFF(s.end_date, NOW()) <= 7 THEN 'Expira em 1 semana'
                    ELSE 'Ativa'
                END as expiration_status
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            JOIN users u ON s.user_id = u.id
            WHERE s.is_active = TRUE
            ORDER BY s.end_date ASC"""
        )
        return subscriptions
    except Exception as e:
        logger.error(f"Erro ao obter assinaturas para exportação: {e}")
        return []
    finally:
        db.close()

# ===== FUNÇÕES PARA MENSAGENS AGENDADAS =====

def create_scheduled_message(message_text, scheduled_date, target_type, target_users=None, created_by=None):
    """Cria uma nova mensagem agendada"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        # Converter target_users para JSON se for uma lista
        target_users_json = None
        if target_users and isinstance(target_users, list):
            target_users_json = json.dumps(target_users)
        elif target_users and isinstance(target_users, str):
            target_users_json = target_users
        
        result = db.execute_query(
            """INSERT INTO scheduled_messages 
            (message_text, scheduled_date, target_type, target_users, created_by) 
            VALUES (%s, %s, %s, %s, %s)""",
            (message_text, scheduled_date, target_type, target_users_json, created_by),
            commit=True
        )
        return result
    except Exception as e:
        logger.error(f"Erro ao criar mensagem agendada: {e}")
        return False
    finally:
        db.close()

def get_scheduled_messages(status=None, limit=None):
    """Obtém mensagens agendadas"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        
        query = "SELECT * FROM scheduled_messages"
        params = []
        
        if status:
            query += " WHERE status = %s"
            params.append(status)
        
        query += " ORDER BY scheduled_date ASC"
        
        if limit:
            query += " LIMIT %s"
            params.append(limit)
        
        messages = db.execute_fetch_all(query, params)
        return messages
    except Exception as e:
        logger.error(f"Erro ao obter mensagens agendadas: {e}")
        return []
    finally:
        db.close()

def update_scheduled_message_status(message_id, status, error_message=None, successful_sends=0, failed_sends=0):
    """Atualiza o status de uma mensagem agendada"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        if status == 'sent':
            db.execute_query(
                """UPDATE scheduled_messages 
                SET status = %s, sent_at = NOW(), successful_sends = %s, failed_sends = %s 
                WHERE id = %s""",
                (status, successful_sends, failed_sends, message_id),
                commit=True
            )
        else:
            db.execute_query(
                """UPDATE scheduled_messages 
                SET status = %s, error_message = %s 
                WHERE id = %s""",
                (status, error_message, message_id),
                commit=True
            )
        return True
    except Exception as e:
        logger.error(f"Erro ao atualizar status da mensagem agendada: {e}")
        return False
    finally:
        db.close()

def cancel_scheduled_message(message_id):
    """Cancela uma mensagem agendada"""
    return update_scheduled_message_status(message_id, 'cancelled')

def get_pending_scheduled_messages():
    """Obtém mensagens agendadas pendentes que devem ser enviadas
    Inclui mensagens com até 30 minutos de atraso para compensar downtime do bot"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        
        messages = db.execute_fetch_all(
            """SELECT * FROM scheduled_messages 
            WHERE status = 'pending' 
            AND scheduled_date <= NOW() + INTERVAL 30 MINUTE
            AND scheduled_date >= NOW() - INTERVAL 30 MINUTE
            ORDER BY scheduled_date ASC"""
        )
        return messages
    except Exception as e:
        logger.error(f"Erro ao obter mensagens pendentes: {e}")
        return []
    finally:
        db.close()

def cleanup_old_scheduled_messages():
    """Remove mensagens agendadas muito antigas (mais de 30 minutos) que não foram processadas"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        result = db.execute_query(
            """UPDATE scheduled_messages 
            SET status = 'cancelled', 
                error_message = 'Mensagem muito antiga - removida automaticamente'
            WHERE status = 'pending' 
            AND scheduled_date < NOW() - INTERVAL 30 MINUTE""",
            commit=True
        )
        
        if result:
            logger.info("🧹 Mensagens agendadas muito antigas foram canceladas automaticamente")
        
        return result
    except Exception as e:
        logger.error(f"Erro ao limpar mensagens antigas: {e}")
        return False
    finally:
        db.close()

def check_duplicate_scheduled_message(target_type, target_users=None):
    """Verifica se já existe uma mensagem agendada pendente para o mesmo tipo de destinatários"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        if target_type == 'specific_users' and target_users:
            # Para usuários específicos, verificar se há mensagem com os mesmos usuários
            target_users_json = json.dumps(target_users) if isinstance(target_users, list) else target_users
            existing = db.execute_fetch_one(
                """SELECT id FROM scheduled_messages 
                WHERE status = 'pending' 
                AND target_type = 'specific_users' 
                AND target_users = %s""",
                (target_users_json,)
            )
        else:
            # Para all_users e vip_users, verificar se há mensagem com o mesmo tipo
            existing = db.execute_fetch_one(
                """SELECT id FROM scheduled_messages 
                WHERE status = 'pending' 
                AND target_type = %s""",
                (target_type,)
            )
        
        return existing is not None
    except Exception as e:
        logger.error(f"Erro ao verificar mensagem duplicada: {e}")
        return False
    finally:
        db.close()

def is_admin_vip(admin_id):
    """Verifica se o admin é VIP"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        admin = db.execute_fetch_one(
            "SELECT is_vip FROM admins WHERE admin_id = %s",
            (admin_id,)
        )
        return admin and admin['is_vip'] == 1
    except Exception as e:
        logger.error(f"Erro ao verificar admin VIP: {e}")
        return False
    finally:
        db.close()

def get_recipients_for_scheduled_message(scheduled_message):
    """Obtém a lista de destinatários para uma mensagem agendada"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        
        target_type = scheduled_message['target_type']
        
        if target_type == 'all_users':
            users = db.execute_fetch_all("SELECT id FROM users")
            return [user['id'] for user in users]
        
        elif target_type == 'vip_users':
            users = db.execute_fetch_all(
                """SELECT u.id FROM users u 
                JOIN subscriptions s ON u.id = s.user_id 
                WHERE s.is_active = TRUE 
                AND (s.is_permanent = TRUE OR s.end_date > NOW())"""
            )
            return [user['id'] for user in users]
        
        elif target_type == 'specific_users':
            if scheduled_message['target_users']:
                try:
                    user_ids = json.loads(scheduled_message['target_users'])
                    return user_ids if isinstance(user_ids, list) else []
                except:
                    return []
        
        return []
    except Exception as e:
        logger.error(f"Erro ao obter destinatários: {e}")
        return []
    finally:
        db.close()

def get_expiring_subscriptions():
    """Obtém assinaturas próximas de expirar (3 dias ou menos)"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return []
        subscriptions = db.execute_fetch_all(
            """SELECT s.*, vp.name as plan_name, vp.price, vp.duration_days, u.username, u.first_name, u.last_name
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            JOIN users u ON s.user_id = u.id
            WHERE s.is_active = TRUE
            AND s.is_permanent = FALSE
            AND s.end_date > NOW()
            AND s.end_date <= DATE_ADD(NOW(), INTERVAL 3 DAY)
            ORDER BY s.end_date ASC"""
        )
        return subscriptions
    except Exception as e:
        logger.error(f"Erro ao obter assinaturas próximas de expirar: {e}")
        return []
    finally:
        db.close()

def update_subscription_notification(subscription_id, notification_type):
    """Atualiza o status de notificação de uma assinatura"""
    db = Database()
    try:
        db.connect()
        if notification_type == "notified_1":
            db.execute_query(
                "UPDATE subscriptions SET notified_1 = TRUE WHERE id = %s",
                (subscription_id,),
                commit=True
            )
        elif notification_type == "notified_2":
            db.execute_query(
                "UPDATE subscriptions SET notified_2 = TRUE WHERE id = %s",
                (subscription_id,),
                commit=True
            )
        elif notification_type == "notified_3":
            db.execute_query(
                "UPDATE subscriptions SET notified_3 = TRUE WHERE id = %s",
                (subscription_id,),
                commit=True
            )
        return True
    except Exception as e:
        logger.error(f"Erro ao atualizar notificação {notification_type} para assinatura {subscription_id}: {e}")
        return False
    finally:
        db.close()

def get_user_stats():
    """Obtém estatísticas dos usuários"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return {'total_users': 0, 'vip_users': 0, 'recent_users': []}
        
        # Total de usuários
        total_result = db.execute_fetch_one("SELECT COUNT(*) as total FROM users")
        total_users = total_result['total'] if total_result else 0
        
        # Usuários VIP
        vip_result = db.execute_fetch_one("SELECT COUNT(*) as total FROM users WHERE is_vip = TRUE")
        vip_users = vip_result['total'] if vip_result else 0
        
        # Últimos usuários
        recent_users = db.execute_fetch_all(
            """SELECT id, username, first_name, last_name, joined_date, is_vip
            FROM users
            ORDER BY joined_date DESC
            LIMIT 5"""
        )
        
        return {
            'total_users': total_users,
            'vip_users': vip_users,
            'recent_users': recent_users
        }
    except Exception as e:
        logger.error(f"Erro ao obter estatísticas: {e}")
        return {'total_users': 0, 'vip_users': 0, 'recent_users': []}
    finally:
        db.close()

def get_all_users():
    """Obtém todos os usuários"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            logger.error("❌ Falha na conexão com banco de dados em get_all_users")
            return []
        
        logger.info("🔍 Executando consulta get_all_users...")
        users = db.execute_fetch_all(
            "SELECT id, username, first_name, last_name, joined_date, is_vip FROM users"
        )
        logger.info(f"✅ get_all_users retornou {len(users)} usuários")
        
        # Log dos primeiros 3 usuários para debug
        for i, user in enumerate(users[:3]):
            logger.info(f"   Usuário {i+1}: ID={user['id']}, Nome={user['first_name']}, VIP={user['is_vip']}")
        
        return users
    except Exception as e:
        logger.error(f"❌ Erro ao obter usuários: {e}")
        return []
    finally:
        db.close()

def get_vip_users():
    """Obtém todos os usuários VIP ativos"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            logger.error("❌ Falha na conexão com banco de dados em get_vip_users")
            return []
        
        logger.info("🔍 Executando consulta get_vip_users...")
        users = db.execute_fetch_all(
            """SELECT DISTINCT u.id, u.username, u.first_name, u.last_name, u.joined_date
            FROM users u
            JOIN subscriptions s ON u.id = s.user_id
            WHERE s.is_active = TRUE
            AND (s.is_permanent = TRUE OR s.end_date > NOW())"""
        )
        logger.info(f"✅ get_vip_users retornou {len(users)} usuários VIP")
        
        # Log dos primeiros 3 usuários VIP para debug
        for i, user in enumerate(users[:3]):
            logger.info(f"   VIP {i+1}: ID={user['id']}, Nome={user['first_name']}")
        
        return users
    except Exception as e:
        logger.error(f"❌ Erro ao obter usuários VIP: {e}")
        return []
    finally:
        db.close()

def is_valid_url(url):
    try:
        parsed = urlparse(url)
        return all([parsed.scheme in ("http", "https"), parsed.netloc])
    except Exception:
        return False

def load_messages_from_db():
    """Carrega mensagens do banco de dados"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return {}
        rows = db.execute_fetch_all(
            "SELECT message_key, message_value FROM bot_messages WHERE language = 'pt-BR'"
        )
        messages = {row['message_key']: row['message_value'] for row in rows}
        return messages
    except Exception as e:
        logger.error(f"Erro ao carregar mensagens do banco: {e}")
        return {}
    finally:
        db.close()

def save_config_to_db(key, value):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        db.execute_query(
            """INSERT INTO bot_config (`config_key`, `config_value`) VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE `config_value` = %s""",
            (key, value, value),
            commit=True
        )
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar configuração no banco: {e}")
        return False
    finally:
        db.close()


def save_message_to_db(message_key, message_value):
    """Salva uma mensagem no banco de dados"""
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        db.execute_query(
            """INSERT INTO bot_messages (message_key, message_value, language) 
            VALUES (%s, %s, 'pt-BR')
            ON DUPLICATE KEY UPDATE message_value = %s""",
            (message_key, message_value, message_value),
            commit=True
        )
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar mensagem {message_key}: {e}")
        return False
    finally:
        db.close()

# =====================================================
# FIM DAS FUNÇÕES AUXILIARES
# =====================================================

# Carregar configurações
def load_config():
    db = Database()
    try:
        db.connect()
        if not db.connection:
            logger.error("Não foi possível conectar ao banco de dados")
            return None
            
        # Usar o novo método que fecha o cursor automaticamente
        rows = db.execute_fetch_all("SELECT config_key, config_value, config_type FROM bot_config")
        
        config = {}
        for row in rows:
            key = row['config_key']
            value = row['config_value']
            config_type = row['config_type']
            # Conversão de tipo
            if config_type == 'boolean':
                config[key] = value.lower() == 'true'
            elif config_type == 'integer':
                config[key] = int(value)
            elif config_type == 'json':
                try:
                    # Limpar caracteres de controle inválidos
                    clean_value = ''.join(char for char in value if ord(char) >= 32 or char in '\t\n\r')
                    config[key] = json.loads(clean_value)
                except json.JSONDecodeError as json_error:
                    logger.warning(f"Erro ao fazer parse JSON para {key}: {json_error}")
                    # Tentar limpar mais caracteres problemáticos
                    try:
                        # Remover caracteres de controle mais agressivamente
                        clean_value = ''.join(char for char in value if ord(char) >= 32)
                        config[key] = json.loads(clean_value)
                    except:
                        logger.error(f"Não foi possível fazer parse JSON para {key}, usando valor como string")
                        config[key] = value
            else:
                config[key] = value
        return config
    except Exception as e:
        logger.error(f"Erro ao carregar configuração: {e}")
        return None
    finally:
        db.close()

def save_config(config):
    db = Database()
    try:
        db.connect()
        for key, value in config.items():
            # Determinar o tipo do valor
            if isinstance(value, bool):
                val = str(value).lower()
                typ = 'boolean'
            elif isinstance(value, int):
                val = str(value)
                typ = 'integer'
            elif isinstance(value, dict) or isinstance(value, list):
                val = json.dumps(value)
                typ = 'json'
            else:
                val = str(value)
                typ = 'string'
            
            db.execute_query(
                """INSERT INTO bot_config (config_key, config_value, config_type) 
                VALUES (%s, %s, %s) 
                ON DUPLICATE KEY UPDATE config_value = %s, config_type = %s""",
                (key, val, typ, val, typ),
                commit=True
            )
        return True
    except Exception as e:
        logger.error(f"Erro ao salvar configuração: {e}")
        return False
    finally:
        db.close()
# Editar uma configuração específica
def edit_config(key, value):
    try:
        logger.info(f"Iniciando edição de {key} com valor: {value}")
        config = load_config()
        if not config:
            logger.error("Não foi possível carregar o config.json")
            return False
        
        # Navega pela estrutura do JSON usando a chave
        keys = key.split('.')
        current = config
        for k in keys[:-1]:
            if k not in current:
                current[k] = {}
            current = current[k]
        
        # Atualiza o valor
        current[keys[-1]] = value
        logger.info(f"Valor atualizado na memória: {current[keys[-1]]}")
        
        # Salva as alterações
        if save_config(config):
            logger.info("Configuração salva com sucesso")
            return True
        else:
            logger.error("Erro ao salvar configuração")
            return False
            
    except Exception as e:
        logger.error(f"Erro ao editar config.json: {e}")
        return False

# Exemplo de uso:
# edit_config('admin_settings.welcome_message', 'Nova mensagem de boas-vindas')
# edit_config('mercadopago.access_token', 'Novo token')
# edit_config('payment_methods.pix_manual.chave_pix', 'Nova chave PIX')

# Verificar pagamento no Mercado Pago
def check_payment(payment_id):
    """Função mantida para compatibilidade - agora usa o sistema de provedores"""
    logger.warning("check_payment está depreciada. Use check_payment_async()")
    return None

# Nova função assíncrona para verificar pagamento
async def check_payment_async(payment_id, provider=None):
    """Verifica pagamento usando o sistema de provedores"""
    try:
        provider_manager = get_pix_provider_manager()
        result = await provider_manager.check_payment_with_fallback(payment_id, provider)
        
        if result:
            logger.info(f"Pagamento verificado com sucesso")
            return result
        else:
            logger.error(f"Falha ao verificar pagamento {payment_id}")
            return None
            
    except Exception as e:
        logger.error(f"Erro ao verificar pagamento: {e}")
        return None

# Registrar assinatura VIP
async def register_vip_subscription(user_id, plan_id, payment_id, context):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        # Obter informações do plano
        plan = db.execute_fetch_one(
            "SELECT * FROM vip_plans WHERE id = %s",
            (plan_id,)
        )
        
        if not plan:
            return False
        
        # Calcular data de expiração
        if plan['duration_days'] == -1:
            end_date = datetime(2099, 12, 31)
            is_permanent = True
        else:
            end_date = datetime.now() + timedelta(days=plan['duration_days'])
            is_permanent = False
        
        # Inserir assinatura
        db.execute_query(
            """INSERT INTO subscriptions 
            (user_id, plan_id, payment_id, payment_method, payment_status, 
             start_date, end_date, is_permanent, is_active) 
            VALUES (%s, %s, %s, 'mercadopago', 'approved', NOW(), %s, %s, TRUE)""",
            (user_id, plan_id, payment_id, end_date, is_permanent),
            commit=True
        )
        
        # Atualizar status VIP do usuário
        db.execute_query(
            "UPDATE users SET is_vip = TRUE WHERE id = %s",
            (user_id,),
            commit=True
        )
        
        # Notificar admin
        config = load_config()
        if config is None:
            logger.error("Falha ao carregar as configurações.")
            return  # ou lidar de forma apropriada

        if config and 'admin_id' in config:
            if not is_admin(int(update.effective_user.id)):
                return
            admin_message = (
                f"🎉 Nova Assinatura VIP!\n\n"
                f"👤 Usuário: {user_id}\n"
                f"💎 Plano: {plan['name']}\n"
                f"💰 Valor: R${plan['price']:.2f}\n"
                f"⏱️ Duração: {'Permanente' if is_permanent else str(plan['duration_days']) + ' dias'}\n"
                f"📅 Expira em: {end_date.strftime('%d/%m/%Y %H:%M')}\n"
                f"💳 ID do Pagamento: {payment_id}"
            )
            await context.bot.send_message(chat_id=admin_id, text=admin_message)
        
        return True
        
    except Exception as e:
        logger.error(f"Erro ao registrar assinatura: {e}")
        return False
    finally:
        db.close()

async def renew_vip_subscription(user_id, plan_id, payment_id, context):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        # Obter informações do plano
        plan = db.execute_fetch_one(
            "SELECT * FROM vip_plans WHERE id = %s",
            (plan_id,)
        )
        
        if not plan:
            return False
        
        # Encontrar assinatura atual ativa
        current_subscription = db.execute_fetch_one(
            """SELECT * FROM subscriptions 
            WHERE user_id = %s 
            AND is_active = TRUE
            AND (is_permanent = TRUE OR end_date > NOW())
            ORDER BY end_date DESC
            LIMIT 1""",
            (user_id,)
        )
        
        if not current_subscription:
            logger.error(f"Tentativa de renovação sem assinatura ativa: usuário {user_id}")
            return False
        
        # Calcular nova data de expiração
        if plan['duration_days'] == -1:
            # Plano permanente
            end_date = datetime(2099, 12, 31)
            is_permanent = True
        else:
            # Renovação - soma os dias à data atual de expiração
            current_end_date = current_subscription['end_date']
            if isinstance(current_end_date, str):
                current_end_date = datetime.strptime(current_end_date, "%Y-%m-%d %H:%M:%S")
            end_date = current_end_date + timedelta(days=plan['duration_days'])
            is_permanent = False
            
            days_left = (current_end_date - datetime.now()).days
            logger.info(f"Renovação detectada. Dias restantes: {days_left}, Novos dias: {plan['duration_days']}, Total: {days_left + plan['duration_days']}")
        
        # Desativar assinatura atual
        db.execute_query(
            "UPDATE subscriptions SET is_active = FALSE WHERE id = %s",
            (current_subscription['id'],),
            commit=True
        )
        
        # Inserir nova assinatura
        db.execute_query(
            """INSERT INTO subscriptions 
            (user_id, plan_id, payment_id, payment_method, payment_status, 
             start_date, end_date, is_permanent, is_active,
             notified_1, notified_2, notified_3, renewal_notified) 
            VALUES (%s, %s, %s, 'mercadopago', 'approved', NOW(), %s, %s, TRUE, FALSE, FALSE, FALSE, FALSE)""",
            (user_id, plan_id, payment_id, end_date, is_permanent),
            commit=True
        )
        
        logger.info(f"Renovação registrada: usuário {user_id}, plano {plan_id}")
        logger.info(f"Nova data de expiração: {end_date}")
        logger.info(f"Notificações de expiração limpas para o usuário {user_id}")

        # Notificar admin
        try:
            config = load_config()
            if config is None:
                logger.error("Falha ao carregar as configurações.")
            return  # ou lidar de forma apropriada
            if config and 'admin_id' in config:
                if not is_admin(int(update.effective_user.id)):
                    return
                admin_message = (
                    f"🔄 Renovação de Assinatura VIP!\n\n"
                    f"👤 Usuário: {user_id}\n"
                    f"💎 Plano: {plan['name']}\n"
                    f"💰 Valor: R${plan['price']:.2f}\n"
                    f"⏱️ Duração: {'Permanente' if plan['duration_days'] == -1 else str(plan['duration_days']) + ' dias'}\n"
                    f"📅 Nova expiração: {end_date.strftime('%d/%m/%Y %H:%M')}\n"
                    f"💳 ID do Pagamento: {payment_id}"
                )
                await context.bot.send_message(chat_id=admin_id, text=admin_message)
        except Exception as e:
            logger.error(f"Erro ao notificar admin sobre renovação: {e}")

        return True
        
    except Exception as e:
        logger.error(f"Erro ao renovar assinatura: {e}")
        return False
    finally:
        db.close()

# Adicionar usuário aos grupos VIP
async def add_user_to_vip_groups(bot, user_id, plan_id):
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    
    # Buscar o plano no banco de dados
    db = Database()
    try:
        db.connect()
        if not db.connection:
            logger.error("Erro ao conectar ao banco de dados")
            return False
        
        # Buscar o plano
        plan = db.execute_fetch_one(
            "SELECT * FROM vip_plans WHERE id = %s",
            (plan_id,)
        )
        
        if not plan:
            logger.error(f"Plano {plan_id} não encontrado no banco de dados")
            return False
        
        # Buscar grupos associados ao plano
        groups = db.execute_fetch_all(
            """SELECT vg.group_id, vg.group_name
            FROM vip_groups vg
            JOIN plan_groups pg ON vg.id = pg.group_id
            WHERE pg.plan_id = %s AND vg.is_active = TRUE""",
            (plan_id,)
        )
        
        if not groups:
            logger.info(f"Nenhum grupo encontrado para o plano {plan_id}")
            return True  # Retorna True mesmo sem grupos
        
        # Calcular duração do link baseada no plano
        if plan['duration_days'] == -1:
            # Plano permanente - link de 30 dias (renovável)
            link_duration = 30
            link_message = "O link expira em 30 dias e pode ser renovado."
        else:
            # Plano temporário - link com duração igual ao plano
            link_duration = plan['duration_days']
            link_message = f"O link expira em {link_duration} dias (duração do seu plano)."
        
        # Adicionar usuário aos grupos
        for group in groups:
            group_id = group['group_id']
            group_name = group['group_name']
            
            try:
                # Verificar se o grupo é um supergrupo
                chat = await bot.get_chat(group_id)
                if chat.type in ['group', 'supergroup', 'channel']:
                    try:
                        # Criar link de convite com duração baseada no plano
                        invite_link = await bot.create_chat_invite_link(
                            chat_id=group_id,
                            name=f"VIP {user_id} - {plan['name']}",
                            expire_date=datetime.now() + timedelta(days=link_duration),
                            member_limit=1,
                            creates_join_request=False
                        )
                        
                        # Enviar link para o usuário
                        await bot.send_message(
                            chat_id=user_id,
                            text=f"⬇ ESTOU PELADINHA TE ESPERANDO 🙈\n\n"
                                 f"😈 Clique em \" VER CANAL \" pra gente começar a brincar 🔥\n\n"
                                 f"💎 VIP VAZADOS VIP 🍑🔥\n\n"
                                 f"📝 O link expira em {plan['duration_days']} dias (duração do seu plano).\n\n"
                                 f"⚠ Este link é único e só pode ser usado uma vez.\n\n"
                                 f"**Link:** {invite_link.invite_link}"
                        )
                        logger.info(f"Link de convite enviado para usuário {user_id} - grupo {group_id} (duração: {link_duration} dias)")
                        
                    except Exception as e:
                        logger.error(f"Erro ao criar link de convite para grupo {group_id}: {e}")
                        # Se falhar, tenta obter link existente
                        try:
                            invite_link = await bot.export_chat_invite_link(chat_id=group_id)
                            await bot.send_message(
                                chat_id=user_id,
                                text=f"⬇ ESTOU PELADINHA TE ESPERANDO 🙈\n\n"
                                     f"😈 Clique em \" VER CANAL \" pra gente começar a brincar 🔥\n\n"
                                     f"💎 VIP VAZADOS VIP 🍑🔥\n\n"
                                     f"📝 O link expira em {plan['duration_days']} dias (duração do seu plano).\n\n"
                                     f"⚠ Este link é único e só pode ser usado uma vez.\n\n"
                                     f"**Link:** {invite_link}"
                            )
                            logger.info(f"Link existente enviado para usuário {user_id} - grupo {group_id}")
                        except Exception as e2:
                            logger.error(f"Erro ao obter link existente: {e2}")
                            # Se tudo falhar, notifica o admin
                            if config and 'admin_id' in config:
                                if not is_admin(int(update.effective_user.id)):
                                    return
                                await bot.send_message(
                                    chat_id=admin_id,
                                    text=f"⚠️ Erro ao gerar link para usuário {user_id} no grupo {group_id}.\n"
                                         f"Erro: {e}\nErro do link: {e2}\n\n"
                                         f"Verifique se o bot tem permissões de administrador no grupo."
                                )
                else:
                    logger.error(f"Grupo {group_id} não é um grupo ou supergrupo válido")
                    # Notifica o admin
                    if config and 'admin_id' in config:
                        if not is_admin(int(update.effective_user.id)):
                            return
                        await bot.send_message(
                            chat_id=admin_id,
                            text=f"⚠️ Grupo {group_id} não é um grupo ou supergrupo válido.\nTipo: {chat.type}"
                        )
                        
            except Exception as e:
                logger.error(f"Erro ao processar grupo {group_id} para usuário {user_id}: {e}")
                # Notifica o admin
                if config and 'admin_id' in config:
                    if not is_admin(int(update.effective_user.id)):
                        return
                    await bot.send_message(
                        chat_id=admin_id,
                        text=f"⚠️ Erro ao processar grupo {group_id} para usuário {user_id}.\nErro: {e}"
                    )
        
        return True
        
    except Exception as e:
        logger.error(f"Erro ao buscar plano {plan_id} no banco de dados: {e}")
        return False
    finally:
        db.close()

# Gerar QR Code PIX do Mercado Pago (MANTIDO PARA COMPATIBILIDADE)
def generate_mercadopago_pix(amount, description, external_reference):
    """Função mantida para compatibilidade - agora usa o sistema de provedores"""
    logger.warning("generate_mercadopago_pix está depreciada. Use generate_pix_automatico()")
    return None

# Nova função unificada para gerar PIX automático
async def generate_pix_automatico(amount, description, external_reference):
    logger.info(f"[DEBUG] generate_pix_automatico chamado: amount={amount}, description={description}, external_reference={external_reference}")
    """Gera PIX usando o sistema de provedores com fallback"""
    try:
        provider_manager = get_pix_provider_manager()
        result = await provider_manager.generate_pix_with_fallback(amount, description, external_reference)
        
        if result:
            logger.info(f"PIX gerado com sucesso usando provedor: {result['provider']}")
            return result
        else:
            logger.error("Falha ao gerar PIX com todos os provedores")
            return None
            
    except Exception as e:
        logger.error(f"Erro ao gerar PIX automático: {e}")
        return None

# Gerar QR Code PIX
def generate_pix_qr_code(payment_data):
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(payment_data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Converter para bytes
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format='PNG')
    img_byte_arr.seek(0)
    return img_byte_arr

# Adicionar usuário às estatísticas
async def add_user_to_stats(user, bot):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            logger.error("❌ Falha na conexão com banco de dados em add_user_to_stats")
            return
        
        logger.info(f"🔍 Verificando usuário {user.id} em add_user_to_stats...")
        
        # Verificar se usuário já existe
        existing_user = db.execute_fetch_one(
            "SELECT id FROM users WHERE id = %s",
            (user.id,)
        )
        
        if not existing_user:
            logger.info(f"➕ Adicionando novo usuário {user.id} ao banco de dados...")
            db.execute_query(
                """INSERT INTO users 
                (id, username, first_name, last_name, joined_date) 
                VALUES (%s, %s, %s, %s, NOW())""",
                (user.id, user.username, user.first_name, user.last_name),
                commit=True
            )
            logger.info(f"✅ Usuário {user.id} adicionado com sucesso")
            
            # Notificar admin
            config = load_config()
            if config is None:
                logger.error("Falha ao carregar as configurações.")
                return  # ou lidar de forma apropriada
            if config and 'admin_id' in config:
                if not is_admin(int(update.effective_user.id)):
                    return
                msg = (
                    f"👤 Novo usuário acessou o bot!\n\n"
                    f"ID: {user.id}\n"
                    f"Nome: {user.first_name or ''} {user.last_name or ''}\n"
                    f"Username: @{user.username if user.username else '-'}\n"
                    f"Data de entrada: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
                )
                await bot.send_message(chat_id=admin_id, text=msg)
        else:
            logger.info(f"ℹ️ Usuário {user.id} já existe no banco de dados")
            
    except Exception as e:
        logger.error(f"❌ Erro ao adicionar usuário {user.id}: {e}")
    finally:
        db.close()

# Atualizar status VIP do usuário
async def update_user_vip_status(user_id, is_vip=True):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return False
        
        # Atualizar status VIP do usuário
        db.execute_query(
            "UPDATE users SET is_vip = %s WHERE id = %s",
            (is_vip, user_id),
            commit=True
        )
        
        logger.info(f"Status VIP atualizado para usuário {user_id}: {is_vip}")
        return True
        
    except Exception as e:
        logger.error(f"Erro ao atualizar status VIP: {e}")
        return False
    finally:
        db.close()

# Comandos do bot
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    config = load_config()

    messages = load_messages_from_db()
    
    # Verificar se messages foi carregado corretamente
    if not messages:
        logger.error("Falha ao carregar mensagens do banco de dados na função start")
        messages = {}  # Usar dicionário vazio como fallback
    
    if not config:
        await update.message.reply_text("Erro ao carregar configurações. Tente novamente mais tarde.")
        return
    
    if config.get('admin_settings', {}).get('maintenance_mode', False):
        await update.message.reply_text("🛠️ O bot está em manutenção. Tente novamente mais tarde.")
        return
    
    await add_user_to_stats(update.effective_user, context.bot)
    
    # Verificar se há arquivo de boas-vindas configurado
    welcome_file_config = config.get('welcome_file', {})
    if welcome_file_config.get('enabled', False) and welcome_file_config.get('file_id'):
        try:
            # Enviar arquivo de boas-vindas
            file_id = welcome_file_config['file_id']
            file_type = welcome_file_config.get('file_type', 'photo')
            caption = messages.get('welcome_message', 'sem mensagem definida')

            
            if file_type == 'photo':
                await context.bot.send_photo(
                    chat_id=update.effective_user.id,
                    photo=file_id,
                    caption=caption
                )
            elif file_type == 'video':
                await context.bot.send_video(
                    chat_id=update.effective_user.id,
                    video=file_id,
                    caption=caption
                )
        except Exception as e:
            logger.error(f"Erro ao enviar arquivo de boas-vindas: {e}")
            # Se falhar, continua com o fluxo normal
    
    db = Database()
    try:
        db.connect()
        if not db.connection:
            await update.message.reply_text("Erro de conexão com o banco de dados. Tente novamente mais tarde.")
            return
        
        # Verificar assinaturas ativas do usuário
        active_subs = db.execute_fetch_all(
            """SELECT s.*, vp.name as plan_name, vp.price, vp.duration_days
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            WHERE s.user_id = %s 
            AND s.is_active = TRUE
            AND (s.is_permanent = TRUE OR s.end_date > NOW())""",
            (update.effective_user.id,)
        )
        
        if active_subs:
            # Mostrar status da assinatura atual
            current_sub = active_subs[0]
            end_date = current_sub['end_date']
            time_left = end_date - datetime.now()
            days_left = time_left.days
            
            keyboard = []
            if days_left <= 3 and not current_sub['is_permanent']:
                keyboard.append([InlineKeyboardButton(
                    "🔄 Renovar Plano Atual",
                    callback_data=f"renew_{current_sub['plan_id']}"
                )])
            
            # Adicionar outros planos disponíveis
            other_plans = db.execute_fetch_all(
                "SELECT * FROM vip_plans WHERE id != %s",
                (current_sub['plan_id'],)
            )
            
            for plan in other_plans:
                keyboard.append([InlineKeyboardButton(
                    f"{plan['name']} - R${plan['price']:.2f}",
                    callback_data=f"plan_{plan['id']}"
                )])
            
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            status_message = f"✨ Você já é VIP!\n\nPlano atual: {current_sub['plan_name']}\n"
            if current_sub['is_permanent']:
                status_message += "Duração: Permanente\n\n"
            else:
                status_message += f"Dias restantes: {days_left}\n\n"
            
            await update.message.reply_text(status_message, reply_markup=reply_markup)
            return
            
        # Se não tiver assinatura, mostrar todos os planos
        plans = db.execute_fetch_all("SELECT * FROM vip_plans")
        
        keyboard = []
        for plan in plans:
            keyboard.append([InlineKeyboardButton(
                f"{plan['name']} - R${plan['price']:.2f}",
                callback_data=f"plan_{plan['id']}"
            )])

        messages = load_messages_from_db()
        start_message = messages.get('start_message', 'Escolha um dos planos VIP disponíveis:')
    
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            start_message,  # Using the message from database
            reply_markup=reply_markup
        )
        
    except Exception as e:
        logger.error(f"Erro no comando /start: {e}")
        await update.message.reply_text("Ocorreu um erro. Tente novamente mais tarde.")
    finally:
        db.close()

async def vip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando para acessar links VIP"""
    config = load_config()
    
    if not config:
        await update.message.reply_text("Erro ao carregar configurações. Tente novamente mais tarde.")
        return
    
    if config.get('admin_settings', {}).get('maintenance_mode', False):
        await update.message.reply_text("🛠️ O bot está em manutenção. Tente novamente mais tarde.")
        return
    
    user_id = update.effective_user.id
    
    # Verificar se o usuário tem assinatura ativa e gerar links
    links_message, error_message = await get_user_vip_links(context.bot, user_id)
    
    if links_message:
        await update.message.reply_text(
            text=links_message,
            parse_mode='Markdown'
        )
        logger.info(f"✅ Links VIP enviados via comando /vip para usuário {user_id}")
    else:
        # Se não tem assinatura ativa, mostrar opções
        keyboard = [
            [InlineKeyboardButton("💎 Ver Planos VIP", callback_data="show_plans")],
            [InlineKeyboardButton("📞 Suporte", url=f"https://t.me/{config.get('admin_user', 'admin')}")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await update.message.reply_text(
            f"❌ {error_message}\n\n"
            f"Para adquirir acesso VIP, clique no botão abaixo:",
            reply_markup=reply_markup
        )

async def handle_show_plans(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler para mostrar planos quando usuário clica em 'Ver Planos VIP'"""
    query = update.callback_query
    await query.answer()
    
    # Verificar modo manutenção
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    if config.get('admin_settings', {}).get('maintenance_mode', False):
        await query.message.reply_text("🛠️ O bot está em manutenção. Tente novamente mais tarde.")
        return
    
    # Mostrar todos os planos disponíveis
    keyboard = []
    db = Database()
    try:
        db.connect()
        plans = db.execute_fetch_all("SELECT * FROM vip_plans")
        for plan in plans:
            keyboard.append([InlineKeyboardButton(
                f"{plan['name']} - R${plan['price']:.2f}",
                callback_data=f"plan_{plan['id']}"
            )])
    finally:
        db.close()
    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.edit_text(
        "Escolha um dos planos VIP disponíveis:",
        reply_markup=reply_markup
    )

async def handle_plan_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler para seleção de planos VIP"""
    query = update.callback_query
    await query.answer()
    
    # Verificar modo manutenção
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    if config.get('admin_settings', {}).get('maintenance_mode', False):
        await query.message.reply_text("🛠️ O bot está em manutenção. Tente novamente mais tarde.")
        return
    
    # Extrair ID do plano do callback
    plan_id = int(query.data.split('_')[1])
    
    # Carregar mensagens do banco de dados
    messages = load_messages_from_db()
    
    # Verificar se messages foi carregado corretamente
    if not messages:
        logger.error("Falha ao carregar mensagens do banco de dados na função handle_plan_selection")
        messages = {}  # Usar dicionário vazio como fallback
    
    plan = await get_plan_by_id(plan_id)
    if not plan:
        await query.message.reply_text("Plano não encontrado.")
        return
    
    # Criar teclado com métodos de pagamento
    keyboard = []
    if config['payment_methods']['pix_automatico']['enabled']:
        keyboard.append([InlineKeyboardButton("💳 PIX Automático", callback_data=f"pix_auto_{plan_id}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    # Criar mensagem baseada no tipo (renovação ou novo plano)
    if query.data.startswith("renew_"):
        message = f"🔄 Renovação do Plano: {plan['name']}\n"
        message += f"💰 Valor: R${plan['price']:.2f}\n"
        message += f"⏱️ Duração: {'Permanente' if plan['duration_days'] == -1 else str(plan['duration_days']) + ' dias'}\n\n"
        message += f"{messages.get('payment_instructions', 'Para renovar, escolha o método de pagamento:')}"
    else:
        message = f"💎 Plano: {plan['name']}\n"
        message += f"💰 Valor: R${plan['price']:.2f}\n"
        message += f"⏱️ Duração: {'Permanente' if plan['duration_days'] == -1 else str(plan['duration_days']) + ' dias'}\n\n"
        message += f"{messages.get('payment_instructions', 'Para pagar, escolha o método de pagamento:')}"
    
    await query.message.edit_text(message, reply_markup=reply_markup)

async def handle_renewal_confirmation(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel_renew":
        keyboard = []
        db = Database()
        try:
            db.connect()
            plans = db.execute_fetch_all("SELECT * FROM vip_plans")
            for plan in plans:
                keyboard.append([InlineKeyboardButton(
                    f"{plan['name']} - R${plan['price']:.2f}",
                    callback_data=f"plan_{plan['id']}"
                )])
        finally:
            db.close()
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "Escolha um dos planos VIP disponíveis:",
            reply_markup=reply_markup
        )
        return
    plan_id = int(query.data.split('_')[2])
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    plan = await get_plan_by_id(plan_id)
    if not plan:
        await query.message.reply_text("Plano não encontrado.")
        return
    keyboard = []
    if config['payment_methods']['pix_automatico']['enabled']:
        keyboard.append([InlineKeyboardButton("💳 PIX Automático", callback_data=f"pix_auto_{plan_id}")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    messages = load_messages_from_db()
    message = f"🔄 Renovação Confirmada!\n\n"
    message += f"Plano: {plan['name']}\n"
    message += f"Valor: R${plan['price']:.2f}\n"
    message += f"Duração: {plan['duration_days']} dias\n\n"
    message += f"{messages.get('payment_instructions', 'Escolha o método de pagamento:')}"
    await query.message.edit_text(message, reply_markup=reply_markup)

async def handle_payment_method(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    # LOG DETALHADO DE ENTRADA
    logger.info(f"[DEBUG] handle_payment_method chamado: user_id={update.effective_user.id}, data={query.data}, chat_id={query.message.chat_id}, message_id={query.message.message_id}")
    await query.answer()
    
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    
    # Verifica modo manutenção
    if config.get('admin_settings', {}).get('maintenance_mode', False):
        await query.message.reply_text("🛠️ O bot está em manutenção. Tente novamente mais tarde.")
        return
    
    # Corrigindo o split do callback_data
    parts = query.data.split('_')
    method = parts[1]  # pix_auto
    plan_id = parts[2]  # ID do plano
    
    # Carregar mensagens do banco de dados
    messages = load_messages_from_db()
    
    plan = await get_plan_by_id(int(plan_id))
    if not plan:
        await query.message.reply_text("Plano não encontrado.")
        return
    
    if method == "auto":
        # Verificar se há provedores PIX automático disponíveis
        provider_manager = get_pix_provider_manager()
        available_providers = provider_manager.get_available_providers()
        
        if not available_providers:
            await query.message.reply_text(
                "❌ Nenhum provedor PIX automático configurado.\n\n"
                "Entre em contato com o administrador para configurar os provedores PIX."
            )
            return
        
        # Gerar PIX usando o sistema de provedores
        pix_data = await generate_pix_automatico(
            plan['price'],
            f"VIP {plan['name']} - {plan['duration_days']} dias",
            f"{update.effective_user.id}_{plan_id}"  # Referência externa
        )
        
        if pix_data:
            # Converter QR Code base64 para imagem
            import base64
            qr_code_bytes = base64.b64decode(pix_data['qr_code_base64'])
            qr_code = io.BytesIO(qr_code_bytes)
            
            # Criar botões "Já Paguei" e "Copiar Código PIX"
            keyboard = [
                [InlineKeyboardButton("✅ Já Paguei", callback_data=f"check_{pix_data['payment_id']}")],
                [InlineKeyboardButton("📋 Copiar Código PIX", callback_data=f"copy_pix_{pix_data['payment_id']}")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            # Criar mensagem com QR Code e código PIX copiável
            pix_code = pix_data['qr_code']
            provider_name = "MercadoPago" if pix_data['provider'] == 'mercadopago' else "CNPay"
            
            caption = (
                f"{messages.get('pix_automatico_instructions', 'Escaneie o QR Code abaixo para pagar automaticamente:')}\n\n"
                f"💰 Valor: R${plan['price']:.2f}\n\n"
                f"📋 **Código PIX para copiar:**\n"
                f"`{pix_code}`\n\n"
                f"📱 **Como pagar:**\n"
                f"1. Escaneie o QR Code acima, OU\n"
                f"2. Copie o código PIX acima e cole no app do seu banco\n\n"
                f"⏳ Aguardando pagamento..."
            )
            
            # Enviar mensagem com QR Code
            message = await query.message.reply_photo(
                photo=qr_code,
                caption=caption,
                reply_markup=reply_markup,
                parse_mode='Markdown'
            )
            
            # Iniciar verificação automática (se job_queue estiver disponível)
            if hasattr(context, 'job_queue') and context.job_queue is not None:
                try:
                    context.job_queue.run_repeating(
                        check_payment_auto,
                        interval=5,
                        first=5,
                        data={
                            'message_id': message.message_id,
                            'chat_id': message.chat_id,
                            'payment_id': pix_data['payment_id'],
                            'user_id': update.effective_user.id,
                            'plan_id': plan_id,
                            'plan': plan,
                            'provider': pix_data['provider']  # Adicionar provedor usado
                        }
                    )
                    logger.info("✅ Verificação automática de pagamento iniciada")
                except Exception as e:
                    logger.warning(f"⚠️ Erro ao iniciar verificação automática: {e}")
                    logger.info("ℹ️ Pagamento será verificado apenas via webhook")
            else:
                logger.warning("⚠️ JobQueue não disponível - pagamento será verificado apenas via webhook")
                logger.info("ℹ️ Para verificação automática, instale: pip install 'python-telegram-bot[job-queue]'")
        else:
            await query.message.reply_text(
                "❌ Erro ao gerar PIX automático.\n\n"
                "Todos os provedores PIX estão indisponíveis no momento.\n"
                "Tente novamente mais tarde."
            )
    else:
        chave_pix = config['payment_methods']['pix_manual']['chave_pix']
        nome_titular = config['payment_methods']['pix_manual']['nome_titular']
        admin_user = config['admin_user']
        
        # Criar mensagem com instruções do PIX
        message = (
            f"💳 *Pagamento via PIX Manual*\n\n"
            f"📝 *Instruções:*\n"
            f"1. Faça o PIX para a chave: `{chave_pix}`\n"
            f"2. Nome do titular: {nome_titular}\n"
            f"3. Após o pagamento, clique no botão abaixo para enviar o comprovante\n\n"
            f"⚠️ *Importante:*\n"
            f"• Envie o comprovante apenas após realizar o pagamento\n"
            f"• Aguarde a confirmação do admin\n"
            f"• O processo pode levar alguns minutos"
        )
        
        # Criar botão para contato com admin
        keyboard = [
            [InlineKeyboardButton("📤 Enviar Comprovante", url=f"https://t.me/{admin_user}")],
            [InlineKeyboardButton("🔙 Voltar", callback_data="back_to_plans")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.message.reply_text(
            text=message,
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )

async def check_payment_auto(context: ContextTypes.DEFAULT_TYPE):
    """Verifica pagamento automaticamente e atualiza status"""
    job = context.job
    data = job.data
    
    try:
        payment_id = data['payment_id']
        user_id = data['user_id']
        plan_id = data['plan_id']
        plan = data['plan']
        provider = data.get('provider', 'mercadopago')
        
        logger.info(f"🔍 Verificando pagamento {payment_id} (provedor: {provider})")
        
        # Para CNPay, verificar se o pagamento já foi processado via webhook
        if provider == 'cnpay':
            # Verificar se o pagamento foi aprovado no banco de dados
            db = Database()
            try:
                db.connect()
                if db.connection:
                    payment_status = db.execute_fetch_one(
                        "SELECT status FROM payments WHERE payment_id = %s",
                        (payment_id,)
                    )
                    
                    if payment_status and payment_status['status'] == 'approved':
                        logger.info(f"✅ Pagamento CNPay {payment_id} já foi processado via webhook!")
                        
                        # Parar verificação automática
                        job.schedule_removal()
                        
                        # Atualizar mensagem
                        try:
                            await context.bot.edit_message_text(
                                chat_id=data['chat_id'],
                                message_id=data['message_id'],
                                text=f"✅ **Pagamento Aprovado!**\n\n"
                                     f"💎 Plano: {plan['name']}\n"
                                     f"💰 Valor: R${plan['price']:.2f}\n"
                                     f"⏱️ Duração: {plan['duration_days']} dias\n\n"
                                     f"🎉 Sua assinatura VIP foi ativada com sucesso!\n\n"
                                     f"Use /start para ver seus planos ativos.",
                                parse_mode='Markdown'
                            )
                        except Exception as e:
                            logger.error(f"Erro ao atualizar mensagem: {e}")
                        
                        return
                    elif payment_status and payment_status['status'] in ['rejected', 'cancelled']:
                        logger.info(f"❌ Pagamento CNPay {payment_id} foi rejeitado/cancelado")
                        
                        # Parar verificação automática
                        job.schedule_removal()
                        
                        # Atualizar mensagem
                        try:
                            await context.bot.edit_message_text(
                                chat_id=data['chat_id'],
                                message_id=data['message_id'],
                                text=f"❌ **Pagamento Rejeitado**\n\n"
                                     f"O pagamento foi rejeitado ou cancelado.\n\n"
                                     f"Tente novamente ou entre em contato com o suporte.",
                                parse_mode='Markdown'
                            )
                        except Exception as e:
                            logger.error(f"Erro ao atualizar mensagem: {e}")
                        
                        return
                    else:
                        # Pagamento ainda pendente, continuar verificando
                        logger.info(f"⏳ Pagamento CNPay {payment_id} ainda pendente, continuando verificação...")
                        return
            except Exception as e:
                logger.error(f"Erro ao verificar status do pagamento CNPay: {e}")
            finally:
                db.close()
            
            # Se chegou aqui, pagamento ainda não foi processado
            logger.info(f"🔔 CNPay detectado - aguardando processamento via webhook para {payment_id}")
            return
        
        # Para outros provedores, verificar normalmente
        payment_info = await check_payment_async(payment_id, provider)
        
        if payment_info and payment_info.get('status') == 'approved':
            logger.info(f"✅ Pagamento {payment_id} aprovado!")
            
            # Parar verificação automática
            job.schedule_removal()
            
            # Registrar assinatura
            success = await register_vip_subscription(user_id, plan_id, payment_id, context)
            
            if success:
                # Atualizar mensagem
                try:
                    await context.bot.edit_message_text(
                        chat_id=data['chat_id'],
                        message_id=data['message_id'],
                        text=f"✅ **Pagamento Aprovado!**\n\n"
                             f"💎 Plano: {plan['name']}\n"
                             f"💰 Valor: R${plan['price']:.2f}\n"
                             f"⏱️ Duração: {plan['duration_days']} dias\n\n"
                             f"🎉 Sua assinatura VIP foi ativada com sucesso!\n\n"
                             f"Use /start para ver seus planos ativos.",
                        parse_mode='Markdown'
                    )
                except Exception as e:
                    logger.error(f"Erro ao atualizar mensagem: {e}")
                    
                # Adicionar usuário aos grupos VIP
                await add_user_to_vip_groups(context.bot, user_id, plan_id)
                
            else:
                logger.error(f"Erro ao registrar assinatura para usuário {user_id}")
                
        elif payment_info and payment_info.get('status') in ['rejected', 'cancelled']:
            logger.info(f"❌ Pagamento {payment_id} rejeitado/cancelado")
            
            # Parar verificação automática
            job.schedule_removal()
            
            # Atualizar mensagem
            try:
                await context.bot.edit_message_text(
                    chat_id=data['chat_id'],
                    message_id=data['message_id'],
                    text=f"❌ **Pagamento Rejeitado**\n\n"
                         f"O pagamento foi rejeitado ou cancelado.\n\n"
                         f"Tente novamente ou entre em contato com o suporte.",
                    parse_mode='Markdown'
                )
            except Exception as e:
                logger.error(f"Erro ao atualizar mensagem: {e}")
                
    except Exception as e:
        logger.error(f"Erro na verificação automática: {e}")

async def check_payment_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    # Verificar se é callback de copiar PIX
    if query.data.startswith('copy_pix_'):
        payment_id = query.data.split('_')[2]
        
        # Buscar dados do pagamento no banco
        db = Database()
        try:
            db.connect()
            if not db.connection:
                await query.answer("❌ Erro de conexão com banco de dados", show_alert=True)
                return
                
            payment_data = db.execute_fetch_one(
                "SELECT qr_code_data FROM payments WHERE payment_id = %s",
                (payment_id,)
            )
            
            if payment_data and payment_data.get('qr_code_data'):
                pix_code = payment_data['qr_code_data']
                
                # Enviar código PIX formatado para cópia fácil
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=f"📋 **Código PIX para copiar:**\n\n"
                         f"```\n{pix_code}\n```\n\n"
                         f"💡 **Instruções:**\n"
                         f"1️⃣ Toque no código acima (entre as linhas)\n"
                         f"2️⃣ Todo o código será selecionado automaticamente\n"
                         f"3️⃣ Pressione Ctrl+C para copiar\n"
                         f"4️⃣ Cole no app do seu banco\n\n"
                         f"✅ **Dica:** O código está formatado para seleção fácil!",
                    parse_mode='Markdown'
                )
                
                await query.answer("📋 Código PIX enviado - toque no código para copiar!", show_alert=True)
            else:
                await query.answer("❌ Código PIX não encontrado", show_alert=True)
                
        except Exception as e:
            logger.error(f"Erro ao buscar código PIX: {e}")
            await query.answer("❌ Erro ao buscar código PIX", show_alert=True)
        finally:
            db.close()
        return
    
    payment_id = query.data.split('_')[1]
    user_id = update.effective_user.id
    
    # Primeiro, verificar se o usuário já tem uma assinatura ativa
    links_message, error_message = await get_user_vip_links(context.bot, user_id)
    
    if links_message:
        # Usuário tem assinatura ativa - mostrar links
        try:
            await query.message.edit_text(
                text=links_message,
                parse_mode='Markdown'
            )
            logger.info(f"✅ Links VIP enviados para usuário {user_id}")
            return
        except Exception as e:
            logger.error(f"Erro ao enviar links VIP: {e}")
            # Se falhar ao editar, enviar nova mensagem
            await context.bot.send_message(
                chat_id=user_id,
                text=links_message,
                parse_mode='Markdown'
            )
            return
    
    # Se não tem assinatura ativa, verificar pagamento
    payment = check_payment(payment_id)
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    
    # Carregar mensagens do banco de dados
    messages = load_messages_from_db()
    
    if payment and payment.get('status') == 'approved':
        # Extrair informações do pagamento
        external_reference = payment.get('external_reference', '')
        if external_reference:
            user_id, plan_id = external_reference.split('_')
            
            # Encontrar o plano
            plan = await get_plan_by_id(int(plan_id))
            if not plan:
                await query.message.reply_text("Plano não encontrado.")
                return
            
            # Registrar assinatura
            if await register_vip_subscription(int(user_id), int(plan_id), payment_id, context):
                # Adicionar usuário aos grupos VIP
                await add_user_to_vip_groups(context.bot, int(user_id), int(plan_id))
                
                try:
                    # Atualizar mensagem com confirmação
                    success_message = f"✅ {messages.get('payment_success', 'Pagamento aprovado!').format(dias=plan['duration_days'])}"
                    await query.message.edit_caption(caption=success_message)
                    
                    # Remover botão
                    await query.message.edit_reply_markup(reply_markup=None)
                except Exception as e:
                    logger.error(f"Erro ao atualizar mensagem: {e}")
                    # Se falhar, tenta enviar uma nova mensagem
                    await query.message.reply_text(success_message)
                
                # Parar verificação automática se existir
                if hasattr(context, 'job_queue') and context.job_queue:
                    for job in context.job_queue.jobs():
                        if job.data.get('payment_id') == payment_id:
                            job.schedule_removal()
    else:
        # Se não tem assinatura ativa e pagamento não foi aprovado
        if error_message:
            await query.answer(error_message, show_alert=True)
        else:
            status = messages.get('payment_pending', 'Aguardando confirmação do pagamento...')
            if payment:
                if payment.get('status') == 'rejected':
                    status = messages.get('payment_error', 'Ocorreu um erro no pagamento. Tente novamente.')
            
            await query.answer(status, show_alert=True)

# Comandos do admin
async def admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    print(f"[DEBUG] update.effective_user.id = {update.effective_user.id} (type={type(update.effective_user.id)})")
    if not is_admin(int(update.effective_user.id)):
        logger.info(f"Usuário {update.effective_user.id} tentou acessar sem permissão.")
        await update.message.reply_text("Acesso negado.")
        return
    
    config = load_config()
    if config is None:
        await update.message.reply_text("Erro ao carregar configurações.")
        return
    
    # Carregar mensagens do banco de dados
    messages = load_messages_from_db()
    
    keyboard = [
        [InlineKeyboardButton("📊 Estatísticas", callback_data="admin_stats")],
        [
            InlineKeyboardButton("⚙️ Configurações", callback_data="admin_settings"),
            InlineKeyboardButton("👥 Usuários VIP", callback_data="admin_vip_users")
        ],
        [InlineKeyboardButton("💎 Planos VIP", callback_data="admin_vip_plans")],
        [InlineKeyboardButton("📝 Mensagens", callback_data="admin_messages")],
        [InlineKeyboardButton("⏰ Agendar Mensagens", callback_data="admin_schedule_messages")],
        [InlineKeyboardButton("🔄 Manutenção", callback_data="admin_maintenance")],
        [InlineKeyboardButton("👤 Gerenciar Admins", callback_data="admin_manage_admins")],
        [InlineKeyboardButton("⚒️ Suporte", url=config.get('support_admin', 'https://t.me/suporte'))]  # Botão de suporte
    ]
    
    # Adicionar botão de broadcast (com emoji de cadeado para admins não-VIP)
    if is_admin_vip(update.effective_user.id):
        keyboard.insert(6, [InlineKeyboardButton("📢 Broadcast", callback_data="admin_broadcast")])
    else:
        keyboard.insert(6, [InlineKeyboardButton("📢🔒 Broadcast (VIP)", callback_data="admin_broadcast_locked")])

    
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        messages.get('admin_welcome', 'Bem-vindo ao painel administrativo.'),
        reply_markup=reply_markup
    )

def generate_database_structure():
    """Gera backup completo do banco de dados (estrutura + dados)"""
    logger.info("[DATABASE] Iniciando geração da estrutura do banco")
    db = Database()
    try:
        conn = db.connect()
        if not conn:
            logger.error("[DATABASE] Falha ao conectar ao banco de dados")
            return "❌ Erro ao conectar ao banco de dados"
        
        cursor = conn.cursor()
        
        # Lista de todas as tabelas
        cursor.execute("SHOW TABLES")
        tables = cursor.fetchall()
        
        sql_backup = "-- Backup Simples do Banco de Dados\n"
        sql_backup += "-- Gerado automaticamente em " + str(datetime.now()) + "\n"
        sql_backup += "-- Versão simplificada sem dependências complexas\n\n"
        sql_backup += "-- Configurações básicas\n"
        sql_backup += "SET SQL_MODE = \"NO_AUTO_VALUE_ON_ZERO\";\n"
        sql_backup += "SET FOREIGN_KEY_CHECKS = 0;\n\n"
        
        for table in tables:
            table_name = table[0]
            
            # Obter estrutura da tabela de forma simplificada
            try:
                sql_backup += f"-- Tabela: {table_name}\n"
                sql_backup += f"DROP TABLE IF EXISTS `{table_name}`;\n"
                
                # Obter estrutura básica da tabela
                cursor.execute(f"DESCRIBE `{table_name}`")
                columns = cursor.fetchall()
                
                if columns:
                    # Construir CREATE TABLE básico
                    sql_backup += f"CREATE TABLE `{table_name}` (\n"
                    
                    column_definitions = []
                    primary_keys = []
                    
                    for col in columns:
                        col_name = col[0]
                        col_type = col[1]
                        col_null = "NOT NULL" if col[2] == "NO" else "NULL"
                        col_key = col[3]
                        col_default = col[4]
                        col_extra = col[5]
                        
                        # Construir definição da coluna
                        col_def = f"  `{col_name}` {col_type}"
                        
                        if col_null == "NOT NULL":
                            col_def += " NOT NULL"
                        
                        if col_default and col_default != "NULL":
                            if col_default.upper() in ["CURRENT_TIMESTAMP", "NOW()"]:
                                col_def += f" DEFAULT {col_default}"
                            elif "DEFAULT_GENERATED" in col_default:
                                # Remover DEFAULT_GENERATED que causa erro de sintaxe
                                col_def += " DEFAULT CURRENT_TIMESTAMP"
                            else:
                                col_def += f" DEFAULT '{col_default}'"
                        
                        if col_extra:
                            # Limpar extras problemáticos
                            if "DEFAULT_GENERATED" in col_extra:
                                col_extra = col_extra.replace("DEFAULT_GENERATED", "").strip()
                            if col_extra and col_extra != "":
                                col_def += f" {col_extra}"
                        
                        column_definitions.append(col_def)
                        
                        if col_key == "PRI":
                            primary_keys.append(f"`{col_name}`")
                    
                    # Limpar definições problemáticas
                    clean_definitions = []
                    for col_def in column_definitions:
                        # Remover DEFAULT_GENERATED de qualquer lugar
                        clean_def = col_def.replace("DEFAULT_GENERATED", "").strip()
                        # Remover espaços duplos
                        clean_def = " ".join(clean_def.split())
                        clean_definitions.append(clean_def)
                    
                    sql_backup += ",\n".join(clean_definitions)
                    
                    if primary_keys:
                        sql_backup += f",\n  PRIMARY KEY ({', '.join(primary_keys)})"
                    
                    sql_backup += "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;\n\n"
                else:
                    sql_backup += f"-- Erro: Não foi possível obter estrutura da tabela {table_name}\n\n"
                    
            except Exception as e:
                logger.warning(f"[DATABASE] Erro ao obter estrutura da tabela {table_name}: {e}")
                sql_backup += f"-- Erro ao obter estrutura da tabela `{table_name}`: {e}\n\n"
                continue
            
            # Obter dados da tabela de forma simplificada
            try:
                cursor.execute(f"SELECT * FROM `{table_name}`")
                rows = cursor.fetchall()
                
                if rows:
                    # Obter nomes das colunas
                    cursor.execute(f"DESCRIBE `{table_name}`")
                    columns = cursor.fetchall()
                    column_names = [col[0] for col in columns]
                    
                    sql_backup += f"-- Dados da tabela {table_name}\n"
                    
                    for row in rows:
                        # Construir INSERT statement simples
                        values = []
                        for value in row:
                            if value is None:
                                values.append("NULL")
                            elif isinstance(value, str):
                                # Escapar aspas simples
                                escaped_value = value.replace("'", "\\'")
                                values.append(f"'{escaped_value}'")
                            elif isinstance(value, (int, float)):
                                values.append(str(value))
                            elif isinstance(value, datetime):
                                values.append(f"'{value.strftime('%Y-%m-%d %H:%M:%S')}'")
                            else:
                                values.append(f"'{str(value)}'")
                        
                        insert_sql = f"INSERT INTO `{table_name}` (`{'`, `'.join(column_names)}`) VALUES ({', '.join(values)});\n"
                        sql_backup += insert_sql
                    
                    sql_backup += "\n"
                else:
                    sql_backup += f"-- Tabela {table_name} está vazia\n\n"
            except Exception as e:
                logger.warning(f"[DATABASE] Erro ao obter dados da tabela {table_name}: {e}")
                sql_backup += f"-- Erro ao obter dados da tabela {table_name}: {e}\n\n"
        
        sql_backup += "-- Restaurar configurações\n"
        sql_backup += "SET FOREIGN_KEY_CHECKS = 1;\n"
        sql_backup += "\n-- Backup concluído com sucesso!\n"
        
        cursor.close()
        return sql_backup
        
    except Exception as e:
        return f"❌ Erro ao gerar backup: {str(e)}"
    finally:
        db.close()

async def test_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando de teste simples"""
    logger.info("🚀 [TEST] Função test_command chamada!")
    await update.message.reply_text("✅ Comando de teste funcionando!")
    logger.info(f"[TEST] Comando /test recebido do usuário {update.effective_user.id}")

async def database_simple(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando de teste para database"""
    await update.message.reply_text("✅ Comando /database_simple funcionando!")
    logger.info(f"[DATABASE_SIMPLE] Comando recebido do usuário {update.effective_user.id}")

async def database(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando para gerar e enviar a estrutura SQL do banco de dados"""
    logger.info("🚀 [DATABASE] Função database chamada!")
    user_id = int(update.effective_user.id)
    logger.info(f"[DATABASE] Comando /database recebido do usuário {user_id}")
    
    # Verificar se é admin
    if not is_admin(user_id):
        logger.info(f"[DATABASE] Usuário {user_id} tentou acessar comando /database sem permissão.")
        await update.message.reply_text("❌ Acesso negado. Apenas administradores podem usar este comando.")
        return
    
    logger.info(f"[DATABASE] Usuário {user_id} autorizado para comando /database")
    
    try:
        # Mostrar mensagem de processamento
        processing_msg = await update.message.reply_text("🔄 Gerando backup completo do banco de dados...")
        
        # Gerar backup SQL completo
        logger.info(f"[DATABASE] Iniciando geração do backup para usuário {user_id}")
        sql_backup = generate_database_structure()
        
        if sql_backup.startswith("❌"):
            logger.error(f"[DATABASE] Erro ao gerar backup: {sql_backup}")
            await processing_msg.edit_text(sql_backup)
            return
        
        logger.info(f"[DATABASE] Backup gerado com sucesso para usuário {user_id}")
        
        # Criar arquivo temporário
        from io import BytesIO
        sql_file = BytesIO(sql_backup.encode('utf-8'))
        sql_file.name = f"database_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.sql"
        
        # Enviar arquivo
        await processing_msg.delete()
        await update.message.reply_document(
            document=sql_file,
            caption="💾 Backup Completo do Banco de Dados\n\n"
                   f"📅 Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n"
                   "📊 Inclui estrutura e todos os dados\n"
                   "🔄 Use este arquivo para restauração completa do banco",
            filename=sql_file.name
        )
        
        logger.info(f"✅ Backup do banco enviado para admin {user_id}")
        
    except Exception as e:
        logger.error(f"❌ Erro no comando /database: {e}")
        await update.message.reply_text(f"❌ Erro ao gerar backup do banco: {str(e)}")

async def handle_admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("[DEBUG] Entrou em handle_admin_callback")
    query = update.callback_query
    await query.answer()
    
    logger.info(f"Callback recebido: {query.data}")
    logger.info(f"[DEBUG] update.effective_user.id = {update.effective_user.id} (type={type(update.effective_user.id)})")
    
    if not is_admin(int(update.effective_user.id)):
        logger.info(f"Usuário {update.effective_user.id} tentou acessar sem permissão.")
        await query.message.reply_text("Acesso negado.")
        return
    
    config = load_config()
    if config is None:
        await query.answer("❌ Erro ao carregar configurações", show_alert=True)
        return
    
    # Carregar mensagens do banco de dados
    messages = load_messages_from_db()
    
    # TRATAMENTO ESPECÍFICO PARA GERENCIAR GRUPOS DO PLANO
    if query.data.startswith("admin_manage_plan_groups_"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.answer("❌ Plano não encontrado!")
            return
        
        # Buscar grupos associados ao plano
        db = Database()
        try:
            db.connect()
            plan_groups = db.execute_fetch_all(
                """SELECT vg.id, vg.group_name, vg.group_id 
                FROM vip_groups vg 
                JOIN plan_groups pg ON vg.id = pg.group_id 
                WHERE pg.plan_id = %s AND vg.is_active = TRUE""",
                (plan_id,)
            )
            
            # Buscar todos os grupos disponíveis
            all_groups = db.execute_fetch_all(
                "SELECT id, group_name, group_id FROM vip_groups WHERE is_active = TRUE"
            )
        finally:
            db.close()
        
        # Criar lista de grupos associados
        associated_group_ids = [group['id'] for group in plan_groups]
        
        keyboard = []
        for group in all_groups:
            is_associated = group['id'] in associated_group_ids
            status_icon = "✅" if is_associated else "❌"
            keyboard.append([
                InlineKeyboardButton(
                    f"{status_icon} {group['group_name']}", 
                    callback_data=f"admin_toggle_plan_group_{plan_id}_{group['id']}"
                )
            ])
        
        keyboard.append([InlineKeyboardButton("➕ Adicionar Novo Grupo/Canal", callback_data=f"admin_add_new_group_{plan_id}")])
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data=f"admin_edit_plan_{plan_id}")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        groups_text = ""
        if plan_groups:
            groups_text = "\n📱 Grupos associados:\n"
            for group in plan_groups:
                groups_text += f"• {group['group_name']}\n"
        else:
            groups_text = "\n📱 Nenhum grupo associado"
        
        await query.message.edit_text(
            f"📱 Gerenciar Grupos do Plano: {plan['name']}\n\n"
            f"Clique nos grupos para associar/desassociar:"
            f"{groups_text}\n\n"
            f"✅ = Associado | ❌ = Não associado",
            reply_markup=reply_markup
        )
        return
    
    # TRATAMENTO ESPECÍFICO PARA ADICIONAR NOVO GRUPO AO PLANO
    if query.data.startswith("admin_add_new_group_"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.answer("❌ Plano não encontrado!")
            return
        
        # Configurar estado para adicionar novo grupo
        context.user_data['adding_group'] = {
            'plan_id': plan_id,
            'plan_name': plan['name'],
            'step': 'group_name'
        }
        
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data=f"admin_manage_plan_groups_{plan_id}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.message.edit_text(
            f"➕ Adicionar Novo Grupo/Canal ao Plano: {plan['name']}\n\n"
            f"Digite o nome do novo grupo/canal:",
            reply_markup=reply_markup
        )
        return

    if query.data == "admin_toggle_maintenance":
        await handle_maintenance_toggle(update, context)
        return
    # Handler for "Gerenciar Admins"
    if query.data == "admin_manage_admins":
        db = Database()
        try:
            db.connect()
            admins = db.execute_fetch_all("SELECT admin_id FROM admins")
        finally:
            db.close()
        admin_ids = [a['admin_id'] for a in admins] if admins else []
        keyboard = []
        for admin_id in admin_ids:
            label = f"👤 {admin_id}"
            if str(admin_id) != str(update.effective_user.id):
                keyboard.append([
                    InlineKeyboardButton(
                        f"Remover {admin_id}",
                        callback_data=f"admin_remove_admin_{admin_id}"
                    )
                ])
            else:
                keyboard.append([
                    InlineKeyboardButton(
                        f"Você ({admin_id})",
                        callback_data="noop"
                    )
                ])
        keyboard.append([InlineKeyboardButton("➕ Adicionar Admin", callback_data="admin_add_admin")])
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "👤 Gerenciar administradores\n\nSelecione uma opção:",
            reply_markup=reply_markup
        )
        return

    # Handler for removing an admin
    if query.data.startswith("admin_remove_admin_"):
        admin_id_to_remove = query.data.replace("admin_remove_admin_", "")
        if str(admin_id_to_remove) == str(update.effective_user.id):
            await query.answer("Você não pode remover a si mesmo!", show_alert=True)
            return
        db = Database()
        try:
            db.connect()
            db.execute_query("DELETE FROM admins WHERE admin_id = %s", (admin_id_to_remove,), commit=True)
        finally:
            db.close()
            await query.answer("Admin removido com sucesso!")

            # Recarrega a lista de admins corretamente, sem recursão
            db = Database()
            try:
                db.connect()
                admins = db.execute_fetch_all("SELECT admin_id FROM admins")
            finally:
                db.close()

            admin_ids = [a['admin_id'] for a in admins] if admins else []
            keyboard = []
            for admin_id in admin_ids:
                label = f"👤 {admin_id}"
                if str(admin_id) != str(update.effective_user.id):
                    keyboard.append([
                        InlineKeyboardButton(
                            f"Remover {admin_id}",
                            callback_data=f"admin_remove_admin_{admin_id}"
                        )
                    ])
                else:
                    keyboard.append([
                        InlineKeyboardButton(
                            f"Você ({admin_id})",
                            callback_data="noop"
                        )
                    ])
            keyboard.append([InlineKeyboardButton("➕ Adicionar Admin", callback_data="admin_add_admin")])
            keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin")])
            reply_markup = InlineKeyboardMarkup(keyboard)

            await query.message.edit_text(
                "👤 Gerenciar administradores\n\nSelecione uma opção:",
                reply_markup=reply_markup
            )

        return

    # Handler to start the add admin process
    if query.data == "admin_add_admin":
        context.user_data['waiting_for_admin_id'] = True
        await query.message.edit_text(
            "Envie o ID do novo admin:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_manage_admins")]
            ])
        )
        return
        
    # TRATAMENTO ESPECÍFICO PARA ALTERNAR ASSOCIAÇÃO DE GRUPO AO PLANO
    if query.data.startswith("admin_toggle_plan_group_"):
        parts = query.data.split('_')
        plan_id = int(parts[-2])
        group_id = int(parts[-1])
        
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.answer("❌ Plano não encontrado!")
            return
        
        db = Database()
        try:
            db.connect()
            
            # Verificar se o grupo está associado
            existing = db.execute_fetch_one(
                "SELECT * FROM plan_groups WHERE plan_id = %s AND group_id = %s",
                (plan_id, group_id)
            )
            
            if existing:
                # Desassociar grupo
                db.execute_query(
                    "DELETE FROM plan_groups WHERE plan_id = %s AND group_id = %s",
                    (plan_id, group_id),
                    commit=True
                )
                await query.answer("❌ Grupo desassociado!")
            else:
                # Associar grupo
                db.execute_query(
                    "INSERT INTO plan_groups (plan_id, group_id) VALUES (%s, %s)",
                    (plan_id, group_id),
                    commit=True
                )
                await query.answer("✅ Grupo associado!")
                
        finally:
            db.close()
        
        # Recarregar o menu de gerenciamento de grupos
        await handle_admin_callback(update, context)
        return
    
    # TRATAMENTO ESPECÍFICO PARA CONFIRMAÇÃO DE REMOÇÃO DE PLANO (DEVE VIR PRIMEIRO)
    if query.data.startswith("admin_confirm_remove_plan_"):
        logger.info(f"[DEBUG] Processando confirmação de remoção do plano {query.data}")
        plan_id = int(query.data.split('_')[-1])
        logger.info(f"[DEBUG] Plan ID extraído: {plan_id}")
        
        plan = await get_plan_by_id(plan_id)
        if not plan:
            logger.error(f"[DEBUG] Plano {plan_id} não encontrado!")
            await query.answer("❌ Plano não encontrado!")
            return
        
        logger.info(f"[DEBUG] Plano encontrado: {plan['name']}")

        config = load_config()
        if config is None:
            await query.answer("❌ Erro ao carregar configurações", show_alert=True)
            return
        
        db = Database()
        try:
            db.connect()
            logger.info(f"[DEBUG] Conectado ao banco de dados")
            # Deletar o plano permanentemente
            db.execute_query(
                "DELETE FROM vip_plans WHERE id = %s",
                (plan_id,),
                commit=True
            )
            logger.info(f"[DEBUG] Plano {plan_id} deletado permanentemente do banco")
        finally:
            db.close()
            logger.info(f"[DEBUG] Conexão com banco fechada")
        
        logger.info(f"[DEBUG] Buscando planos ativos para atualizar menu")
        # Voltar para o menu de planos (sem chamada recursiva)
        db = Database()
        try:
            db.connect()
            plans = db.execute_fetch_all("SELECT * FROM vip_plans")
        finally:
            db.close()
        
        keyboard = []
        for plan_item in plans:
            keyboard.append([
                InlineKeyboardButton(f"✏️ {plan_item['name']} (R${plan_item['price']:.2f})", callback_data=f"admin_edit_plan_{plan_item['id']}"),
                InlineKeyboardButton("🗑️", callback_data=f"admin_remove_plan_{plan_item['id']}")
            ])
        keyboard.append([InlineKeyboardButton("➕ Adicionar Novo Plano", callback_data="admin_add_plan")])
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        logger.info(f"[DEBUG] Editando mensagem com confirmação")
        await query.message.edit_text(
            f"✅ **Plano '{plan['name']}' removido com sucesso!**\n\n"
            f"💎 Gerenciar Planos VIP\n\n"
            f"Selecione um plano para editar ou remova/adicione novos planos:",
            reply_markup=reply_markup
        )
        logger.info(f"[DEBUG] Mensagem editada com sucesso")
        return
    
    # Verificar se é um callback de edição de configurações
    if query.data == "admin_edit_bot_token":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🔑 Editar Token do Bot\n\n"
            f"Token atual: {config['bot_token']}\n\n"
            "Envie o novo token do bot:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'bot_token'
        return
    elif query.data == "admin_edit_mp_token":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "💳 Editar Token do MercadoPago\n\n"
            f"Token atual: {config['mercadopago']['access_token']}\n\n"
            "Envie o novo token do MercadoPago:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'mp_token'
        return
    elif query.data == "admin_edit_pix_key":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📱 Editar Chave PIX\n\n"
            f"Chave atual: {config['payment_methods']['pix_manual']['chave_pix']}\n\n"
            "Envie a nova chave PIX:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'pix_key'
        return
    elif query.data == "admin_edit_pix_name":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "👤 Editar Nome do Titular PIX\n\n"
            f"Nome atual: {config['payment_methods']['pix_manual']['nome_titular']}\n\n"
            "Envie o novo nome do titular:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'pix_name'
        return
    
    
    elif query.data == "admin_upload_welcome_file":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_welcome_file")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📎 Enviar Novo Arquivo de Boas-vindas\n\n"
            "Envie uma foto ou vídeo que será usado como arquivo de boas-vindas.\n\n"
            "⚠️ O arquivo deve ser menor que 50MB.",
            reply_markup=reply_markup
        )
        context.user_data['waiting_for_welcome_file'] = True
        return
    
    elif query.data == "admin_remove_welcome_file":
        keyboard = [
            [InlineKeyboardButton("✅ Confirmar", callback_data="admin_confirm_remove_welcome_file")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="admin_welcome_file")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🗑️ Remover Arquivo de Boas-vindas\n\n"
            "Tem certeza que deseja remover o arquivo de boas-vindas?\n"
            "Esta ação não pode ser desfeita.",
            reply_markup=reply_markup
        )
        return
    

    # Verificar se é um callback de broadcast trancado
    if query.data == "admin_broadcast_locked":
        # Apagar a mensagem atual
        await query.message.delete()
        
        # Enviar nova mensagem com informações sobre liberação de recursos
        keyboard = [
            [InlineKeyboardButton("💎 Quero ser Premium", callback_data="admin_upgrade_vip")],
            [InlineKeyboardButton("⬅️ Voltar ao Menu", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        await query.message.reply_text(
            "🔒 **Função Exclusiva para Administradores VIP**\n\n"
            "Para liberar todos os recursos do bot e ter acesso completo a todas as funcionalidades, "
            "torne-se um administrador VIP!\n\n"
            "**Recursos VIP incluem:**\n"
            "• 📢 Broadcast para todos os usuários\n"
            "• 📹 Envio de vídeos em massa\n"
            "• ⭕ Vídeos circulares\n"
            "• 📊 Relatórios avançados\n"
            "• ⚙️ Configurações exclusivas\n"
            "• 🎯 Ferramentas de marketing\n\n"
            "**Valor:** R$ 50,00/mês\n\n"
            "💬 Entre em contato com o suporte para mais informações:",
            reply_markup=reply_markup,
            parse_mode='Markdown'
        )
        return
    
    # Verificar se é um callback de upgrade para VIP
    if query.data == "admin_upgrade_vip":
        # Responder imediatamente para evitar timeout
        await query.answer("🔄 Gerando QR Code de pagamento...", show_alert=True)
        
        # Gerar QR Code para pagamento do VIP do admin
        admin_id = update.effective_user.id
        amount = 50.00  # Valor do VIP mensal
        description = f"Upgrade VIP Admin - {admin_id}"
        external_reference = f"admin_vip_{admin_id}_{int(datetime.now().timestamp())}"
        
        # Gerar PIX
        pix_result = await generate_pix_automatico(amount, description, external_reference)
        
        # Debug: verificar o resultado
        logger.info(f"🔍 Resultado do PIX: {pix_result}")
        
        if pix_result and pix_result.get('qr_code'):
            # Salvar dados do pagamento
            db = Database()
            try:
                db.connect()
                if db.connection:
                    db.execute_query(
                        """INSERT INTO admin_vip_payments 
                        (admin_id, amount, description, external_reference, pix_code, created_at, status) 
                        VALUES (%s, %s, %s, %s, %s, NOW(), 'pending')""",
                        (admin_id, amount, description, external_reference, pix_result['pix_code']),
                        commit=True
                    )
            except Exception as e:
                logger.error(f"Erro ao salvar pagamento VIP admin: {e}")
            finally:
                db.close()
            
            # Enviar QR Code
            keyboard = [
                [InlineKeyboardButton("🔄 Verificar Pagamento", callback_data="admin_check_vip_payment")],
                [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            # Converter base64 para bytes
            import base64
            qr_image_bytes = base64.b64decode(pix_result['qr_code_base64'])
            
            await query.message.reply_photo(
                photo=qr_image_bytes,
                caption=f"💎 **Upgrade para Admin VIP**\n\n"
                       f"**Valor:** R$ {amount:.2f}\n"
                       f"**Descrição:** {description}\n"
                       f"**Referência:** `{external_reference}`\n\n"
                       f"📱 **Escaneie o QR Code acima para pagar via PIX**\n\n"
                       f"⏰ O pagamento será processado automaticamente em alguns minutos.\n"
                       f"🔄 Clique em 'Verificar Pagamento' após realizar o pagamento.",
                reply_markup=reply_markup,
                parse_mode='Markdown'
            )
        else:
            error_msg = "❌ Erro ao gerar QR Code de pagamento."
            if pix_result:
                error_msg += f"\n\nDetalhes: {pix_result.get('error', 'Erro desconhecido')}"
            else:
                error_msg += "\n\nNenhum resultado retornado pela função de geração de PIX."
            
            await query.message.reply_text(
                error_msg,
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]])
            )
        return
    
    # Verificar se é um callback de verificação de pagamento VIP
    if query.data == "admin_check_vip_payment":
        admin_id = update.effective_user.id
        
        # Verificar se há pagamento pendente
        db = Database()
        try:
            db.connect()
            if db.connection:
                payment = db.execute_fetch_one(
                    """SELECT * FROM admin_vip_payments 
                    WHERE admin_id = %s AND status = 'pending' 
                    ORDER BY created_at DESC LIMIT 1""",
                    (admin_id,)
                )
                
                if payment:
                    # Verificar status do pagamento (simulação - você pode integrar com seu sistema de pagamento)
                    # Por enquanto, vamos simular que o pagamento foi aprovado
                    await query.answer("🔄 Verificando pagamento...", show_alert=True)
                    
                    # Simular aprovação do pagamento
                    db.execute_query(
                        """UPDATE admin_vip_payments 
                        SET status = 'approved', approved_at = NOW() 
                        WHERE id = %s""",
                        (payment['id'],),
                        commit=True
                    )
                    
                    # Atualizar admin para VIP
                    db.execute_query(
                        """UPDATE admins 
                        SET is_vip = 1 
                        WHERE admin_id = %s""",
                        (admin_id,),
                        commit=True
                    )
                    
                    await query.message.reply_text(
                        "🎉 **Parabéns! Você agora é um Admin VIP!**\n\n"
                        "✅ Seu pagamento foi aprovado\n"
                        "🔓 Todos os recursos foram liberados\n"
                        "📢 Agora você pode usar o Broadcast e outras funções exclusivas\n\n"
                        "🔄 Recarregue o menu para ver as novas funcionalidades!",
                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Voltar ao Menu", callback_data="admin_back")]]),
                        parse_mode='Markdown'
                    )
                else:
                    await query.answer("❌ Nenhum pagamento pendente encontrado", show_alert=True)
        except Exception as e:
            logger.error(f"Erro ao verificar pagamento VIP: {e}")
            await query.answer("❌ Erro ao verificar pagamento", show_alert=True)
        finally:
            db.close()
        return
    
    # Verificar se é um callback de agendamento de mensagens
    if query.data == "admin_schedule_messages":
        # Menu de agendamento de mensagens
        keyboard = [
            [InlineKeyboardButton("➕ Nova Mensagem", callback_data="admin_schedule_new")],
            [InlineKeyboardButton("📋 Mensagens Pendentes", callback_data="admin_schedule_pending")],
            [InlineKeyboardButton("📊 Histórico", callback_data="admin_schedule_history")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "⏰ Agendamento de Mensagens\n\n"
            "Gerencie mensagens agendadas para envio automático:\n\n"
            "➕ Nova Mensagem: Criar nova mensagem agendada\n"
            "📋 Pendentes: Ver mensagens agendadas pendentes\n"
            "📊 Histórico: Ver histórico de mensagens enviadas",
            reply_markup=reply_markup
        )
        return
    
    elif query.data == "admin_schedule_new":
        # Iniciar criação de nova mensagem agendada
        context.user_data['scheduling_step'] = 'message_text'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_schedule_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "➕ Nova Mensagem Agendada\n\n"
            "Digite o texto da mensagem que deseja agendar:",
            reply_markup=reply_markup
        )
        return
    
    elif query.data == "admin_schedule_pending":
        # Mostrar mensagens pendentes
        pending_messages = get_scheduled_messages(status='pending', limit=10)
        
        if not pending_messages:
            keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_schedule_messages")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.message.edit_text(
                "📋 Mensagens Pendentes\n\n"
                "Nenhuma mensagem agendada pendente.",
                reply_markup=reply_markup
            )
            return
        
        text = "📋 Mensagens Pendentes\n\n"
        keyboard = []
        
        for msg in pending_messages:
            scheduled_time = msg['scheduled_date'].strftime('%d/%m/%Y %H:%M')
            target_text = {
                'all_users': 'Todos os usuários',
                'vip_users': 'Usuários VIP',
                'specific_users': 'Usuários específicos'
            }.get(msg['target_type'], msg['target_type'])
            
            text += f"🆔 ID: {msg['id']}\n"
            text += f"📅 Agendada para: {scheduled_time}\n"
            text += f"👥 Destinatários: {target_text}\n"
            text += f"📝 Mensagem: {msg['message_text'][:50]}{'...' if len(msg['message_text']) > 50 else ''}\n\n"
            
            keyboard.append([
                InlineKeyboardButton(f"❌ Cancelar {msg['id']}", callback_data=f"admin_schedule_cancel_{msg['id']}"),
                InlineKeyboardButton(f"👁️ Ver {msg['id']}", callback_data=f"admin_schedule_view_{msg['id']}")
            ])
        
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin_schedule_messages")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(text, reply_markup=reply_markup)
        return
    
    elif query.data == "admin_schedule_history":
        # Mostrar histórico de mensagens
        sent_messages = get_scheduled_messages(status='sent', limit=10)
        
        if not sent_messages:
            keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_schedule_messages")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            await query.message.edit_text(
                "📊 Histórico de Mensagens\n\n"
                "Nenhuma mensagem enviada ainda.",
                reply_markup=reply_markup
            )
            return
        
        text = "📊 Histórico de Mensagens\n\n"
        keyboard = []
        
        for msg in sent_messages:
            sent_time = msg['sent_at'].strftime('%d/%m/%Y %H:%M') if msg['sent_at'] else 'N/A'
            target_text = {
                'all_users': 'Todos os usuários',
                'vip_users': 'Usuários VIP',
                'specific_users': 'Usuários específicos'
            }.get(msg['target_type'], msg['target_type'])
            
            text += f"🆔 ID: {msg['id']}\n"
            text += f"📅 Enviada em: {sent_time}\n"
            text += f"👥 Destinatários: {target_text}\n"
            text += f"✅ Enviadas: {msg['successful_sends']}\n"
            text += f"❌ Falhas: {msg['failed_sends']}\n"
            text += f"📝 Mensagem: {msg['message_text'][:50]}{'...' if len(msg['message_text']) > 50 else ''}\n\n"
        
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin_schedule_messages")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(text, reply_markup=reply_markup)
        return
    
    elif query.data.startswith("admin_schedule_cancel_"):
        # Cancelar mensagem agendada
        message_id = int(query.data.split('_')[-1])
        if cancel_scheduled_message(message_id):
            await query.answer("✅ Mensagem cancelada com sucesso!")
            # Atualizar a lista
            await query.message.edit_text(
                "📋 Mensagens Pendentes\n\n"
                "Mensagem cancelada com sucesso!",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_schedule_pending")]])
            )
        else:
            await query.answer("❌ Erro ao cancelar mensagem!")
        return
    
    elif query.data.startswith("admin_schedule_view_"):
        # Ver detalhes da mensagem
        message_id = int(query.data.split('_')[-1])
        messages = get_scheduled_messages()
        message = next((m for m in messages if m['id'] == message_id), None)
        
        if not message:
            await query.answer("❌ Mensagem não encontrada!")
            return
        
        scheduled_time = message['scheduled_date'].strftime('%d/%m/%Y %H:%M')
        target_text = {
            'all_users': 'Todos os usuários',
            'vip_users': 'Usuários VIP',
            'specific_users': 'Usuários específicos'
        }.get(message['target_type'], message['target_type'])
        
        text = f"👁️ Detalhes da Mensagem #{message['id']}\n\n"
        text += f"📅 Agendada para: {scheduled_time}\n"
        text += f"👥 Destinatários: {target_text}\n"
        text += f"📊 Status: {message['status'].upper()}\n"
        if message['successful_sends']:
            text += f"✅ Enviadas: {message['successful_sends']}\n"
        if message['failed_sends']:
            text += f"❌ Falhas: {message['failed_sends']}\n"
        text += f"\n📝 Mensagem:\n{message['message_text']}"
        
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_schedule_pending")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(text, reply_markup=reply_markup)
        return
    
    elif query.data == "admin_schedule_target_all":
        # Selecionar todos os usuários
        context.user_data['scheduled_target_type'] = 'all_users'
        context.user_data['scheduling_step'] = 'scheduled_date'
        
        # Gerar exemplo com data atual
        example_time = datetime.now()
        example_str = example_time.strftime("%d/%m/%Y %H:%M")
        
        await query.message.edit_text(
            "✅ Destinatários: Todos os usuários\n\n"
            "Agora digite a data e hora para envio da mensagem:\n\n"
            "Formato: DD/MM/AAAA HH:MM\n"
            f"Exemplo: {example_str}"
        )
        return
    
    elif query.data == "admin_schedule_target_vip":
        # Selecionar usuários VIP
        context.user_data['scheduled_target_type'] = 'vip_users'
        context.user_data['scheduling_step'] = 'scheduled_date'
        
        # Gerar exemplo com data atual
        example_time = datetime.now()
        example_str = example_time.strftime("%d/%m/%Y %H:%M")
        
        await query.message.edit_text(
            "✅ Destinatários: Usuários VIP\n\n"
            "Agora digite a data e hora para envio da mensagem:\n\n"
            "Formato: DD/MM/AAAA HH:MM\n"
            f"Exemplo: {example_str}"
        )
        return
    
    elif query.data == "admin_schedule_target_specific":
        # Selecionar usuários específicos
        context.user_data['scheduled_target_type'] = 'specific_users'
        context.user_data['scheduling_step'] = 'specific_users'
        
        await query.message.edit_text(
            "✅ Destinatários: Usuários Específicos\n\n"
            "Digite os IDs dos usuários separados por vírgula ou espaço:\n\n"
            "Exemplo: 123456789, 987654321, 555666777"
        )
        return

    # Verificar se é um callback de broadcast
    if query.data == "admin_broadcast":
        # Menu de broadcast
        keyboard = [
            [InlineKeyboardButton("📢 Enviar para Todos", callback_data="admin_broadcast_all")],
            [InlineKeyboardButton("👥 Enviar para VIPs", callback_data="admin_broadcast_vip")],
            [InlineKeyboardButton("📹 Enviar Vídeo para Todos", callback_data="admin_broadcast_video_all")],
            [InlineKeyboardButton("📹 Enviar Vídeo para VIPs", callback_data="admin_broadcast_video_vip")],
            [InlineKeyboardButton("⭕ Enviar Vídeo Circular para Todos", callback_data="admin_broadcast_videonote_all")],
            [InlineKeyboardButton("⭕ Enviar Vídeo Circular para VIPs", callback_data="admin_broadcast_videonote_vip")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📢 Broadcast\n\nEscolha o tipo de broadcast:\n\n"
            "📹 Vídeo Normal: Formato retangular tradicional\n"
            "⭕ Vídeo Circular: Formato circular (video_note)",
            reply_markup=reply_markup
        )
        return
    
    elif query.data == "admin_broadcast_videonote_all":
        # Preparar para enviar vídeo circular para todos
        context.user_data['broadcast_type'] = 'videonote_all'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "⭕ Enviar vídeo circular para todos os usuários\n\n"
            "📋 Requisitos do vídeo circular:\n"
            "• Formato quadrado (ex: 240x240)\n"
            "• Duração máxima: 60 segundos\n"
            "• Será exibido como círculo no app\n\n"
            "Envie o vídeo que deseja compartilhar:",
            reply_markup=reply_markup
        )
        return
    
    elif query.data == "admin_broadcast_videonote_vip":
        # Preparar para enviar vídeo circular para VIPs
        context.user_data['broadcast_type'] = 'videonote_vip'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "⭕ Enviar vídeo circular para usuários VIP\n\n"
            "📋 Requisitos do vídeo circular:\n"
            "• Formato quadrado (ex: 240x240)\n"
            "• Duração máxima: 60 segundos\n"
            "• Será exibido como círculo no app\n\n"
            "Envie o vídeo que deseja compartilhar:",
            reply_markup=reply_markup
        )
        return
    
    elif query.data == "admin_broadcast_video_all":
        # Preparar para enviar vídeo para todos
        context.user_data['broadcast_type'] = 'video_all'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📹 Enviar vídeo para todos os usuários\n\n"
            "Primeiro, envie o vídeo que deseja compartilhar:",
            reply_markup=reply_markup
        )
        return
    
    elif query.data == "admin_broadcast_video_vip":
        # Preparar para enviar vídeo para VIPs
        context.user_data['broadcast_type'] = 'video_vip'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📹 Enviar vídeo para usuários VIP\n\n"
            "Primeiro, envie o vídeo que deseja compartilhar:",
            reply_markup=reply_markup
        )
        return
    
    # Verificar se é um callback de configuração de provedores PIX
    elif query.data == "admin_pix_providers":
        # Menu de configuração de provedores PIX
        provider_manager = get_pix_provider_manager()
        available_providers = provider_manager.get_available_providers()
        default_provider = config.get('pix_provider', 'mercadopago')
        
        # Status dos provedores
        mercadopago_enabled = config.get('mercadopago_enabled', False)
        cnpay_enabled = config.get('cnpay_enabled', False)
        
        keyboard = [
            [InlineKeyboardButton(
                f"{'🟢' if cnpay_enabled else '🔴'} CNPay",
                callback_data="admin_toggle_cnpay"
            )],
            [InlineKeyboardButton("🔧 Configurar CNPay", callback_data="admin_config_cnpay")],
            [InlineKeyboardButton("🎯 Definir Provedor Padrão", callback_data="admin_set_default_provider")],
            [InlineKeyboardButton("🧪 Testar Conexões", callback_data="admin_test_providers")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        status_text = "🔧 CONFIGURAÇÃO DE PROVEDORES PIX\n\n"
        status_text += f"📱 Provedores Disponíveis:\n"
        status_text += f"   🏦 CNPay: {'✅ Ativo' if cnpay_enabled else '❌ Inativo'}\n\n"
        status_text += f"🎯 Provedor Padrão: {default_provider.title()}\n"
        status_text += f"📊 Provedores Configurados: {len(available_providers)}\n\n"
        status_text += "Escolha uma opção:"
        
        await query.message.edit_text(status_text, reply_markup=reply_markup)
        return
    
    elif query.data == "admin_toggle_cnpay":
        # Alternar status do CNPay
        current_cnpay_status = config.get('cnpay_enabled', False)
        config['cnpay_enabled'] = not current_cnpay_status
        new_status_message = "ativado" if not current_cnpay_status else "desativado"
        status_message = f"✅ CNPay {new_status_message}!"

        # Se estamos ativando o CNPay, podemos defini-lo como padrão
        if config['cnpay_enabled']:
            config['pix_provider'] = 'cnpay'
        
        if save_config(config):
            await query.answer(status_message)
            
            # Recarregar menu
            provider_manager = get_pix_provider_manager()
            available_providers = provider_manager.get_available_providers()
            default_provider = config.get('pix_provider', 'mercadopago')
            cnpay_enabled = config.get('cnpay_enabled', False)
            
            keyboard = [
                [InlineKeyboardButton(
                    f"{'🟢' if cnpay_enabled else '🔴'} CNPay",
                    callback_data="admin_toggle_cnpay"
                )],
                [InlineKeyboardButton("🔧 Configurar CNPay", callback_data="admin_config_cnpay")],
                [InlineKeyboardButton("🎯 Definir Provedor Padrão", callback_data="admin_set_default_provider")],
                [InlineKeyboardButton("🧪 Testar Conexões", callback_data="admin_test_providers")],
                [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            status_text = "🔧 CONFIGURAÇÃO DE PROVEDORES PIX\n\n"
            status_text += f"📱 Provedores Disponíveis:\n"
            status_text += f"   🏦 CNPay: {'✅ Ativo' if cnpay_enabled else '❌ Inativo'}\n\n"
            status_text += f"🎯 Provedor Padrão: {default_provider.title()}\n"
            status_text += f"📊 Provedores Configurados: {len(available_providers)}\n\n"
            status_text += "Escolha uma opção:"
            
            await query.message.edit_text(status_text, reply_markup=reply_markup)
        else:
            await query.answer("❌ Erro ao salvar configuração")
        return
    
    elif query.data == "admin_config_mercadopago":
        # Configurar MercadoPago
        keyboard = [
            [InlineKeyboardButton("🔑 Token de Acesso", callback_data="admin_edit_mp_token")],
            [InlineKeyboardButton("🌍 Ambiente", callback_data="admin_edit_mp_environment")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_pix_providers")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        current_token = config.get('mercadopago_access_token', 'Não configurado')
        current_env = config.get('mercadopago_environment', 'production')
        
        config_text = "🔧 CONFIGURAÇÃO MERCADOPAGO\n\n"
        config_text += f"🔑 Token: {'********' if current_token != 'Não configurado' else 'Não configurado'}\n"
        config_text += f"🌍 Ambiente: {current_env}\n\n"
        config_text += "Escolha o que deseja configurar:"
        
        await query.message.edit_text(config_text, reply_markup=reply_markup)
        return
    
    elif query.data == "admin_config_cnpay":
        # Configurar CNPay
        keyboard = [
            [InlineKeyboardButton("🔑 API Key", callback_data="admin_edit_cnpay_key")],
            [InlineKeyboardButton("🔐 API Secret", callback_data="admin_edit_cnpay_secret")],
            [InlineKeyboardButton("🌍 Ambiente", callback_data="admin_edit_cnpay_environment")],
            [InlineKeyboardButton("🌐 Webhook URL", callback_data="admin_edit_cnpay_webhook")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_pix_providers")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        current_key = config.get('cnpay_api_key', 'Não configurado')
        current_env = config.get('cnpay_environment', 'sandbox')
        current_webhook = config.get('cnpay_webhook_url', 'Não configurado')
        
        config_text = "🔧 CONFIGURAÇÃO CNPAY\n\n"
        config_text += f"🔑 API Key: {'********' if current_key != 'Não configurado' else 'Não configurado'}\n"
        config_text += f"🔐 API Secret: {'********' if config.get('cnpay_api_secret') else 'Não configurado'}\n"
        config_text += f"🌍 Ambiente: {current_env}\n"
        config_text += f"🌐 Webhook: {current_webhook}\n\n"
        config_text += "Escolha o que deseja configurar:"
        
        await query.message.edit_text(config_text, reply_markup=reply_markup)
        return
    
    elif query.data == "admin_set_default_provider":
        # Definir provedor padrão
        current_default = config.get('pix_provider', 'mercadopago')
        mercadopago_enabled = config.get('mercadopago_enabled', False)
        cnpay_enabled = config.get('cnpay_enabled', False)
        
        keyboard = []
        if cnpay_enabled:
            keyboard.append([InlineKeyboardButton(
                f"{'✅ ' if current_default == 'cnpay' else ''}CNPay",
                callback_data="admin_set_provider_cnpay"
            )])
        
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin_pix_providers")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        text = "🎯 DEFINIR PROVEDOR PADRÃO\n\n"
        text += f"Provedor atual: {current_default.title()}\n\n"
        text += "Escolha o novo provedor padrão:"
        
        await query.message.edit_text(text, reply_markup=reply_markup)
        return
    
    elif query.data.startswith("admin_set_provider_"):
        # Definir provedor específico como padrão
        provider = query.data.split('_')[-1]
        config['pix_provider'] = provider
        
        if save_config(config):
            await query.answer(f"✅ {provider.title()} definido como padrão!")
            # Recarregar o menu
            await handle_admin_callback(update, context)
        else:
            await query.answer("❌ Erro ao salvar configuração")
        return
    
    elif query.data == "admin_test_providers":
        # Testar conexões dos provedores
        await query.answer("🧪 Testando conexões...")
        
        test_results = []
        provider_manager = get_pix_provider_manager()
        
        for provider_name, provider in provider_manager.providers.items():
            try:
                # Teste simples de conexão
                if provider_name == 'mercadopago':
                    if provider.config.get('mercadopago_access_token'):
                        test_results.append(f"✅ MercadoPago: Configurado")
                    else:
                        test_results.append(f"❌ MercadoPago: Token não configurado")
                elif provider_name == 'cnpay':
                    if provider.api_key and provider.api_secret:
                        test_results.append(f"✅ CNPay: Configurado")
                    else:
                        test_results.append(f"❌ CNPay: Credenciais não configuradas")
            except Exception as e:
                test_results.append(f"❌ {provider_name.title()}: Erro - {str(e)}")
        
        if not test_results:
            test_results.append("❌ Nenhum provedor configurado")
        
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_pix_providers")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        test_text = "🧪 TESTE DE CONEXÕES\n\n"
        test_text += "\n".join(test_results)
        
        await query.message.edit_text(test_text, reply_markup=reply_markup)
        return
    elif query.data == "admin_settings":
        # Menu de configurações
        keyboard = [
            [InlineKeyboardButton("🔑 Token do Bot", callback_data="admin_edit_bot_token")],
            [InlineKeyboardButton("🔧 Provedores PIX", callback_data="admin_pix_providers")],
            [InlineKeyboardButton("📎 Arquivo de Boas-vindas", callback_data="admin_welcome_file")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "⚙️ Configurações\n\nEscolha uma opção para editar:",
            reply_markup=reply_markup
        )
        return
    if query.data == "admin_messages":
        # Limpa o estado de edição se existir
        if 'editing' in context.user_data:
            del context.user_data['editing']
        keyboard = [
            [InlineKeyboardButton("🏁 Mensagem de Início", callback_data="admin_edit_start_message")],
            [InlineKeyboardButton("👋 Mensagem de Boas-vindas", callback_data="admin_edit_welcome_message")],
            [InlineKeyboardButton("💎 Mensagem de Pagamento", callback_data="admin_edit_payment_message")],
            [InlineKeyboardButton("✅ Mensagem de Sucesso", callback_data="admin_edit_success_message")],
            [InlineKeyboardButton("❌ Mensagem de Erro", callback_data="admin_edit_error_message")],
            [InlineKeyboardButton("📝 Instruções PIX", callback_data="admin_edit_pix_instructions")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        text = "📝 Mensagens do Bot\n\nMensagens atuais:\n\n"
        text += f"🏁 Início: {messages.get('start_message', 'Não definida')[:50]}...\n\n"
        text += f"👋 Boas-vindas: {messages.get('welcome_message', 'Não definida')[:50]}...\n\n"
        text += f"💎 Pagamento: {messages.get('payment_instructions', 'Não definida')[:50]}...\n\n"
        text += f"✅ Sucesso: {messages.get('payment_success', 'Não definida')[:50]}...\n\n"
        text += f"❌ Erro: {messages.get('payment_error', 'Não definida')[:50]}...\n\n"
        text += f"📝 PIX: {messages.get('pix_automatico_instructions', 'Não definida')[:50]}...\n\n"
        text += "Escolha uma mensagem para editar:"

        await query.message.edit_text(text, reply_markup=reply_markup)

        return
    elif query.data == "admin_edit_start_message":
        logger.info("[DEBUG] Entrou no bloco admin_edit_start_message (handle_admin_edit)")
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            # Apagar a mensagem anterior
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de início para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="🏁 Editar Mensagem de Início\n\n"
                     f"Mensagem atual:\n{messages.get('start_message', 'Não definida')}\n\n"
                     "Envie a nova mensagem de início:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de início enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de início: {e}")
        context.user_data['editing'] = 'start_message'
        return
    elif query.data == "admin_edit_welcome_message":
        logger.info("[DEBUG] Entrou no bloco admin_edit_welcome_message (handle_admin_edit)")
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            # Apagar a mensagem anterior
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de boas-vindas para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="👋 Editar Mensagem de Boas-vindas\n\n"
                     f"Mensagem atual:\n{messages.get('welcome_message', 'Não definida')}\n\n"
                     "Envie a nova mensagem de boas-vindas:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de boas-vindas enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de boas-vindas: {e}")
        context.user_data['editing'] = 'welcome_message'
        return
    elif query.data == "admin_edit_payment_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de pagamento para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="💎 Editar Mensagem de Pagamento\n\n"
                     f"Mensagem atual:\n{messages.get('payment_instructions', 'Não definida')}\n\n"
                     "Envie a nova mensagem de pagamento:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de pagamento enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de pagamento: {e}")
        context.user_data['editing'] = 'payment_instructions'
        return
    elif query.data == "admin_edit_success_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de sucesso para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="✅ Editar Mensagem de Sucesso\n\n"
                     f"Mensagem atual:\n{messages.get('payment_success', 'Não definida')}\n\n"
                     "Envie a nova mensagem de sucesso:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de sucesso enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de sucesso: {e}")
        context.user_data['editing'] = 'payment_success'
        return
    elif query.data == "admin_edit_error_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de erro para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="❌ Editar Mensagem de Erro\n\n"
                     f"Mensagem atual:\n{messages.get('payment_error', 'Não definida')}\n\n"
                     "Envie a nova mensagem de erro:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de erro enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de erro: {e}")
        context.user_data['editing'] = 'payment_error'
        return
    elif query.data == "admin_edit_pix_instructions":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de instruções PIX para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="📝 Editar Instruções PIX\n\n"
                     f"Mensagem atual:\n{messages.get('pix_automatico_instructions', 'Não definida')}\n\n"
                     "Envie a nova mensagem de instruções PIX:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de instruções PIX enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de instruções PIX: {e}")
        context.user_data['editing'] = 'pix_automatico_instructions'
        return
    elif query.data == "admin_broadcast_all":
        # Preparar para enviar para todos
        context.user_data['broadcast_type'] = 'all'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📢 Enviar mensagem para todos os usuários\n\n"
            "Digite a mensagem que deseja enviar:",
            reply_markup=reply_markup
        )
        return
    elif query.data == "admin_broadcast_vip":
        # Preparar para enviar para VIPs
        context.user_data['broadcast_type'] = 'vip'
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📢 Enviar mensagem para usuários VIP\n\n"
            "Digite a mensagem que deseja enviar:",
            reply_markup=reply_markup
        )
        return
    

    
    # Se não for broadcast, continua com o código existente
    action = '_'.join(query.data.split('_')[1:])
    logger.info(f"Ação: {action}")
    
    if action == "stats":
        # Mostrar estatísticas
        stats = get_user_stats()
        
        text = "📊 Estatísticas do Bot\n\n"
        text += f"Total de Usuários: {stats['total_users']}\n"
        text += f"Total de VIPs: {stats['vip_users']}\n"
        text += f"Última Atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n\n"
        text += "👥 Últimos Usuários:\n"
        
        # Mostrar os últimos 5 usuários
        for user in stats['recent_users']:
            text += f"\nID: {user['id']}"
            if user['username']:
                text += f"\nUsername: @{user['username']}"
            text += f"\nNome: {user['first_name']}"
            if user['last_name']:
                text += f" {user['last_name']}"
            text += f"\nData: {user['joined_date']}"
            text += f"\nVIP: {'✅' if user.get('is_vip', False) else '❌'}\n"
        
        keyboard = [
            [InlineKeyboardButton("📊 Exportar Excel", callback_data="admin_export_excel")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(text, reply_markup=reply_markup)
        
    elif action == "vip_users":
        # Listar usuários VIP
        active_subscriptions = get_all_active_subscriptions()
        
        if active_subscriptions:
            text = "👥 Usuários VIP Ativos:\n\n"
            for sub in active_subscriptions:
                text += f"ID: {sub['user_id']}\n"
                text += f"Nome: {sub['first_name']} {sub['last_name'] or ''}\n"
                text += f"Plano: {sub['plan_name']}\n"
                text += f"Expira em: {sub['end_date']}\n\n"
        else:
            text = "Nenhum usuário VIP ativo."
        
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(text, reply_markup=reply_markup)
        
    elif action == "maintenance":
        # Modo manutenção
        keyboard = [
            [InlineKeyboardButton(
                "🔴 Desativar Manutenção" if config.get('maintenance_mode', False) else "🟢 Ativar Manutenção",
                callback_data="admin_toggle_maintenance"
            )],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        status = "ativado" if config.get('maintenance_mode', False) else "desativado"
        await query.message.edit_text(
            f"🔄 Modo Manutenção\n\nStatus atual: {status}",
            reply_markup=reply_markup
        )
        
    elif action == "back":
        # Limpa o estado de edição se existir
        if 'editing' in context.user_data:
            del context.user_data['editing']
        # Menu principal com layout melhorado
        keyboard = [
            [InlineKeyboardButton("📊 Estatísticas", callback_data="admin_stats")],
            [
                InlineKeyboardButton("⚙️ Configurações", callback_data="admin_settings"),
                InlineKeyboardButton("👥 Usuários VIP", callback_data="admin_vip_users")
            ],
            [InlineKeyboardButton("💎 Planos VIP", callback_data="admin_vip_plans")],
            [InlineKeyboardButton("📝 Mensagens", callback_data="admin_messages")],
            [InlineKeyboardButton("⏰ Agendar Mensagens", callback_data="admin_schedule_messages")],
            [InlineKeyboardButton("🔄 Manutenção", callback_data="admin_maintenance")],
            [InlineKeyboardButton("👤 Gerenciar Admins", callback_data="admin_manage_admins")],
            [InlineKeyboardButton("⚒️ Suporte", url=config.get('support_admin', 'https://t.me/suporte'))]  # Botão de suporte
        ]
        
        # Adicionar botão de broadcast (com emoji de cadeado para admins não-VIP)
        if is_admin_vip(update.effective_user.id):
            keyboard.insert(6, [InlineKeyboardButton("📢 Broadcast", callback_data="admin_broadcast")])
        else:
            keyboard.insert(6, [InlineKeyboardButton("📢🔒 Broadcast (VIP)", callback_data="admin_broadcast_locked")])

        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🔧 Painel Administrativo\n\nEscolha uma opção:",
            reply_markup=reply_markup
        )

    elif query.data == "admin_edit_cnpay_environment":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_config_cnpay")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🌍 Editar Ambiente do CNPay\n\n"
            f"Ambiente atual: {config.get('cnpay_environment', 'sandbox')}\n\n"
            "Envie o novo ambiente (sandbox ou production):",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'cnpay_environment'
        return

    elif query.data == "admin_edit_cnpay_webhook":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_config_cnpay")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🌐 Editar Webhook URL do CNPay\n\n"
            f"Webhook atual: {config.get('cnpay_webhook_url', 'Não configurado')}\n\n"
            "Envie a nova URL do webhook:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'cnpay_webhook_url'
        return
   
    elif query.data == "admin_welcome_file":
        # Menu de arquivo de boas-vindas
        welcome_file_config = config.get('welcome_file', {})
        is_enabled = welcome_file_config.get('enabled', False)
        file_type = welcome_file_config.get('file_type', 'photo')
        print(messages)
        caption = messages.get('start_message', 'sem mensagem definida!')
        print(caption)
        
        keyboard = [
            [InlineKeyboardButton(
                f"{'🔴' if not is_enabled else '🟢'} {'Desativar' if is_enabled else 'Ativar'} Arquivo",
                callback_data="admin_toggle_welcome_file"
            )],
            [InlineKeyboardButton("📎 Enviar Novo Arquivo", callback_data="admin_upload_welcome_file")],
            [InlineKeyboardButton("📝 Editar Legenda", callback_data="admin_edit_welcome_caption")],
            [InlineKeyboardButton("🗑️ Remover Arquivo", callback_data="admin_remove_welcome_file")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        status_text = "📎 Arquivo de Boas-vindas\n\n"
        if is_enabled:
            status_text += f"✅ Status: Ativado\n"
            status_text += f"📁 Tipo: {file_type.title()}\n"
            status_text += f"📝 Legenda: {caption}\n"
        else:
            status_text += f"❌ Status: Desativado\n"
        
        status_text += "\nEscolha uma opção:"
        
        await query.message.edit_text(
            status_text,
            reply_markup=reply_markup
        )
        return

    # --- INÍCIO: Menu de gerenciamento de planos VIP ---
    elif query.data == "admin_vip_plans":
        db = Database()
        try:
            db.connect()
            plans = db.execute_fetch_all("SELECT * FROM vip_plans")
        finally:
            db.close()
        keyboard = []
        for plan in plans:
            keyboard.append([
                InlineKeyboardButton(f"✏️ {plan['name']} (R${plan['price']:.2f})", callback_data=f"admin_edit_plan_{plan['id']}"),
                InlineKeyboardButton("🗑️", callback_data=f"admin_remove_plan_{plan['id']}")
            ])
        keyboard.append([InlineKeyboardButton("➕ Adicionar Novo Plano", callback_data="admin_add_plan")])
        keyboard.append([InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")])
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "💎 Gerenciar Planos VIP\n\nSelecione um plano para editar ou remova/adicione novos planos:",
            reply_markup=reply_markup
        )
        return
    
    # Tratamento para remover plano
    elif query.data.startswith("admin_remove_plan_"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.answer("❌ Plano não encontrado!")
            return
        
        # Verificar se há usuários ativos com este plano
        db = Database()
        try:
            db.connect()
            active_users = db.execute_fetch_all(
                "SELECT COUNT(*) as count FROM subscriptions WHERE plan_id = %s AND is_active = TRUE",
                (plan_id,)
            )
            user_count = active_users[0]['count'] if active_users else 0
        finally:
            db.close()
        
        keyboard = [
            [InlineKeyboardButton("✅ Confirmar Remoção", callback_data=f"admin_confirm_remove_plan_{plan_id}")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="admin_vip_plans")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        warning_text = f"🗑️ Remover Plano: {plan['name']}\n\n"
        warning_text += f"💰 Preço: R${plan['price']:.2f}\n"
        warning_text += f"⏱️ Duração: {plan['duration_days']} dias\n"
        warning_text += f"👥 Usuários ativos: {user_count}\n\n"
        
        if user_count > 0:
            warning_text += "⚠️ ATENÇÃO: Este plano possui usuários ativos!\n"
            warning_text += "A remoção pode afetar as assinaturas existentes.\n\n"
        
        warning_text += "Tem certeza que deseja remover este plano?\n"
        warning_text += "Esta ação não pode ser desfeita."
        
        await query.message.edit_text(warning_text, reply_markup=reply_markup)
        return
    
    elif query.data == "admin_upload_welcome_file":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_welcome_file")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "📎 Enviar Novo Arquivo de Boas-vindas\n\n"
            "Envie uma foto ou vídeo que será usado como arquivo de boas-vindas.\n\n"
            "⚠️ O arquivo deve ser menor que 50MB.",
            reply_markup=reply_markup
        )
        context.user_data['waiting_for_welcome_file'] = True
        return
    
    
    elif query.data == "admin_confirm_remove_welcome_file":
        try:
            logger.info("Iniciando remoção do arquivo de boas-vindas...")
            
            # NÃO responder o callback aqui, pois já foi respondido em handle_admin_callback
            # Load config and update in memory first
            config = load_config()

            if not config:
                logger.error("Falha ao carregar as configurações.")
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="❌ Erro ao carregar as configurações. Tente novamente."
                )
                return
                
            caption = messages.get('start_message', 'sem mensagem definida!')

            # Update welcome file configuration
            config['welcome_file'] = {
                'enabled': False,
                'file_id': '',
                'file_type': 'photo',
                'caption': caption
            }
            
            # Save config and handle result
            if save_config(config):
                logger.info("Arquivo de boas-vindas removido com sucesso")
                keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_welcome_file")]]
                reply_markup = InlineKeyboardMarkup(keyboard)
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Arquivo de boas-vindas removido com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                logger.error("Falha ao salvar as configurações após remoção do arquivo")
                keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_welcome_file")]]
                reply_markup = InlineKeyboardMarkup(keyboard)
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="❌ Ocorreu um erro ao remover o arquivo. Tente novamente.",
                    reply_markup=reply_markup
                )
            
            
        except Exception as e:
            logger.error(f"Erro ao processar remoção do arquivo de boas-vindas: {e}")
            try:
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=f"❌ Ocorreu um erro ao processar: {str(e)}"
                )
            except Exception as e2:
                logger.error(f"Erro ao enviar mensagem de erro: {e2}")
        return
    
    
    # Tratamento para adicionar novo plano
    elif query.data == "admin_add_plan":
        context.user_data['adding_plan'] = {'step': 'name'}
        keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_vip_plans")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "➕ Adicionar Novo Plano VIP\n\n"
            "Digite o nome do novo plano:",
            reply_markup=reply_markup
        )
        return
    # --- FIM: Menu de gerenciamento de planos VIP ---

    if query.data == "admin_export_excel":
        await query.message.edit_text("📊 Gerando relatório completo...")
        
        try:
            # Obter dados
            stats = get_user_stats()
            all_users = get_all_users()
            subscriptions = get_subscriptions_for_export()
            expiring_subs = get_expiring_subscriptions()
            
            # Criar workbook
            wb = openpyxl.Workbook()
            
            # Remover aba padrão
            wb.remove(wb.active)
            
            # === ABA 1: RESUMO EXECUTIVO ===
            ws_summary = wb.create_sheet("📊 Resumo Executivo")
            
            # Estatísticas gerais
            ws_summary.append(["RELATÓRIO DE ASSINATURAS VIP", ""])
            ws_summary.append(["Data da Exportação:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
            ws_summary.append([""])
            ws_summary.append(["ESTATÍSTICAS GERAIS", ""])
            ws_summary.append(["Total de Usuários:", stats['total_users']])
            ws_summary.append(["Total de VIPs Ativos:", stats['vip_users']])
            ws_summary.append(["Total de Assinaturas:", len(subscriptions)])
            ws_summary.append(["Assinaturas Expirando (≤3 dias):", len(expiring_subs)])
            ws_summary.append([""])
            
            # Estatísticas por plano
            plan_stats = {}
            for sub in subscriptions:
                plan_name = sub['plan_name']
                if plan_name not in plan_stats:
                    plan_stats[plan_name] = {'count': 0, 'revenue': 0}
                plan_stats[plan_name]['count'] += 1
                plan_stats[plan_name]['revenue'] += float(sub['price'] or 0)
            
            ws_summary.append(["ESTATÍSTICAS POR PLANO", ""])
            ws_summary.append(["Plano", "Quantidade", "Receita Total (R$)"])
            total_revenue = 0
            for plan_name, data in plan_stats.items():
                ws_summary.append([plan_name, data['count'], f"R$ {data['revenue']:.2f}"])
                total_revenue += data['revenue']
            ws_summary.append(["", "", ""])
            ws_summary.append(["RECEITA TOTAL:", f"R$ {total_revenue:.2f}"])
            
            # === ABA 2: ASSINATURAS DETALHADAS ===
            ws_subs = wb.create_sheet("📋 Assinaturas Detalhadas")
            
            # Cabeçalho
            headers = [
                "ID Assinatura", "ID Usuário", "Username", "Nome Completo",
                "Plano", "Preço (R$)", "Duração (dias)", "Método Pagamento",
                "Status Pagamento", "Data Início", "Data Fim", "Dias Restantes",
                "Dias Pagos", "Total Dias", "Status Expiração", "Permanente",
                "Data Criação"
            ]
            ws_subs.append(headers)
            
            # Dados das assinaturas
            for sub in subscriptions:
                full_name = f"{sub['first_name'] or ''} {sub['last_name'] or ''}".strip()
                
                row = [
                    sub['subscription_id'],
                    sub['user_id'],
                    sub['username'] or '',
                    full_name,
                    sub['plan_name'],
                    f"R$ {float(sub['price'] or 0):.2f}",
                    sub['duration_days'] if sub['duration_days'] != -1 else "Permanente",
                    sub['payment_method'].replace('_', ' ').title(),
                    sub['payment_status'].title(),
                    sub['start_date'].strftime('%d/%m/%Y %H:%M') if sub['start_date'] else '',
                    sub['end_date'].strftime('%d/%m/%Y %H:%M') if sub['end_date'] else 'Permanente',
                    sub['days_remaining'] if sub['days_remaining'] != 999999 else "∞",
                    sub['days_paid'] if sub['days_paid'] != 999999 else "∞",
                    sub['total_days'] if sub['total_days'] != 999999 else "∞",
                    sub['expiration_status'],
                    "SIM" if sub['is_permanent'] else "NÃO",
                    sub['created_at'].strftime('%d/%m/%Y %H:%M') if sub['created_at'] else ''
                ]
                ws_subs.append(row)
            
            # === ABA 3: EXPIRANDO EM BREVE ===
            ws_expiring = wb.create_sheet("⚠️ Expirando em Breve")
            
            if expiring_subs:
                ws_expiring.append([
                    "ID Usuário", "Username", "Nome", "Plano", "Dias Restantes",
                    "Data Expiração", "Status", "Valor (R$)"
                ])
                
                for sub in expiring_subs:
                    full_name = f"{sub['first_name'] or ''} {sub['last_name'] or ''}".strip()
                    
                    # Calcular dias restantes
                    end_date = sub['end_date']
                    if isinstance(end_date, str):
                        end_date = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
                    days_left = (end_date - datetime.now()).days
                    
                    ws_expiring.append([
                        sub['user_id'],
                        sub['username'] or '',
                        full_name,
                        sub['plan_name'],
                        days_left,
                        sub['end_date'].strftime('%d/%m/%Y %H:%M') if sub['end_date'] else '',
                        "Expirando" if days_left <= 3 else "Próximo de expirar",
                        f"R$ {float(sub['price'] or 0):.2f}"
                    ])
            else:
                ws_expiring.append(["Nenhuma assinatura expirando em breve!"])
            
            # === ABA 4: TODOS OS USUÁRIOS ===
            ws_users = wb.create_sheet("👥 Todos os Usuários")
            
            ws_users.append([
                "ID", "Username", "Nome", "Sobrenome", "Data de Entrada", "É VIP"
            ])
            
            for user in all_users:
                ws_users.append([
                    user['id'],
                    user['username'] or '',
                    user['first_name'] or '',
                    user['last_name'] or '',
                    user['joined_date'].strftime('%d/%m/%Y %H:%M') if user['joined_date'] else '',
                    'SIM' if user.get('is_vip', False) else 'NÃO'
                ])
            
            # Aplicar formatação e ajustar colunas
            for ws in [ws_summary, ws_subs, ws_expiring, ws_users]:
                # Ajustar largura das colunas
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Máximo de 50 caracteres
                    ws.column_dimensions[col_letter].width = adjusted_width
                
                # Formatação do cabeçalho
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Formatação especial para assinaturas expirando
            if ws_expiring.max_row > 1:
                for row in range(2, ws_expiring.max_row + 1):
                    days_cell = ws_expiring[f'E{row}']  # Coluna dias restantes
                    if days_cell.value and isinstance(days_cell.value, int):
                        if days_cell.value <= 1:
                            days_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        elif days_cell.value <= 3:
                            days_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # Salvar em memória
            file_stream = BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            
            await query.message.reply_document(
                document=file_stream,
                filename=f"relatorio_vip_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                caption="📊 Relatório completo de assinaturas VIP gerado com sucesso!\n\n"
                       "📋 Inclui:\n"
                       "• Resumo executivo com estatísticas\n"
                       "• Assinaturas detalhadas com dias pagos/restantes\n"
                       "• Lista de assinaturas expirando em breve\n"
                       "• Todos os usuários do sistema"
            )
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório Excel: {e}")
            await query.message.edit_text(f"❌ Erro ao gerar relatório: {str(e)}")
        
        return

    # Nova funcionalidade: Exportar apenas assinaturas expirando
    if query.data == "admin_export_expiring":
        await query.message.edit_text("⚠️ Gerando relatório de assinaturas expirando...")
        
        try:
            expiring_subs = get_expiring_subscriptions()
            
            if not expiring_subs:
                await query.message.edit_text("✅ Nenhuma assinatura expirando em breve!")
                return
            
            # Criar workbook simples para assinaturas expirando
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Assinaturas Expirando"
            
            # Cabeçalho
            ws.append(["RELATÓRIO DE ASSINATURAS EXPIRANDO"])
            ws.append(["Data da Exportação:", datetime.now().strftime('%d/%m/%Y %H:%M:%S')])
            ws.append(["Total de Assinaturas:", len(expiring_subs)])
            ws.append([""])
            
            # Cabeçalhos das colunas
            headers = [
                "ID Usuário", "Username", "Nome Completo", "Plano", 
                "Dias Restantes", "Data Expiração", "Valor (R$)", "Status"
            ]
            ws.append(headers)
            
            # Dados
            for sub in expiring_subs:
                full_name = f"{sub['first_name'] or ''} {sub['last_name'] or ''}".strip()
                
                # Calcular dias restantes
                end_date = sub['end_date']
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
                days_left = (end_date - datetime.now()).days
                
                # Determinar status
                if days_left <= 0:
                    status = "EXPIRADA"
                elif days_left == 1:
                    status = "EXPIRA HOJE"
                elif days_left == 2:
                    status = "EXPIRA AMANHÃ"
                else:
                    status = f"EXPIRA EM {days_left} DIAS"
                
                row = [
                    sub['user_id'],
                    sub['username'] or '',
                    full_name,
                    sub['plan_name'],
                    days_left,
                    sub['end_date'].strftime('%d/%m/%Y %H:%M') if sub['end_date'] else '',
                    f"R$ {float(sub['price'] or 0):.2f}",
                    status
                ]
                ws.append(row)
            
            # Formatação
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Formatação do cabeçalho
            if ws.max_row > 4:  # Mais que 4 linhas (título + data + total + vazio)
                for cell in ws[5]:  # Linha dos cabeçalhos
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Formatação condicional para urgência
            for row in range(6, ws.max_row + 1):  # A partir da linha 6 (dados)
                days_cell = ws[f'E{row}']  # Coluna dias restantes
                status_cell = ws[f'H{row}']  # Coluna status
                
                if days_cell.value is not None:
                    if days_cell.value <= 0:
                        # Vermelho para expiradas
                        days_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        status_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif days_cell.value == 1:
                        # Amarelo escuro para expira hoje
                        days_cell.fill = PatternFill(start_color="FFDD44", end_color="FFDD44", fill_type="solid")
                        status_cell.fill = PatternFill(start_color="FFDD44", end_color="FFDD44", fill_type="solid")
                    elif days_cell.value <= 3:
                        # Amarelo claro para expira em breve
                        days_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                        status_cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            
            # Salvar
            file_stream = BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)
            
            await query.message.reply_document(
                document=file_stream,
                filename=f"assinaturas_expirando_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                caption=f"⚠️ Relatório de assinaturas expirando!\n\n"
                       f"📊 Total: {len(expiring_subs)} assinaturas\n"
                       f"🔴 Expirando em ≤3 dias\n\n"
                       f"💡 Use este relatório para:\n"
                       f"• Enviar lembretes aos usuários\n"
                       f"• Planejar campanhas de renovação\n"
                       f"• Acompanhar receita em risco"
            )
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de expiração: {e}")
            await query.message.edit_text(f"❌ Erro ao gerar relatório: {str(e)}")
        
        return

async def handle_admin_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("[DEBUG] Entrou em handle_admin_edit")
    query = update.callback_query
    await query.answer()
    logger.info(f"Callback de edição recebido: {query.data}")
    logger.info(f"[DEBUG] Valor exato do query.data: '{query.data}'")
    
    # Carregar configurações iniciais
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        await query.message.reply_text("❌ Erro ao carregar as configurações. Tente novamente.")
        return
        
    logger.info(f"[DEBUG] query.data.startswith('admin_edit_plan_name_input_'): {query.data.startswith('admin_edit_plan_name_input_')}")

    if not is_admin(int(update.effective_user.id)):
        print("Usuário não é admin! Ignorando mensagem.")
        logger.info(f"Usuário {update.effective_user.id} tentou acessar sem permissão.")
        await query.message.reply_text("Acesso negado.")
        return

    print("Usuário é admin! Processando mensagem.")

    messages = load_messages_from_db()
    
    # Verificar se messages foi carregado corretamente
    if not messages:
        logger.error("Falha ao carregar mensagens do banco de dados na função handle_admin_edit")
        messages = {}  # Usar dicionário vazio como fallback

    
    # Bloco de tratamento do callback admin_edit_welcome_message
    if query.data == "admin_edit_welcome_message":
        logger.info("[DEBUG] Entrou no bloco admin_edit_welcome_message (handle_admin_edit)")
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            # Apagar a mensagem anterior
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de boas-vindas para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="👋 Editar Mensagem de Boas-vindas\n\n"
                     f"Mensagem atual:\n{messages.get('welcome_message', 'Não definida')}\n\n"
                     "Envie a nova mensagem de boas-vindas:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de boas-vindas enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de boas-vindas: {e}")
        context.user_data['editing'] = 'welcome_message'
        return
    elif query.data == "admin_edit_welcome_caption":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_welcome_file")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        welcome_file_config = config.get('welcome_file', {})
        current_caption = messages.get('welcome_message', 'sem mensagem definida!')

        await query.message.edit_text(
            "📝 Editar Legenda do Arquivo de Boas-vindas\n\n"
            f"Legenda atual: {current_caption}\n\n"
            "Envie a nova legenda para o arquivo de boas-vindas:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'welcome_caption'
        return
    elif query.data == "admin_edit_start_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de início para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="🏁 Editar Mensagem de Início\n\n"
                     f"Mensagem atual:\n{messages.get('start_message', 'Não definida')}\n\n"
                     "Envie a nova mensagem de início:",
                reply_markup=reply_markup,
                parse_mode='HTML'
            )
            context.user_data['editing_message_id'] = msg.message_id
            context.user_data['editing'] = 'start_message'
            logger.info("[DEBUG] Mensagem de edição de início enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de início: {e}")
        return
    elif query.data == "admin_edit_payment_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de pagamento para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="💎 Editar Mensagem de Pagamento\n\n"
                     f"Mensagem atual:\n{messages.get('payment_instructions', 'Não definida')}\n\n"
                     "Envie a nova mensagem de pagamento:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de pagamento enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de pagamento: {e}")
        context.user_data['editing'] = 'payment_instructions'
        return
    elif query.data == "admin_edit_success_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de sucesso para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="✅ Editar Mensagem de Sucesso\n\n"
                     f"Mensagem atual:\n{messages.get('payment_success', 'Não definida')}\n\n"
                     "Envie a nova mensagem de sucesso:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de sucesso enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de sucesso: {e}")
        context.user_data['editing'] = 'payment_success'
        return
    elif query.data == "admin_edit_error_message":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de erro para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="❌ Editar Mensagem de Erro\n\n"
                     f"Mensagem atual:\n{messages.get('payment_error', 'Não definida')}\n\n"
                     "Envie a nova mensagem de erro:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de erro enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de erro: {e}")
        context.user_data['editing'] = 'payment_error'
        return
    elif query.data == "admin_edit_pix_instructions":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        try:
            try:
                await query.message.delete()
                logger.info("[DEBUG] Mensagem anterior apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem anterior: {e}")
            logger.info("[DEBUG] Tentando enviar mensagem de edição de instruções PIX para o admin")
            msg = await context.bot.send_message(
                chat_id=update.effective_user.id,
                text="📝 Editar Instruções PIX\n\n"
                     f"Mensagem atual:\n{messages.get('pix_automatico_instructions', 'Não definida')}\n\n"
                     "Envie a nova mensagem de instruções PIX:",
                reply_markup=reply_markup
            )
            context.user_data['editing_message_id'] = msg.message_id
            logger.info("[DEBUG] Mensagem de edição de instruções PIX enviada com sucesso")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao enviar mensagem de edição de instruções PIX: {e}")
        context.user_data['editing'] = 'pix_automatico_instructions'
        return
        # Verifica se é uma edição específica do plano
   

    elif query.data.startswith("admin_edit_plan_duration_input_"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.message.reply_text("Plano não encontrado.")
            return
            
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data=f"admin_edit_plan_{plan_id}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            f"⏱️ Editar Duração do Plano\n\n"
            f"Duração atual: {plan['duration_days']} dias\n\n"
            "Envie a nova duração em dias (apenas números):",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = f"plan_duration_{plan_id}"
        
    elif query.data.startswith("admin_edit_plan_name_input_"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.message.reply_text("Plano não encontrado.")
            return
            
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data=f"admin_edit_plan_{plan_id}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            f"📝 Editar Nome do Plano\n\n"
            f"Nome atual: {plan['name']}\n\n"
            "Envie o novo nome do plano:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = f"plan_name_{plan_id}"
        
    elif query.data.startswith("admin_edit_plan_price_input_"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.message.reply_text("Plano não encontrado.")
            return
            
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data=f"admin_edit_plan_{plan_id}")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            f"💰 Editar Preço do Plano\n\n"
            f"Preço atual: R${plan['price']:.2f}\n\n"
            "Envie o novo preço (apenas números):",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = f"plan_price_{plan_id}"
        
    elif query.data.startswith("admin_edit_plan_") and not query.data.endswith("_input"):
        plan_id = int(query.data.split('_')[-1])
        plan = await get_plan_by_id(plan_id)
        if not plan:
            await query.message.reply_text("Plano não encontrado.")
            return
            
        keyboard = [
            [InlineKeyboardButton("📝 Nome", callback_data=f"admin_edit_plan_name_input_{plan_id}")],
            [InlineKeyboardButton("💰 Preço", callback_data=f"admin_edit_plan_price_input_{plan_id}")],
            [InlineKeyboardButton("⏱️ Duração (dias)", callback_data=f"admin_edit_plan_duration_input_{plan_id}")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_vip_plans")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            f"💎 Editar Plano: {plan['name']}\n\n"
            f"Preço atual: R${plan['price']:.2f}\n"
            f"Duração atual: {plan['duration_days']} dias\n\n"
            "Escolha o que deseja editar:",
            reply_markup=reply_markup
        )
        return
        
    
    elif query.data == "admin_edit_cnpay_key":
        logger.info("Editar API Key do CNPay")
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_config_cnpay")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🔑 Editar API Key do CNPay\n\n"
            f"API Key atual: {config.get('cnpay_api_key', 'Não configurada')}\n\n"
            "Envie a nova API Key:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'cnpay_api_key'
        return

    elif query.data == "admin_edit_cnpay_secret":
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_config_cnpay")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            "🔐 Editar API Secret do CNPay\n\n"
            f"API Secret atual: {'********' if config.get('cnpay_api_secret') else 'Não configurada'}\n\n"
            "Envie o novo API Secret:",
            reply_markup=reply_markup
        )
        context.user_data['editing'] = 'cnpay_api_secret'
        return

async def handle_admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info(f"[DEBUG] handle_admin_text chamado! Texto recebido: {getattr(update.message, 'text', None)} | context.user_data: {context.user_data}")
    
    # --- FLUXO DE AGENDAMENTO DE MENSAGENS ---
    if context.user_data.get('scheduling_step') == 'message_text':
        # Salvar texto da mensagem
        context.user_data['scheduled_message_text'] = update.message.text
        context.user_data['scheduling_step'] = 'target_type'
        
        keyboard = [
            [InlineKeyboardButton("👥 Todos os Usuários", callback_data="admin_schedule_target_all")],
            [InlineKeyboardButton("💎 Usuários VIP", callback_data="admin_schedule_target_vip")],
            [InlineKeyboardButton("👤 Usuários Específicos", callback_data="admin_schedule_target_specific")],
            [InlineKeyboardButton("❌ Cancelar", callback_data="admin_schedule_messages")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "✅ Mensagem salva!\n\n"
            "Agora escolha o tipo de destinatários:",
            reply_markup=reply_markup
        )
        return
    
    elif context.user_data.get('scheduling_step') == 'specific_users':
        # Processar lista de usuários específicos
        user_text = update.message.text.strip()
        try:
            # Tentar interpretar como lista de IDs separados por vírgula ou espaço
            user_ids = []
            for part in user_text.replace(',', ' ').split():
                if part.strip().isdigit():
                    user_ids.append(int(part.strip()))
            
            if not user_ids:
                await update.message.reply_text(
                    "❌ Nenhum ID válido encontrado. Digite os IDs dos usuários separados por vírgula ou espaço:\n\n"
                    "Exemplo: 123456789, 987654321, 555666777"
                )
                return
            
            context.user_data['scheduled_target_users'] = user_ids
            context.user_data['scheduling_step'] = 'scheduled_date'
            
            # Gerar exemplo com data atual + 30 minutos
            example_time = datetime.now() + timedelta(minutes=30)
            example_str = example_time.strftime("%d/%m/%Y %H:%M")
            
            await update.message.reply_text(
                f"✅ {len(user_ids)} usuários selecionados!\n\n"
                "Agora digite a data e hora para envio da mensagem:\n\n"
                "Formato: DD/MM/AAAA HH:MM\n"
                f"Exemplo: {example_str}"
            )
        except Exception as e:
            await update.message.reply_text(
                "❌ Erro ao processar IDs. Digite os IDs dos usuários separados por vírgula ou espaço:\n\n"
                "Exemplo: 123456789, 987654321, 555666777"
            )
        return
    
    elif context.user_data.get('scheduling_step') == 'scheduled_date':
        # Processar data e hora agendada
        date_text = update.message.text.strip()
        now = datetime.now()  # Definir no início para evitar erro no except
        try:
            # Tentar diferentes formatos de data
            scheduled_date = None
            for fmt in ['%d/%m/%Y %H:%M', '%d-%m-%Y %H:%M', '%d/%m/%Y %H:%M:%S']:
                try:
                    scheduled_date = datetime.strptime(date_text, fmt)
                    break
                except ValueError:
                    continue
            
            if not scheduled_date:
                # Gerar exemplo com data atual + 30 minutos
                example_time = datetime.now() + timedelta(minutes=30)
                example_str = example_time.strftime("%d/%m/%Y %H:%M")
                
                await update.message.reply_text(
                    "❌ Formato de data inválido. Use o formato:\n\n"
                    "DD/MM/AAAA HH:MM\n"
                    f"Exemplo: {example_str}"
                )
                return
            
            # Verificar se a data não é no passado
            if scheduled_date < now:
                # Gerar exemplo com data atual + 1 minuto
                example_time = now + timedelta(minutes=1)
                example_str = example_time.strftime("%d/%m/%Y %H:%M")
                
                await update.message.reply_text(
                    "❌ A data deve ser no futuro.\n\n"
                    "Digite uma data válida:\n\n"
                    "DD/MM/AAAA HH:MM\n"
                    f"Exemplo: {example_str}"
                )
                return
            
            # Verificar se já existe mensagem agendada para o mesmo tipo de destinatários
            message_text = context.user_data['scheduled_message_text']
            target_type = context.user_data.get('scheduled_target_type', 'all_users')
            target_users = context.user_data.get('scheduled_target_users')
            created_by = update.effective_user.id
            
            if check_duplicate_scheduled_message(target_type, target_users):
                target_text = {
                    'all_users': 'todos os usuários',
                    'vip_users': 'usuários VIP',
                    'specific_users': f"{len(target_users) if target_users else 0} usuários específicos"
                }.get(target_type, target_type)
                
                await update.message.reply_text(
                    f"⚠️ Já existe uma mensagem agendada pendente para {target_text}.\n\n"
                    f"📋 Verifique as mensagens pendentes no menu de agendamento ou aguarde o envio da mensagem atual."
                )
                return
            
            # Criar a mensagem agendada
            if create_scheduled_message(message_text, scheduled_date, target_type, target_users, created_by):
                scheduled_time = scheduled_date.strftime('%d/%m/%Y %H:%M')
                target_text = {
                    'all_users': 'todos os usuários',
                    'vip_users': 'usuários VIP',
                    'specific_users': f"{len(target_users) if target_users else 0} usuários específicos"
                }.get(target_type, target_type)
                
                await update.message.reply_text(
                    f"✅ Mensagem agendada com sucesso!\n\n"
                    f"📅 Data/Hora: {scheduled_time}\n"
                    f"👥 Destinatários: {target_text}\n"
                    f"📝 Mensagem: {message_text[:100]}{'...' if len(message_text) > 100 else ''}\n\n"
                    f"A mensagem será enviada automaticamente no horário agendado."
                )
                
                # Limpar dados do contexto
                for key in ['scheduling_step', 'scheduled_message_text', 'scheduled_target_type', 'scheduled_target_users']:
                    context.user_data.pop(key, None)
            else:
                await update.message.reply_text(
                    "❌ Erro ao agendar mensagem. Tente novamente."
                )
        except Exception as e:
            logger.error(f"Erro ao processar data agendada: {e}")
            # Gerar exemplo com data atual
            example_time = now
            example_str = example_time.strftime("%d/%m/%Y %H:%M")
            
            await update.message.reply_text(
                "❌ Erro ao processar data. Use o formato:\n\n"
                "DD/MM/AAAA HH:MM\n"
                f"Exemplo: {example_str}"
            )
        return
    
    # --- FLUXO DE ADIÇÃO DE ADMIN ---
    if context.user_data.get("waiting_for_admin_id"):
        logger.info(f"[DEBUG] Entrou no bloco waiting_for_admin_id. Texto: {getattr(update.message, 'text', None)}")
        admin_id = update.message.text.strip()
        if not admin_id.isdigit():
            await update.message.reply_text("❌ O ID deve conter apenas números. Tente novamente:")
            return
        context.user_data["pending_admin_id"] = admin_id
        context.user_data["waiting_for_admin_id"] = False
        context.user_data["waiting_for_admin_username"] = True
        await update.message.reply_text("Agora envie o username do novo admin (sem o @):")
        return

    if context.user_data.get("waiting_for_admin_username"):
        logger.info(f"[DEBUG] Entrou no bloco waiting_for_admin_username. Texto: {getattr(update.message, 'text', None)}")
        username = update.message.text.strip().lstrip("@")
        admin_id = context.user_data.get("pending_admin_id")
        if not username:
            await update.message.reply_text("❌ O username não pode estar vazio. Tente novamente:")
            return
        # Salvar admin no banco
        try:
            add_admin(admin_id, update.effective_user.id, username=username)
            await update.message.reply_text(f"✅ Novo admin adicionado com sucesso!\nID: {admin_id}\nUsername: @{username}")
            logger.info(f"[DEBUG] Admin adicionado com sucesso: {admin_id} @{username}")
        except Exception as e:
            logger.error(f"[DEBUG] Erro ao adicionar admin: {e}")
            await update.message.reply_text(f"❌ Erro ao adicionar admin: {e}")
        # Limpar flags
        context.user_data.pop("waiting_for_admin_username", None)
        context.user_data.pop("pending_admin_id", None)
        # Voltar para o menu de admins
        await handle_admin_callback(update, context)
        return

    if (
        'editing' not in context.user_data and 
        'broadcast_type' not in context.user_data and 
        'waiting_for_broadcast_text' not in context.user_data and
        'adding_group' not in context.user_data and
        'adding_plan' not in context.user_data and
        'waiting_for_admin_id' not in context.user_data and
        'waiting_for_admin_username' not in context.user_data
    ):
        return
    logger.info(f"Texto recebido para edição: {update.message.text}")

    if not is_admin(int(update.effective_user.id)):
        logger.error("[handle_admin_text] Usuário não é admin! Ignorando mensagem.")
        return

    if context.user_data.get("adding_admin"):
        # Solicita o ID do novo admin
        await update.message.reply_text("Por favor, envie o ID do Telegram do novo admin (apenas números):")
        context.user_data["adding_admin"] = False
        context.user_data["waiting_for_admin_id"] = True
        return
    
    # --- NOVO FLUXO AJUSTADO ---
    # Se está esperando escolha do botão
    if context.user_data.get('waiting_for_button_choice', False):
        escolha = update.message.text.strip().lower()
        if escolha in ['sim', 's', 'yes', 'y']:
            context.user_data['waiting_for_button_choice'] = False
            context.user_data['waiting_for_button_text'] = True
            await update.message.reply_text('Digite o texto do botão:')
            return
        elif escolha in ['não', 'nao', 'n', 'no']:
            context.user_data['waiting_for_button_choice'] = False
            context.user_data['button_text'] = None
            context.user_data['button_url'] = None
            # Enviar broadcast usando a mensagem já salva
            await enviar_broadcast(update, context)
            return
        else:
            await update.message.reply_text('Por favor, responda "Sim" ou "Não". Deseja adicionar um botão de redirecionamento?')
            return
    # Se está esperando texto do botão
    if context.user_data.get('waiting_for_button_text', False):
        context.user_data['button_text'] = update.message.text.strip()
        context.user_data['waiting_for_button_text'] = False
        context.user_data['waiting_for_button_url'] = True
        await update.message.reply_text('Agora envie o link do botão (começando com https://):')
        return
    # Se está esperando link do botão
    if context.user_data.get('waiting_for_button_url', False):
        url = update.message.text.strip()
        if not is_valid_url(url):
            await update.message.reply_text('O link deve ser válido e começar com http:// ou https://. Tente novamente:')
            return
        context.user_data['button_url'] = url
        context.user_data['waiting_for_button_url'] = False
        # Enviar broadcast usando a mensagem já salva
        await enviar_broadcast(update, context)
        return
    # --- FIM NOVO FLUXO ---

    # Se for broadcast, salvar a mensagem inicial e perguntar sobre o botão
    if 'broadcast_type' in context.user_data and not context.user_data.get('waiting_for_button_choice', False) and not context.user_data.get('waiting_for_button_text', False) and not context.user_data.get('waiting_for_button_url', False):
        context.user_data['broadcast_message_text'] = update.message.text
        context.user_data['waiting_for_button_choice'] = True
        await update.message.reply_text('Deseja adicionar um botão de redirecionamento? (Sim/Não)')
        return

    # Fluxo antigo para edição de mensagens/configs
    if 'editing' in context.user_data:
        editing_type = context.user_data['editing']
        
        # Edição de preço de plano
        if editing_type.startswith('plan_price_'):
            try:
                plan_id = int(editing_type.split('_')[-1])
                new_price = float(update.message.text.replace(',', '.'))
                
                db = Database()
                try:
                    db.connect()
                    db.execute_query(
                        "UPDATE vip_plans SET price = %s WHERE id = %s",
                        (new_price, plan_id),
                        commit=True
                    )
                finally:
                    db.close()
                
                await update.message.reply_text(f"✅ Preço do plano atualizado para R${new_price:.2f}!")
                del context.user_data['editing']
                
                # Voltar para o menu de edição do plano
                plan = await get_plan_by_id(plan_id)
                if plan:
                    keyboard = [
                        [InlineKeyboardButton("📝 Nome", callback_data=f"admin_edit_plan_name_input_{plan_id}")],
                        [InlineKeyboardButton("💰 Preço", callback_data=f"admin_edit_plan_price_input_{plan_id}")],
                        [InlineKeyboardButton("⏱️ Duração (dias)", callback_data=f"admin_edit_plan_duration_input_{plan_id}")],
                        [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_vip_plans")]
                    ]
                    reply_markup = InlineKeyboardMarkup(keyboard)
                    await update.message.reply_text(
                        f"💎 Editar Plano: {plan['name']}\n\n"
                        f"Preço atual: R${plan['price']:.2f}\n"
                        f"Duração atual: {plan['duration_days']} dias\n\n"
                        "Escolha o que deseja editar:",
                        reply_markup=reply_markup
                    )
                return
            except ValueError:
                await update.message.reply_text("❌ Preço inválido! Digite apenas números (ex: 49.90)")
                return
            except Exception as e:
                logger.error(f"Erro ao atualizar preço do plano: {e}")
                await update.message.reply_text("❌ Erro ao atualizar o preço do plano!")
                del context.user_data['editing']
                return
        
        # Edição de nome de plano
        elif editing_type.startswith('plan_name_'):
            try:
                plan_id = int(editing_type.split('_')[-1])
                new_name = update.message.text.strip()
                
                if not new_name:
                    await update.message.reply_text("❌ Nome não pode estar vazio!")
                    return
                
                db = Database()
                try:
                    db.connect()
                    db.execute_query(
                        "UPDATE vip_plans SET name = %s WHERE id = %s",
                        (new_name, plan_id),
                        commit=True
                    )
                finally:
                    db.close()
                
                await update.message.reply_text(f"✅ Nome do plano atualizado para '{new_name}'!")
                del context.user_data['editing']
                
                # Voltar para o menu de edição do plano
                plan = await get_plan_by_id(plan_id)
                if plan:
                    keyboard = [
                        [InlineKeyboardButton("📝 Nome", callback_data=f"admin_edit_plan_name_input_{plan_id}")],
                        [InlineKeyboardButton("💰 Preço", callback_data=f"admin_edit_plan_price_input_{plan_id}")],
                        [InlineKeyboardButton("⏱️ Duração (dias)", callback_data=f"admin_edit_plan_duration_input_{plan_id}")],
                        [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_vip_plans")]
                    ]
                    reply_markup = InlineKeyboardMarkup(keyboard)
                    await update.message.reply_text(
                        f"💎 Editar Plano: {plan['name']}\n\n"
                        f"Preço atual: R${plan['price']:.2f}\n"
                        f"Duração atual: {plan['duration_days']} dias\n\n"
                        "Escolha o que deseja editar:",
                        reply_markup=reply_markup
                    )
                return
            except Exception as e:
                logger.error(f"Erro ao atualizar nome do plano: {e}")
                await update.message.reply_text("❌ Erro ao atualizar o nome do plano!")
                del context.user_data['editing']
                return
        
        # Edição de duração de plano
        elif editing_type.startswith('plan_duration_'):
            try:
                plan_id = int(editing_type.split('_')[-1])
                new_duration = int(update.message.text.strip())
                
                db = Database()
                try:
                    db.connect()
                    db.execute_query(
                        "UPDATE vip_plans SET duration_days = %s WHERE id = %s",
                        (new_duration, plan_id),
                        commit=True
                    )
                finally:
                    db.close()
                
                duration_text = "Permanente" if new_duration == -1 else f"{new_duration} dias"
                await update.message.reply_text(f"✅ Duração do plano atualizada para {duration_text}!")
                del context.user_data['editing']
                
                # Voltar para o menu de edição do plano
                plan = await get_plan_by_id(plan_id)
                if plan:
                    keyboard = [
                        [InlineKeyboardButton("📝 Nome", callback_data=f"admin_edit_plan_name_input_{plan_id}")],
                        [InlineKeyboardButton("💰 Preço", callback_data=f"admin_edit_plan_price_input_{plan_id}")],
                        [InlineKeyboardButton("⏱️ Duração (dias)", callback_data=f"admin_edit_plan_duration_input_{plan_id}")],
                        [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_vip_plans")]
                    ]
                    reply_markup = InlineKeyboardMarkup(keyboard)
                    await update.message.reply_text(
                        f"💎 Editar Plano: {plan['name']}\n\n"
                        f"Preço atual: R${plan['price']:.2f}\n"
                        f"Duração atual: {plan['duration_days']} dias\n\n"
                        "Escolha o que deseja editar:",
                        reply_markup=reply_markup
                    )
                return
            except ValueError:
                await update.message.reply_text("❌ Duração inválida! Digite apenas números.")
                return
            except Exception as e:
                logger.error(f"Erro ao atualizar duração do plano: {e}")
                await update.message.reply_text("❌ Erro ao atualizar a duração do plano!")
                del context.user_data['editing']
                return


    # Adicionar fluxo de adicionar novo plano
    if 'adding_plan' in context.user_data:
        step = context.user_data['adding_plan']['step']
        if step == 'name':
            context.user_data['adding_plan']['name'] = update.message.text.strip()
            context.user_data['adding_plan']['step'] = 'price'
            await update.message.reply_text("Digite o preço do plano (apenas números, ex: 49.90):")
            return
        elif step == 'price':
            try:
                price = float(update.message.text.replace(',', '.'))
                context.user_data['adding_plan']['price'] = price
                context.user_data['adding_plan']['step'] = 'duration'
                await update.message.reply_text("Digite a duração do plano em dias (ou -1 para permanente):")
            except Exception:
                await update.message.reply_text("Preço inválido. Digite apenas números, ex: 49.90")
            return
        elif step == 'duration':
            try:
                duration = int(update.message.text.strip())
                context.user_data['adding_plan']['duration'] = duration
                context.user_data['adding_plan']['step'] = 'description'
                await update.message.reply_text("Digite a descrição do plano:")
            except Exception:
                await update.message.reply_text("Duração inválida. Digite apenas números.")
            return
        elif step == 'description':
            description = update.message.text.strip()
            context.user_data['adding_plan']['description'] = description
            context.user_data['adding_plan']['step'] = 'create_group'
            
            await update.message.reply_text(
                "Deseja criar um novo grupo VIP para este plano?\n\n"
                "Digite 'sim' para criar um grupo ou 'não' para continuar sem grupo:"
            )
            return
        elif step == 'create_group':
            create_group = update.message.text.strip().lower()
            if create_group in ['sim', 's', 'yes', 'y']:
                context.user_data['adding_plan']['step'] = 'group_name'
                await update.message.reply_text("Digite o nome do novo grupo VIP:")
                return
            elif create_group in ['não', 'nao', 'n', 'no']:
                # Pular criação de grupo e ir direto para finalizar
                data = context.user_data['adding_plan']
                db = Database()
                try:
                    db.connect()
                    db.execute_query(
                        "INSERT INTO vip_plans (name, price, duration_days, description, is_active) VALUES (%s, %s, %s, %s, 1)",
                        (data['name'], data['price'], data['duration'], data['description']),
                        commit=True
                    )
                finally:
                    db.close()
                
                await update.message.reply_text(f"Plano '{data['name']}' adicionado com sucesso!")
                del context.user_data['adding_plan']
                await handle_admin_callback(update, context)
                return
            else:
                await update.message.reply_text("Resposta inválida. Digite 'sim' ou 'não':")
                return
        elif step == 'group_name':
            group_name = update.message.text.strip()
            context.user_data['adding_plan']['group_name'] = group_name
            context.user_data['adding_plan']['step'] = 'group_id'
            await update.message.reply_text(
                "Digite o ID do grupo no Telegram (número negativo, ex: -1001234567890):\n\n"
                "Para obter o ID do grupo:\n"
                "1. Adicione o bot ao grupo\n"
                "2. Envie /id no grupo\n"
                "3. Copie o ID que aparecer"
            )
            return
        elif step == 'group_id':
            try:
                group_id = int(update.message.text.strip())
                if group_id >= 0:
                    await update.message.reply_text("ID do grupo deve ser negativo. Tente novamente:")
                    return
                
                context.user_data['adding_plan']['group_id'] = group_id
                data = context.user_data['adding_plan']
                
                # Inserir o plano e o grupo
                db = Database()
                try:
                    db.connect()
                    
                    # Inserir o plano
                    db.execute_query(
                        "INSERT INTO vip_plans (name, price, duration_days, description, is_active) VALUES (%s, %s, %s, %s, 1)",
                        (data['name'], data['price'], data['duration'], data['description']),
                        commit=True
                    )
                    
                    # Pegar o ID do plano inserido
                    plan_result = db.execute_fetch_one("SELECT LAST_INSERT_ID() as plan_id")
                    plan_id = plan_result['plan_id']
                    
                    # Inserir o grupo VIP
                    db.execute_query(
                        "INSERT INTO vip_groups (group_id, group_name, is_active) VALUES (%s, %s, TRUE)",
                        (group_id, data['group_name']),
                        commit=True
                    )
                    
                    # Pegar o ID do grupo inserido
                    group_result = db.execute_fetch_one("SELECT LAST_INSERT_ID() as group_id")
                    vip_group_id = group_result['group_id']
                    
                    # Associar o plano ao grupo
                    db.execute_query(
                        "INSERT INTO plan_groups (plan_id, group_id) VALUES (%s, %s)",
                        (plan_id, vip_group_id),
                        commit=True
                    )
                    
                finally:
                    db.close()
                
                await update.message.reply_text(
                    f"Plano '{data['name']}' e grupo '{data['group_name']}' criados com sucesso!\n\n"
                    f"ID do grupo: {group_id}\n"
                    f"Nome do grupo: {data['group_name']}"
                )
                del context.user_data['adding_plan']
                await handle_admin_callback(update, context)
                return
                
            except ValueError:
                await update.message.reply_text("ID inválido. Digite apenas números (ex: -1001234567890):")
                return

    # Adicionar fluxo de edição de mensagem de boas-vindas
    if 'editing' in context.user_data:
        editing_type = context.user_data['editing']
        novo_texto = update.message.text.strip()
        logger.info(f"[DEBUG] Tipo de edição recebido: {editing_type}")
        keyboard = [[InlineKeyboardButton("⬅️ Voltar", callback_data="admin_messages")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        # Apagar a mensagem de edição (se existir)
        if 'editing_message_id' in context.user_data:
            try:
                await context.bot.delete_message(
                    chat_id=update.effective_user.id,
                    message_id=context.user_data['editing_message_id']
                )
                logger.info("[DEBUG] Mensagem de edição apagada com sucesso")
            except Exception as e:
                logger.warning(f"[DEBUG] Não foi possível apagar a mensagem de edição: {e}")
            del context.user_data['editing_message_id']
        
        # Salvar e confirmar a edição
        if editing_type == 'start_message':
            if save_message_to_db('start_message', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Mensagem de boas-vindas atualizada com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar a mensagem de boas-vindas!")
            del context.user_data['editing']
            return
        elif editing_type == 'payment_instructions':
            if save_message_to_db('payment_instructions', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Mensagem de pagamento atualizada com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar a mensagem de pagamento!")
            del context.user_data['editing']
            return
        elif editing_type == 'payment_success':
            if save_message_to_db('payment_success', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Mensagem de sucesso atualizada com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar a mensagem de sucesso!")
            del context.user_data['editing']
            return
        elif editing_type == 'payment_error':
            if save_message_to_db('payment_error', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Mensagem de erro atualizada com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar a mensagem de erro!")
            del context.user_data['editing']
            return
        elif editing_type == 'pix_automatico_instructions':
            if save_message_to_db('pix_automatico_instructions', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Instruções PIX atualizadas com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar as instruções PIX!")
            del context.user_data['editing']
            return

        elif editing_type == 'welcome_caption':
            if save_message_to_db('welcome_message', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Legenda do arquivo de boas-vindas atualizada com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar a legenda do arquivo de boas-vindas!")
            del context.user_data['editing']
            return
        elif editing_type == 'welcome_message':
            if save_message_to_db('welcome_message', novo_texto):
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text="✅ Mensagem de boas-vindas atualizada com sucesso!",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar a mensagem de boas-vindas!")
            del context.user_data['editing']
            return
        elif editing_type == "cnpay_api_key":
            if save_config_to_db('cnpay_api_key', novo_texto):
                success_message = "✅ API Key do CNPay atualizada com sucesso!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=success_message,
                    reply_markup=reply_markup
                )
            else:
                error_message = "❌ Erro ao atualizar a API Key do CNPay!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=error_message,
                    reply_markup=reply_markup
                )
        elif editing_type == "cnpay_api_secret":
            if save_config_to_db('cnpay_api_secret', novo_texto):
                success_message = "✅ API Secret do CNPay atualizada com sucesso!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=success_message,
                    reply_markup=reply_markup
                )
            else:
                error_message = "❌ Erro ao atualizar a API Secret do CNPay!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=error_message,
                    reply_markup=reply_markup
                )
        elif editing_type == "cnpay_environment":
            if save_config_to_db('cnpay_environment', novo_texto):
                success_message = "✅ Ambiente do CNPay atualizado com sucesso!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=success_message,
                    reply_markup=reply_markup
                )
            else:
                error_message = "❌ Erro ao atualizar o ambiente do CNPay!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=error_message,
                    reply_markup=reply_markup
                )
        elif editing_type == "cnpay_webhook_url":
            if save_config_to_db('cnpay_webhook_url', novo_texto):
                success_message = "✅ Webhook URL do CNPay atualizada com sucesso!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=success_message,
                    reply_markup=reply_markup
                )
            else:
                error_message = "❌ Erro ao atualizar o Webhook URL do CNPay!"
                await context.bot.send_message(
                    chat_id=update.effective_user.id,
                    text=error_message,
                    reply_markup=reply_markup
                )

        
# Função auxiliar para enviar o broadcast usando os dados do contexto
async def enviar_broadcast(update, context):
    broadcast_type = context.user_data.get('broadcast_type')
    message_text = context.user_data.get('broadcast_message_text', '')
    button_text = context.user_data.get('button_text')
    button_url = context.user_data.get('button_url')
    try:
        all_users = get_all_users()
        vip_users = get_vip_users()
        if broadcast_type in ['all', 'video_all', 'videonote_all']:
            recipients = [user['id'] for user in all_users]
        else:
            recipients = [user['id'] for user in vip_users]
        is_video_broadcast = broadcast_type.startswith('video_') or broadcast_type.startswith('videonote_')
        success_count = 0
        error_count = 0
        if is_video_broadcast and 'broadcast_video' in context.user_data:
            video_info = context.user_data['broadcast_video']
            video_file_id = video_info['file_id']
            is_videonote = video_info.get('is_videonote', False)
            video_type_text = "vídeo circular" if is_videonote else "vídeo"
            progress_message = await update.message.reply_text(
                f"📹 Enviando {video_type_text} + mensagem para {len(recipients)} usuários...\n"
                f"✅ Enviados: 0\n"
                f"❌ Erros: 0"
            )
            for user_id in recipients:
                try:
                    if is_videonote:
                        await context.bot.send_video_note(
                            chat_id=user_id,
                            video_note=video_file_id
                        )
                        if message_text.strip() or button_text:
                            if button_text and button_url:
                                reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton(button_text, url=button_url)]])
                                await context.bot.send_message(
                                    chat_id=user_id,
                                    text=message_text if message_text.strip() else button_text,
                                    reply_markup=reply_markup
                                )
                            else:
                                await context.bot.send_message(
                                    chat_id=user_id,
                                    text=message_text
                                )
                    else:
                        if button_text and button_url:
                            reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton(button_text, url=button_url)]])
                            await context.bot.send_video(
                                chat_id=user_id,
                                video=video_file_id,
                                caption=message_text,
                                reply_markup=reply_markup
                            )
                        else:
                            await context.bot.send_video(
                                chat_id=user_id,
                                video=video_file_id,
                                caption=message_text
                            )
                    success_count += 1
                except Exception as e:
                    logger.error(f"   ❌ Erro ao enviar {video_type_text} para {user_id}: {e}")
                    error_count += 1
                if (success_count + error_count) % 10 == 0:
                    await progress_message.edit_text(
                        f"📹 Enviando {video_type_text} + mensagem para {len(recipients)} usuários...\n"
                        f"✅ Enviados: {success_count}\n"
                        f"❌ Erros: {error_count}"
                    )
            await progress_message.edit_text(
                f"📹 Broadcast com {video_type_text} concluído!\n\n"
                f"✅ {video_type_text.title()}s enviados: {success_count}\n"
                f"❌ Erros: {error_count}\n\n"
                f"Tipo: {'Todos os usuários' if broadcast_type.endswith('_all') else 'Usuários VIP'}"
            )
            del context.user_data['broadcast_type']
            del context.user_data['broadcast_video']
            if 'waiting_for_broadcast_text' in context.user_data:
                del context.user_data['waiting_for_broadcast_text']
        else:
            progress_message = await update.message.reply_text(
                f"📢 Enviando mensagem para {len(recipients)} usuários...\n"
                f"✅ Enviados: 0\n"
                f"❌ Erros: 0"
            )
            for user_id in recipients:
                try:
                    if button_text and button_url:
                        reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton(button_text, url=button_url)]])
                        await context.bot.send_message(
                            chat_id=user_id,
                            text=message_text,
                            reply_markup=reply_markup
                        )
                    else:
                        await context.bot.send_message(
                            chat_id=user_id,
                            text=message_text
                        )
                    success_count += 1
                except Exception as e:
                    logger.error(f"Erro ao enviar mensagem para {user_id}: {e}")
                    error_count += 1
                if (success_count + error_count) % 10 == 0:
                    await progress_message.edit_text(
                        f"📢 Enviando mensagem para {len(recipients)} usuários...\n"
                        f"✅ Enviados: {success_count}\n"
                        f"❌ Erros: {error_count}"
                    )
            await progress_message.edit_text(
                f"📢 Broadcast concluído!\n\n"
                f"✅ Mensagens enviadas: {success_count}\n"
                f"❌ Erros: {error_count}\n\n"
                f"Tipo: {'Todos os usuários' if broadcast_type == 'all' else 'Usuários VIP'}"
            )
            del context.user_data['broadcast_type']
            if 'waiting_for_broadcast_text' in context.user_data:
                del context.user_data['waiting_for_broadcast_text']
        # Voltar ao menu de broadcast
        keyboard = [
            [InlineKeyboardButton("📢 Enviar para Todos", callback_data="admin_broadcast_all")],
            [InlineKeyboardButton("👥 Enviar para VIPs", callback_data="admin_broadcast_vip")],
            [InlineKeyboardButton("📹 Enviar Vídeo para Todos", callback_data="admin_broadcast_video_all")],
            [InlineKeyboardButton("📹 Enviar Vídeo para VIPs", callback_data="admin_broadcast_video_vip")],
            [InlineKeyboardButton("⭕ Enviar Vídeo Circular para Todos", callback_data="admin_broadcast_videonote_all")],
            [InlineKeyboardButton("⭕ Enviar Vídeo Circular para VIPs", callback_data="admin_broadcast_videonote_vip")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(
            "📢 Broadcast\n\nEscolha o tipo de broadcast:\n\n"
            "📹 Vídeo Normal: Formato retangular tradicional\n"
            "⭕ Vídeo Circular: Formato circular (video_note)",
            reply_markup=reply_markup
        )
    except Exception as e:
        logger.error(f"Erro ao realizar broadcast: {e}")
        await update.message.reply_text(
            f"❌ Erro ao realizar broadcast: {str(e)}\n\n"
            "Tente novamente mais tarde."
        )
        if 'broadcast_type' in context.user_data:
            del context.user_data['broadcast_type']
        if 'broadcast_video' in context.user_data:
            del context.user_data['broadcast_video']
        if 'waiting_for_broadcast_text' in context.user_data:
            del context.user_data['waiting_for_broadcast_text']

def safe_cleanup(temp_dir, max_attempts=3, delay=1):
    for attempt in range(max_attempts):
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                return True
        except Exception as e:
            if attempt == max_attempts - 1:
                logger.error(f"Falha ao limpar {temp_dir}: {e}")
                return False
            time.sleep(delay)
            
async def handle_admin_files(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler para receber arquivos (fotos e vídeos) do admin"""
    
    user_id = str(update.effective_user.id)
    messages = load_messages_from_db()
    
    # Verificar se messages foi carregado corretamente
    if not messages:
        logger.error("Falha ao carregar mensagens do banco de dados")
        messages = {}  # Usar dicionário vazio como fallback

    if not is_admin(user_id):
        logger.warning(f"Usuário não autorizado tentou acessar: {user_id}")
        await update.message.reply_text("🚫 Você não tem permissão para acessar o painel administrativo.")
        return
    
    # Carregar configurações
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        await update.message.reply_text("❌ Erro ao carregar configurações. Tente novamente.")
        return
    
    # Verificar se está aguardando arquivo de boas-vindas
    if context.user_data.get('waiting_for_welcome_file', False):
        try:
            file_id = None
            file_type = None
            
            # Verificar se é foto
            if update.message.photo:
                file_id = update.message.photo[-1].file_id
                file_type = 'photo'
            # Verificar se é vídeo
            elif update.message.video:
                file_id = update.message.video.file_id
                file_type = 'video'
            else:
                await update.message.reply_text("❌ Por favor, envie uma foto ou vídeo.")
                return
            
            # Obter caption com fallback seguro
            welcome_caption = messages.get('welcome_message', 'Bem-vindo!')
            if not welcome_caption or welcome_caption == 'sem mensagem definida!':
                welcome_caption = 'Bem-vindo!'
            
            # Salvar arquivo na configuração
            if 'welcome_file' not in config:
                config['welcome_file'] = {
                    'enabled': True,
                    'file_id': file_id,
                    'file_type': file_type,
                    'caption': welcome_caption
                }
            else:
                config['welcome_file']['file_id'] = file_id
                config['welcome_file']['file_type'] = file_type
                config['welcome_file']['enabled'] = True
                # Manter caption existente se não houver nova
                if 'caption' not in config['welcome_file']:
                    config['welcome_file']['caption'] = welcome_caption
            
            if save_config(config):
                # Limpar estado
                del context.user_data['waiting_for_welcome_file']
                
                # Confirmar sucesso
                keyboard = [[InlineKeyboardButton("⬅️ Voltar ao Menu", callback_data="admin_welcome_file")]]
                reply_markup = InlineKeyboardMarkup(keyboard)
                
                current_caption = config['welcome_file'].get('caption', welcome_caption)
                
                await update.message.reply_text(
                    f"✅ Arquivo de boas-vindas configurado com sucesso!\n\n"
                    f"📁 Tipo: {file_type.title()}\n"
                    f"📝 Legenda atual: {current_caption}\n\n"
                    f"O arquivo será enviado para novos usuários que usarem /start",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text("❌ Erro ao salvar arquivo. Tente novamente.")
                
        except Exception as e:
            logger.error(f"Erro ao processar arquivo de boas-vindas: {e}")
            await update.message.reply_text("❌ Erro ao processar arquivo. Tente novamente.")
            # Limpar estado em caso de erro
            if 'waiting_for_welcome_file' in context.user_data:
                del context.user_data['waiting_for_welcome_file']
    
    # Verificar se está aguardando vídeo para broadcast
    elif context.user_data.get('broadcast_type', '').startswith('video_') or context.user_data.get('broadcast_type', '').startswith('videonote_'):
        try:
            # Verificar se é vídeo
            if not update.message.video:
                await update.message.reply_text("❌ Por favor, envie um vídeo.")
                return
            
            video_file_id = update.message.video.file_id
            video_duration = update.message.video.duration
            video_size = update.message.video.file_size
            video_width = update.message.video.width
            video_height = update.message.video.height
            
            broadcast_type = context.user_data['broadcast_type']
            is_videonote = broadcast_type.startswith('videonote_')
            
            # Validações específicas para video_note
            if is_videonote:
                # Se não for quadrado, processa automaticamente
                if video_width != video_height:
                    try:
                        await update.message.reply_text(
                            f"⏳ Processando vídeo para formato quadrado...\n\n"
                            f"📏 Dimensões atuais: {video_width}x{video_height}\n"
                            f"🔄 Recortando automaticamente..."
                        )
                        
                        # Verificar se o módulo de processamento está disponível
                        if not VIDEO_PROCESSOR_AVAILABLE:
                            await update.message.reply_text(
                                f"❌ Módulo de processamento de vídeo não disponível.\n\n"
                                f"📏 Dimensões atuais: {video_width}x{video_height}\n"
                                f"📋 Requisito: Largura = Altura (ex: 240x240)"
                            )
                            return
                        
                        # Baixar o vídeo para arquivo temporário
                        video_file = await update.message.video.get_file()
                        temp_dir = tempfile.mkdtemp()
                        input_path = os.path.join(temp_dir, "input.mp4")
                        
                        logger.info(f"📥 Baixando vídeo para: {input_path}")
                        await video_file.download_to_drive(input_path)
                        
                        # Verificar se o arquivo foi baixado
                        if not os.path.exists(input_path):
                            await update.message.reply_text("❌ Erro ao baixar vídeo.")
                            shutil.rmtree(temp_dir)
                            return
                        
                        logger.info(f"🎬 Iniciando processamento do vídeo: {input_path}")
                        
                        # Processar para quadrado
                        processed_path = process_video_for_telegram(input_path)
                        
                        logger.info(f"📤 Resultado do processamento: {processed_path}")
                        
                        if processed_path and os.path.exists(processed_path):
                            # Enviar mensagem de sucesso
                            await update.message.reply_text("✅ Vídeo recortado automaticamente para formato quadrado!")
                            
                            logger.info(f"📤 Fazendo upload do vídeo processado: {processed_path}")
                            
                            # Fazer upload do vídeo processado e obter novo file_id
                            with open(processed_path, 'rb') as f:
                                sent = await context.bot.send_video_note(
                                    chat_id=update.effective_user.id, 
                                    video_note=f
                                )
                                
                                if sent and sent.video_note:
                                    video_file_id = sent.video_note.file_id
                                    video_width = 240  # Valor padrão para video notes
                                    video_height = 240
                                    video_size = sent.video_note.file_size
                                    video_duration = sent.video_note.duration
                                else:
                                    logger.error(f"❌ Erro ao processar vídeo automaticamente.")
                                    await update.message.reply_text("❌ Erro ao processar vídeo automaticamente.")
                                    safe_cleanup(temp_dir)
                                    return
                            
                            
                            logger.info(f"✅ Vídeo processado automaticamente: {video_width}x{video_height}")
                            
                        else:
                            logger.error(f"❌ Processamento falhou - processed_path: {processed_path}")
                            await update.message.reply_text(
                                f"❌ Erro ao processar vídeo automaticamente.\n\n"
                                f"📏 Dimensões atuais: {video_width}x{video_height}\n"
                                f"📋 Requisito: Largura = Altura (ex: 240x240)"
                            )
                            shutil.rmtree(temp_dir)
                            return
                            
                    except Exception as e:
                        logger.error(f"Erro ao processar vídeo automaticamente: {e}")
                        import traceback
                        logger.error(f"Traceback: {traceback.format_exc()}")
                        await update.message.reply_text(
                            f"❌ Erro ao processar vídeo automaticamente.\n\n"
                            f"📏 Dimensões atuais: {video_width}x{video_height}\n"
                            f"📋 Requisito: Largura = Altura (ex: 240x240)"
                        )
                        if 'temp_dir' in locals():
                            shutil.rmtree(temp_dir)
                        return
                
                # Verificar duração (máximo 60 segundos para video_note)
                if video_duration > 60:
                    await update.message.reply_text(
                        f"❌ O vídeo circular deve ter no máximo 60 segundos!\n\n"
                        f"⏱️ Duração atual: {video_duration} segundos\n"
                        f"📋 Máximo permitido: 60 segundos"
                    )
                    return
                
                # Verificar tamanho do arquivo (máximo 8MB para video_note)
                if video_size and video_size > 8 * 1024 * 1024:
                    await update.message.reply_text(
                        f"❌ O vídeo circular é muito grande!\n\n"
                        f"📦 Tamanho atual: {video_size // (1024*1024)} MB\n"
                        f"📋 Máximo permitido: 8 MB"
                    )
                    return
            
            # Verificar tamanho do vídeo normal (máximo 50MB)
            if not is_videonote and video_size and video_size > 50 * 1024 * 1024:
                await update.message.reply_text("❌ O vídeo é muito grande. Máximo permitido: 50MB")
                return
            
            # Salvar informações do vídeo no contexto
            context.user_data['broadcast_video'] = {
                'file_id': video_file_id,
                'duration': video_duration,
                'size': video_size,
                'width': video_width,
                'height': video_height,
                'is_videonote': is_videonote
            }
            
            # Aguardar texto da mensagem
            context.user_data['waiting_for_broadcast_text'] = True
            
            # Confirmar recebimento do vídeo
            keyboard = [[InlineKeyboardButton("❌ Cancelar", callback_data="admin_broadcast")]]
            reply_markup = InlineKeyboardMarkup(keyboard)
            
            target = "todos os usuários" if broadcast_type.endswith('_all') else "usuários VIP"
            video_type = "circular" if is_videonote else "normal"
            
            # Mensagem de confirmação personalizada
            if is_videonote:
                await update.message.reply_text(
                    f"✅ Vídeo circular recebido!\n\n"
                    f"⭕ Tipo: Vídeo Circular\n"
                    f"📹 Duração: {video_duration} segundos\n"
                    f"📏 Dimensões: {video_width}x{video_height} (quadrado)\n"
                    f"📦 Tamanho: {video_size // (1024*1024) if video_size else 'N/A'} MB\n\n"
                    f"Agora digite o texto da mensagem que será enviada junto com o vídeo circular para {target}:",
                    reply_markup=reply_markup
                )
            else:
                await update.message.reply_text(
                    f"✅ Vídeo recebido!\n\n"
                    f"📹 Tipo: Vídeo Normal\n"
                    f"📹 Duração: {video_duration} segundos\n"
                    f"📏 Dimensões: {video_width}x{video_height}\n"
                    f"📦 Tamanho: {video_size // (1024*1024) if video_size else 'N/A'} MB\n\n"
                    f"Agora digite o texto da mensagem que será enviada junto com o vídeo para {target}:",
                    reply_markup=reply_markup
                )
            
        except Exception as e:
            logger.error(f"Erro ao processar vídeo de broadcast: {e}")
            await update.message.reply_text("❌ Erro ao processar vídeo. Tente novamente.")
            # Limpar estado em caso de erro
            if 'broadcast_type' in context.user_data:
                del context.user_data['broadcast_type']
            if 'broadcast_video' in context.user_data:
                del context.user_data['broadcast_video']
            if 'waiting_for_broadcast_text' in context.user_data:
                del context.user_data['waiting_for_broadcast_text']

async def handle_welcome_file_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    config = load_config()
    messages = load_messages_from_db()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    if not is_admin(int(update.effective_user.id)):
        return
    
    # Inicializa welcome_file se não existir
    if 'welcome_file' not in config:
        config['welcome_file'] = {
            'enabled': False,
            'file_id': '',
            'file_type': 'photo',
            'caption': messages.get('welcome_caption', 'sem mensagem definida!')
        }
    
    # Alternar status do arquivo de boas-vindas
    current_status = config['welcome_file'].get('enabled', False)
    config['welcome_file']['enabled'] = not current_status
    new_status = config['welcome_file']['enabled']
    message = load_messages_from_db()
    
    # Salvar configurações
    if save_config(config):
        # Atualizar mensagem
        status = "ativado" if new_status else "desativado"
        await query.answer(f"✅ Arquivo de boas-vindas {status}!")
        
        # Recarregar configuração após salvar
        config = load_config()
        if config is None:
            logger.error("Falha ao carregar as configurações.")
            return  # ou lidar de forma apropriada
        welcome_file_config = config.get('welcome_file', {})
        is_enabled = welcome_file_config.get('enabled', False)
        file_type = welcome_file_config.get('file_type', 'photo')
        caption = welcome_file_config.get('caption', message.get('start_message', 'sem mensagem definida!'))
        
        # Recriar o menu com o status atualizado
        keyboard = [
            [InlineKeyboardButton(
                f"{'🔴' if not is_enabled else '🟢'} {'Desativar' if is_enabled else 'Ativar'} Arquivo",
                callback_data="admin_toggle_welcome_file"
            )],
            [InlineKeyboardButton("📎 Enviar Novo Arquivo", callback_data="admin_upload_welcome_file")],
            [InlineKeyboardButton("📝 Editar Legenda", callback_data="admin_edit_welcome_caption")],
            [InlineKeyboardButton("🗑️ Remover Arquivo", callback_data="admin_remove_welcome_file")],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_settings")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        status_text = "📎 Arquivo de Boas-vindas\n\n"
        if is_enabled:
            status_text += f"✅ Status: Ativado\n"
            status_text += f"📁 Tipo: {file_type.title()}\n"
            status_text += f"📝 Legenda: {caption}\n"
        else:
            status_text += f"❌ Status: Desativado\n"
        
        status_text += "\nEscolha uma opção:"
        
        try:
            await query.message.edit_text(
                status_text,
                reply_markup=reply_markup
            )
        except Exception as e:
            # Se falhar ao editar, tenta enviar uma nova mensagem
            await query.message.reply_text(
                f"✅ Arquivo de boas-vindas {status}!\n\n{status_text}",
                reply_markup=reply_markup
            )
    else:
        await query.answer("❌ Erro ao salvar configuração")

async def handle_maintenance_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        await query.message.reply_text("❌ Erro ao carregar configurações.")
        return

    if not is_admin(int(update.effective_user.id)):
        await query.message.reply_text("Acesso negado.")
        return

    # Garante que 'admin_settings' existe
    if 'admin_settings' not in config:
        config['admin_settings'] = {}

    # Alternar o modo de manutenção
    current_mode = config['admin_settings'].get('maintenance_mode', False)
    new_mode = not current_mode
    config['admin_settings']['maintenance_mode'] = new_mode

    # Salvar no campo admin_settings
    admin_success = save_config_to_db('admin_settings', json.dumps(config['admin_settings']))
    
    # Atualizar também a chave separada 'maintenance_mode' (se você quiser mantê-la sincronizada)
    simple_success = save_config_to_db('maintenance_mode', str(new_mode))  # salve como string: "True"/"False"

    success = admin_success and simple_success
    status = "🟢 Ativado" if new_mode else "🔴 Desativado"

    if success:
        keyboard = [
            [InlineKeyboardButton(
                "🔴 Desativar Manutenção" if new_mode else "🟢 Ativar Manutenção",
                callback_data="admin_toggle_maintenance"
            )],
            [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.message.edit_text(
            f"🔄 Modo Manutenção\n\nStatus atual: {status}",
            reply_markup=reply_markup
        )
        logger.info(f"Modo manutenção {'ativado' if new_mode else 'desativado'} pelo admin {update.effective_user.id}")
    else:
        await query.message.reply_text("❌ Erro ao salvar configuração. Tente novamente.")
  
async def handle_payment_toggle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    logger.info(f"Callback de toggle recebido: {query.data}")
    
    try:
        config = load_config()
        if config is None:
            logger.error("Falha ao carregar as configurações.")
            return  # ou lidar de forma apropriada
        logger.info(f"Config carregada: {config}")
        
        if str(update.effective_user.id):
            logger.error("Acesso negado - ID não corresponde")
            return
        
        # Extrai o tipo de PIX do callback
        if "pix_auto" in query.data:
            action = "auto"
        elif "pix_manual" in query.data:
            action = "manual"
        else:
            logger.error(f"Callback inválido: {query.data}")
            return
            
        logger.info(f"Ação de toggle: {action}")
        
        # Alterna o estado do método correto
        if action == "auto":
            current_state = config['payment_methods']['pix_automatico']['enabled']
            logger.info(f"Estado atual do PIX Automático: {current_state}")
            config['payment_methods']['pix_automatico']['enabled'] = not current_state
            new_state = config['payment_methods']['pix_automatico']['enabled']
            logger.info(f"Novo estado do PIX Automático: {new_state}")
            method_name = "Automático"
        else:  # manual
            current_state = config['payment_methods']['pix_manual']['enabled']
            logger.info(f"Estado atual do PIX Manual: {current_state}")
            config['payment_methods']['pix_manual']['enabled'] = not current_state
            new_state = config['payment_methods']['pix_manual']['enabled']
            logger.info(f"Novo estado do PIX Manual: {new_state}")
            method_name = "Manual"
        
        # Salva a configuração
        logger.info("Tentando salvar configuração...")
        if save_config(config):
            logger.info("Configuração salva com sucesso")
            # Atualiza a mensagem
            keyboard = [
                [InlineKeyboardButton("🔑 Token do Bot", callback_data="admin_edit_bot_token")],
                [InlineKeyboardButton("💳 Token MercadoPago", callback_data="admin_edit_mp_token")],
                [InlineKeyboardButton("📱 Chave PIX", callback_data="admin_edit_pix_key")],
                [InlineKeyboardButton("👤 Nome Titular PIX", callback_data="admin_edit_pix_name")],
                [InlineKeyboardButton(
                    f"{'🔴' if not config['payment_methods']['pix_automatico']['enabled'] else '🟢'} PIX Automático",
                    callback_data="admin_toggle_pix_auto"
                )],
                [InlineKeyboardButton(
                    f"{'🔴' if not config['payment_methods']['pix_manual']['enabled'] else '🟢'} PIX Manual",
                    callback_data="admin_toggle_pix_manual"
                )],
                [InlineKeyboardButton("📎 Arquivo de Boas-vindas", callback_data="admin_welcome_file")],
                [InlineKeyboardButton("⬅️ Voltar", callback_data="admin_back")]
            ]
            reply_markup = InlineKeyboardMarkup(keyboard)
            status = "ativado" if new_state else "desativado"
            await query.message.edit_text(
                f"⚙️ Configurações\n\nPIX {method_name} {status}!",
                reply_markup=reply_markup
            )
        else:
            logger.error("Falha ao salvar configuração")
            await query.message.reply_text("❌ Erro ao salvar configuração. Tente novamente.")
            
    except Exception as e:
        logger.error(f"Erro ao alternar PIX {action}: {e}")
        await query.message.reply_text("❌ Erro ao alternar método de pagamento. Tente novamente.")


async def process_access_delivery_queue(context):
    """Processa a fila de entrega de acesso VIP"""
    try:
        logger.info(f"🔄 Job process_access_delivery_queue executado - Fila vazia: {access_delivery_queue.empty()}")
        
        while not access_delivery_queue.empty():
            event = access_delivery_queue.get()
            user_id = event['user_id']
            plan_id = event['plan_id']
            
            logger.info(f"🎯 Processando entrega de acesso VIP para usuário {user_id} (plano {plan_id})")
            
            try:
                await add_user_to_vip_groups(context.bot, user_id, plan_id)
                logger.info(f"✅ Entrega de acesso VIP concluída para usuário {user_id} (plano {plan_id})")
            except Exception as e:
                logger.error(f"❌ Erro ao entregar acesso VIP para usuário {user_id}: {e}")
                import traceback
                logger.error(f"Traceback: {traceback.format_exc()}")
        
        if access_delivery_queue.empty():
            logger.info("📭 Fila vazia - nenhum item para processar")
            
    except Exception as e:
        logger.error(f"❌ Erro ao processar fila de entrega de acesso: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")

async def check_expired_subscriptions(context: ContextTypes.DEFAULT_TYPE):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return
        
        # Obter assinaturas expiradas
        expired_subs = db.execute_fetch_all(
            """SELECT s.*, vp.name as plan_name, vp.price 
            FROM subscriptions s
            JOIN vip_plans vp ON s.plan_id = vp.id
            WHERE s.is_active = TRUE 
            AND s.is_permanent = FALSE
            AND s.end_date <= NOW()"""
        )
        
        for sub in expired_subs:
            try:
                # Desativar assinatura
                db.execute_query(
                    "UPDATE subscriptions SET is_active = FALSE WHERE id = %s",
                    (sub['id'],),
                    commit=True
                )
                
                # Atualizar status do usuário
                db.execute_query(
                    "UPDATE users SET is_vip = FALSE WHERE id = %s",
                    (sub['user_id'],),
                    commit=True
                )
                
                # Remover usuário dos grupos VIP
                await remove_user_from_vip_groups(context.bot, sub['user_id'], sub['plan_id'])
                
                # Notificar usuário
                await context.bot.send_message(
                    chat_id=sub['user_id'],
                    text=f"⚠️ Sua assinatura VIP expirou!\n\n"
                         f"Plano: {sub['plan_name']}\n"
                         f"Data de expiração: {sub['end_date']}\n\n"
                         f"🚫 Você foi removido dos grupos VIP.\n\n"
                         f"Para continuar com acesso VIP, adquira um novo plano usando /start"
                )
                
                logger.info(f"Assinatura expirada processada: usuário {sub['user_id']}, plano {sub['plan_id']}")
                
            except Exception as e:
                logger.error(f"Erro ao processar assinatura expirada {sub['id']}: {e}")
                
    except Exception as e:
        logger.error(f"Erro ao verificar assinaturas expiradas: {e}")
    finally:
        db.close()

async def process_scheduled_messages(context: ContextTypes.DEFAULT_TYPE):
    """Processa mensagens agendadas que devem ser enviadas"""
    try:
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logger.info(f"🕐 [{current_time}] Iniciando processamento de mensagens agendadas...")
        
        # Limpar mensagens muito antigas primeiro
        cleanup_old_scheduled_messages()
        
        # Obter mensagens pendentes
        pending_messages = get_pending_scheduled_messages()
        
        if not pending_messages:
            logger.info(f"📭 [{current_time}] Nenhuma mensagem agendada pendente.")
            return
        
        logger.info(f"📨 [{current_time}] Processando {len(pending_messages)} mensagens agendadas...")
        
        for message in pending_messages:
            try:
                # Verificar se a mensagem está atrasada
                scheduled_time = message['scheduled_date']
                if isinstance(scheduled_time, str):
                    scheduled_time = datetime.strptime(scheduled_time, "%Y-%m-%d %H:%M:%S")
                
                current_time = datetime.now()
                delay_minutes = (current_time - scheduled_time).total_seconds() / 60
                
                if delay_minutes > 0:
                    logger.info(f"⏰ Processando mensagem agendada ID: {message['id']} (atrasada por {delay_minutes:.1f} minutos)")
                else:
                    logger.info(f"🎯 Processando mensagem agendada ID: {message['id']} (no horário)")
                
                # Obter destinatários
                recipients = get_recipients_for_scheduled_message(message)
                
                if not recipients:
                    logger.warning(f"Nenhum destinatário encontrado para mensagem {message['id']}")
                    update_scheduled_message_status(
                        message['id'], 
                        'failed', 
                        error_message="Nenhum destinatário encontrado"
                    )
                    continue
                
                logger.info(f"Enviando mensagem {message['id']} para {len(recipients)} destinatários...")
                
                # Enviar mensagem para todos os destinatários
                successful_sends = 0
                failed_sends = 0
                
                for user_id in recipients:
                    try:
                        await context.bot.send_message(
                            chat_id=user_id,
                            text=message['message_text']
                        )
                        successful_sends += 1
                        
                        # Pequena pausa para evitar rate limiting
                        await asyncio.sleep(0.1)
                        
                    except Exception as e:
                        logger.error(f"Erro ao enviar mensagem para usuário {user_id}: {e}")
                        failed_sends += 1
                
                # Atualizar status da mensagem
                update_scheduled_message_status(
                    message['id'],
                    'sent',
                    successful_sends=successful_sends,
                    failed_sends=failed_sends
                )
                
                logger.info(f"Mensagem {message['id']} processada: {successful_sends} sucessos, {failed_sends} falhas")
                
            except Exception as e:
                logger.error(f"Erro ao processar mensagem agendada {message['id']}: {e}")
                update_scheduled_message_status(
                    message['id'],
                    'failed',
                    error_message=str(e)
                )
        
        logger.info(f"✅ [{current_time}] Processamento de mensagens agendadas concluído.")
        
    except Exception as e:
        logger.error(f"Erro no processamento de mensagens agendadas: {e}")

async def check_expiring_subscriptions(context: ContextTypes.DEFAULT_TYPE):
    """Verifica e notifica assinaturas próximas de expirar."""
    try:
        logger.info("Iniciando verificação de assinaturas próximas de expirar...")
        
        # Obter assinaturas próximas de expirar do banco de dados
        expiring_subscriptions = get_expiring_subscriptions()
        
        # Carregar configuração
        config = load_config()
        if config is None:
            logger.error("Falha ao carregar as configurações.")
            return  # ou lidar de forma apropriada
        current_time = datetime.now()
        logger.info(f"Verificando assinaturas em: {current_time}")
        logger.info(f"Total de assinaturas próximas de expirar: {len(expiring_subscriptions)}")
        
        for sub in expiring_subscriptions:
            try:
                # Calcular dias restantes
                end_date = sub['end_date']
                if isinstance(end_date, str):
                    end_date = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
                time_left = end_date - current_time
                days_left = time_left.days
                hours_left = time_left.seconds // 3600
                
                logger.info(f"Verificando assinatura do usuário {sub['user_id']}:")
                logger.info(f"Dias restantes: {days_left}")
                logger.info(f"Horas restantes: {hours_left}")
                
                # Verificar se deve notificar (3, 2 ou 1 dia, ou menos de 24 horas)
                should_notify = False
                notification_key = None
                
                if days_left == 0 and hours_left <= 24:
                    should_notify = True
                    notification_key = "notified_1"
                elif days_left == 1:
                    should_notify = True
                    notification_key = "notified_1"
                elif days_left == 2:
                    should_notify = True
                    notification_key = "notified_2"
                elif days_left == 3:
                    should_notify = True
                    notification_key = "notified_3"
                
                # Verificar se já foi notificado
                already_notified = sub.get(notification_key, False) if notification_key else False
                
                if should_notify and not already_notified:
                    # Notificar usuário
                    try:
                        message = f"⚠️ Sua assinatura VIP está próxima de expirar!\n\n"
                        message += f"Plano: {sub['plan_name']}\n"
                        if days_left == 0:
                            message += f"Horas restantes: {hours_left}\n"
                        else:
                            message += f"Dias restantes: {days_left}\n"
                        message += f"Data de expiração: {sub['end_date']}\n\n"
                        message += f"Para renovar seu acesso VIP, use /start e escolha um novo plano! 🎉"
                        
                        await context.bot.send_message(
                            chat_id=sub['user_id'],
                            text=message
                        )
                        logger.info(f"Notificação enviada para usuário {sub['user_id']}")
                        
                        # Marcar como notificado no banco de dados
                        update_subscription_notification(sub['id'], notification_key)
                        logger.info(f"Usuário {sub['user_id']} marcado como notificado para {notification_key}")
                        
                    except Exception as e:
                        logger.error(f"Erro ao notificar usuário {sub['user_id']}: {e}")
            
            except Exception as e:
                logger.error(f"Erro ao processar assinatura próxima de expirar: {e}")
        
        logger.info("Verificação de assinaturas próximas de expirar concluída!")
            
    except Exception as e:
        logger.error(f"Erro ao verificar assinaturas próximas de expirar: {e}")

async def initial_check(context: ContextTypes.DEFAULT_TYPE):
    """Verificação inicial de assinaturas quando o bot inicia."""
    logger.info("Iniciando verificação inicial de assinaturas...")
    
    # Verificar assinaturas expiradas
    await check_expired_subscriptions(context)
    
    # Verificar assinaturas próximas de expirar
    await check_expiring_subscriptions(context)
    
    logger.info("Verificação inicial concluída!")

async def handle_back_to_plans(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    
    # Limpar estado do usuário
    if 'waiting_for_proof' in context.user_data:
        del context.user_data['waiting_for_proof']
    
    # Retornar para a lista de planos
    keyboard = []
    db = Database()
    try:
        db.connect()
        plans = db.execute_fetch_all("SELECT * FROM vip_plans")
        for plan in plans:
            keyboard.append([InlineKeyboardButton(
                f"{plan['name']} - R${plan['price']:.2f}",
                callback_data=f"plan_{plan['id']}"
            )])
    finally:
        db.close()
    reply_markup = InlineKeyboardMarkup(keyboard)
    await query.message.edit_text(
        "Escolha um dos planos VIP disponíveis:",
        reply_markup=reply_markup
    )

async def check_bot_initialization(bot, config):
    """Verifica a inicialização do bot e envia relatório ao admin."""
    try:
        # Verificar se o token é válido
        bot_info = await bot.get_me()
        logger.info(f"Bot iniciado com sucesso: @{bot_info.username}")
        
        # Verificar dependências
        missing_deps = []
        try:
            import qrcode
        except ImportError:
            missing_deps.append("qrcode")
        try:
            import mercadopago
        except ImportError:
            missing_deps.append("mercadopago")
        try:
            from PIL import Image
        except ImportError:
            missing_deps.append("Pillow")
            
        # Verificar arquivos de configuração
        missing_files = []
        if not os.path.exists('config.json'):
            missing_files.append("config.json")
        if not os.path.exists('messages.txt'):
            missing_files.append("messages.txt")
            
        # Verificar estrutura do config.json
        config_errors = []
        if 'bot_token' not in config:
            config_errors.append("Token do bot não encontrado")
        if 'admin_id' not in config:
            config_errors.append("ID do admin não encontrado")
        if 'payment_methods' not in config:
            config_errors.append("Configurações de pagamento não encontradas")
        # Removida verificação de vip_plans pois agora está no banco de dados
            
        # Preparar mensagem de status
        status_message = f"🤖 *Status de Inicialização do Bot*\n\n"
        status_message += f"✅ Bot iniciado: @{bot_info.username}\n"
        
        if missing_deps:
            status_message += f"\n❌ Dependências faltando:\n"
            for dep in missing_deps:
                status_message += f"• {dep}\n"
                
        if missing_files:
            status_message += f"\n❌ Arquivos faltando:\n"
            for file in missing_files:
                status_message += f"• {file}\n"
                
        if config_errors:
            status_message += f"\n❌ Erros de configuração:\n"
            for error in config_errors:
                status_message += f"• {error}\n"
                
        if not (missing_deps or missing_files or config_errors):
            status_message += "\n✅ Todas as verificações passaram com sucesso!"
            
        # Enviar mensagem ao admin
        try:
            for admin_id in get_all_admin_ids():
                await bot.send_message(chat_id=admin_id, text=status_message, parse_mode='Markdown')
            Logger.info("Relatório de inicialização enviado ao admin")
        except Exception as e:
            logger.error(f"Erro ao enviar relatório ao admin: {e}")
            
    except Exception as e:
        logger.error(f"Erro ao verificar inicialização: {e}")
        try:
            for admin_id in get_all_admin_ids():
                await bot.send_message(chat_id=admin_id, text=f"❌ *Erro na inicialização do bot*\n\nErro: {str(e)}", parse_mode="Markdown")
        except:
            logger.error("Não foi possível enviar mensagem de erro ao admin")


def main():
    try:
        # Iniciar o webhook do CNPay em thread separada
        webhook_thread = threading.Thread(target=start_cnpay_webhook, daemon=True)
        webhook_thread.start()
        
        # Verificar conexão com o banco de dados
        logger.info("🔍 Verificando conexão com o banco de dados...")
        from db_config import DB_CONFIG
        
        # Verificar se a configuração existe
        if DB_CONFIG is None:
            logger.error("❌ Configuração do banco de dados não encontrada")
            logger.error("📋 Configure as variáveis de ambiente necessárias no Railway:")
            logger.error("   - DB_HOST")
            logger.error("   - DB_PORT") 
            logger.error("   - DB_USER")
            logger.error("   - DB_PASSWORD")
            logger.error("   - DB_NAME (opcional, padrão: bot_vip)")
            logger.error("💡 O bot não pode funcionar sem conexão com o banco de dados")
            return
        
        # Mostrar informações da configuração
        logger.info(f"📋 Configuração do banco:")
        logger.info(f"   Host: {DB_CONFIG.get('host', 'Não definido')}")
        logger.info(f"   Porta: {DB_CONFIG.get('port', 'Não definida')}")
        logger.info(f"   Usuário: {DB_CONFIG.get('user', 'Não definido')}")
        logger.info(f"   Banco: {DB_CONFIG.get('database', 'Não definido')}")
        
        db = Database()
        try:
            connection = db.connect()
            if connection and connection.is_connected():
                cursor = connection.cursor()
                cursor.execute("SELECT VERSION()")
                version = cursor.fetchone()
                cursor.close()
                logger.info(f"✅ Conectado ao MySQL versão: {version[0]}")
                logger.info(f"✅ Banco de dados configurado corretamente na porta {DB_CONFIG.get('port', 'Não definida')}")
            else:
                logger.error("❌ Falha na conexão com o banco de dados")
                logger.error("💡 Verifique se a porta está correta e o MySQL está rodando")
        except Exception as e:
            logger.error(f"❌ Erro ao conectar ao banco de dados: {e}")
            logger.error("💡 Execute 'python setup_database.py' para reconfigurar")
        finally:
            db.close()

        # Carregar configurações
        config = load_config()
        if config is None:
            logger.error("Falha ao carregar as configurações.")
            return  # ou lidar de forma apropriada
        if not config or 'bot_token' not in config:
            logger.error("Token do bot não encontrado na configuração.")
            return

        # Inicializar o bot
        application = Application.builder().token(config['bot_token']).build()
        
        # Definir as instâncias globais
        set_application_instance(application)
        set_bot_instance(application.bot)
        
        # Criar contexto compartilhado
        shared_context = create_bot_context()
        if shared_context:
            logger.info("✅ Contexto compartilhado criado para outras threads")
            # Configurar contexto compartilhado global
            set_shared_context(application.bot, application, shared_context)
        else:
            logger.warning("⚠️ Não foi possível criar contexto compartilhado")

        # Adicionar handlers
        logger.info("🔧 Registrando handlers de comandos...")
        application.add_handler(CommandHandler("start", start))
        application.add_handler(CommandHandler("vip", vip))
        application.add_handler(CommandHandler("admin", admin))
        application.add_handler(CommandHandler("database", database))
        application.add_handler(CommandHandler("database_simple", database_simple))
        application.add_handler(CommandHandler("test", test_command))
        application.add_handler(CommandHandler("test_users", test_users))  # Comando temporário para debug
        logger.info("✅ Handlers de comandos registrados com sucesso!")
        application.add_handler(CallbackQueryHandler(handle_plan_selection, pattern="^plan_"))
        application.add_handler(CallbackQueryHandler(handle_plan_selection, pattern="^renew_"))
        application.add_handler(CallbackQueryHandler(handle_renewal_confirmation, pattern="^confirm_renew_"))
        application.add_handler(CallbackQueryHandler(handle_payment_method, pattern="^pix_"))
        application.add_handler(CallbackQueryHandler(check_payment_manual, pattern="^check_"))
        application.add_handler(CallbackQueryHandler(check_payment_manual, pattern="^copy_pix_"))
        application.add_handler(CallbackQueryHandler(handle_back_to_plans, pattern="^back_to_plans$"))
        application.add_handler(CallbackQueryHandler(handle_show_plans, pattern="^show_plans$"))
        application.add_handler(CallbackQueryHandler(handle_admin_edit, pattern="^admin_edit_"))
        application.add_handler(CallbackQueryHandler(handle_admin_callback, pattern="^admin_(?!edit_)"))
        application.add_handler(CallbackQueryHandler(handle_maintenance_toggle, pattern="^admin_toggle_maintenance$"))
        application.add_handler(CallbackQueryHandler(handle_payment_toggle, pattern="^admin_toggle_pix_"))
        application.add_handler(CallbackQueryHandler(handle_welcome_file_toggle, pattern="^admin_toggle_welcome_file$"))
        application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_admin_text))
        application.add_handler(MessageHandler(filters.PHOTO | filters.VIDEO, handle_admin_files))

        # Adicionar jobs periódicos (se job_queue estiver disponível)
        job_queue = application.job_queue
        if job_queue is not None:
            try:
                job_queue.run_repeating(check_expired_subscriptions, interval=3*60, first=10)
                job_queue.run_repeating(check_expiring_subscriptions, interval=60*60, first=20)
                job_queue.run_repeating(process_access_delivery_queue, interval=5, first=5)
                job_queue.run_repeating(process_scheduled_messages, interval=60, first=30)  # Verificar mensagens agendadas a cada minuto
                job_queue.run_once(initial_check, when=5)
                logger.info("✅ Jobs periódicos configurados com sucesso")
            except Exception as e:
                logger.warning(f"⚠️ Erro ao configurar jobs periódicos: {e}")
                logger.info("ℹ️ Bot funcionará sem jobs periódicos (verificações manuais)")
        else:
            logger.warning("⚠️ JobQueue não disponível. Bot funcionará sem verificações automáticas")
            logger.info("ℹ️ Para habilitar verificações automáticas, instale: pip install 'python-telegram-bot[job-queue]'")

        # Handler de erros
        application.add_error_handler(error_handler)

        # Iniciar o bot
        application.run_polling()

    except Exception as e:
        logger.error(f"Erro ao iniciar o bot: {e}")
        time.sleep(5)
        main()
async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Tratamento de erros do bot."""
    logger.error(f"Erro não tratado: {context.error}")
    
    # Se for erro de conflito, apenas logar e não tentar reiniciar
    if "Conflict" in str(context.error):
        logger.warning("⚠️ Detectado conflito de instâncias do bot. Aguardando resolução automática...")
        # Não tentar reiniciar - deixar o Telegram resolver o conflito
        return
    else:
        # Para outros erros, apenas logar
        logger.error(f"Erro: {context.error}")
        if update:
            logger.error(f"Update: {update}")

# Remover usuário dos grupos VIP
async def remove_user_from_vip_groups(bot, user_id, plan_id):
    """Remove usuário dos grupos VIP quando a assinatura expira"""
    config = load_config()
    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    
    # Buscar o plano no banco de dados
    db = Database()
    try:
        db.connect()
        if not db.connection:
            logger.error("Erro ao conectar ao banco de dados")
            return False
        
        # Buscar grupos associados ao plano
        groups = db.execute_fetch_all(
            """SELECT vg.group_id, vg.group_name
            FROM vip_groups vg
            JOIN plan_groups pg ON vg.id = pg.group_id
            WHERE pg.plan_id = %s AND vg.is_active = TRUE""",
            (plan_id,)
        )
        
        if not groups:
            logger.info(f"Nenhum grupo encontrado para o plano {plan_id}")
            return True  # Retorna True mesmo sem grupos
        
        # Remover usuário dos grupos
        for group in groups:
            group_id = group['group_id']
            group_name = group['group_name']
            
            try:
                # Verificar se o grupo é um supergrupo
                chat = await bot.get_chat(group_id)
                if chat.type in ['group', 'supergroup', 'channel']:
                    try:
                        # Remover usuário do grupo
                        await bot.ban_chat_member(
                            chat_id=group_id,
                            user_id=user_id,
                            until_date=datetime.now() + timedelta(seconds=30)  # Ban temporário de 30 segundos
                        )
                        
                        logger.info(f"Usuário {user_id} removido do grupo {group_id} ({group_name})")
                        
                        # Notificar admin sobre a remoção
                        if config and 'admin_id' in config:
                            if not is_admin(int(update.effective_user.id)):
                                return
                            await bot.send_message(
                                chat_id=admin_id,
                                text=f"🚫 Usuário removido do grupo VIP\n\n"
                                     f"👤 Usuário: {user_id}\n"
                                     f"📱 Grupo: {group_name}\n"
                                     f"💎 Plano: {plan_id}\n"
                                     f"⏰ Motivo: Assinatura expirada"
                            )
                        
                    except Exception as e:
                        logger.error(f"Erro ao remover usuário {user_id} do grupo {group_id}: {e}")
                        # Se falhar, notifica o admin
                        if config and 'admin_id' in config:
                            if not is_admin(int(update.effective_user.id)):
                                return
                            await bot.send_message(
                                chat_id=admin_id,
                                text=f"⚠️ Erro ao remover usuário do grupo VIP\n\n"
                                     f"👤 Usuário: {user_id}\n"
                                     f"📱 Grupo: {group_name}\n"
                                     f"💎 Plano: {plan_id}\n"
                                     f"❌ Erro: {e}\n\n"
                                     f"Verifique se o bot tem permissões de administrador no grupo."
                            )
                else:
                    logger.error(f"Grupo {group_id} não é um grupo ou supergrupo válido")
                    
            except Exception as e:
                logger.error(f"Erro ao processar grupo {group_id} para remoção do usuário {user_id}: {e}")
                # Notifica o admin
                if config and 'admin_id' in config:
                    if not is_admin(int(update.effective_user.id)):
                        return
                    await bot.send_message(
                        chat_id=admin_id,
                        text=f"⚠️ Erro ao processar grupo {group_id} para remoção do usuário {user_id}.\nErro: {e}"
                    )
        
        return True
        
    except Exception as e:
        logger.error(f"Erro ao buscar grupos do plano {plan_id} para remoção: {e}")
        return False
    finally:
        db.close()

# =====================================================
# SISTEMA DE MÚLTIPLOS PROVEDORES PIX
# =====================================================

class PixProvider:
    """Classe base para provedores PIX"""
    
    def __init__(self, config):
        self.config = config
    
    async def generate_pix(self, amount, description, external_reference):
        """Gera PIX - deve ser implementado pelos provedores"""
        raise NotImplementedError
    
    async def check_payment(self, payment_id):
        """Verifica pagamento - deve ser implementado pelos provedores"""
        raise NotImplementedError

class MercadoPagoProvider(PixProvider):
    """Provedor MercadoPago"""
    
    async def generate_pix(self, amount, description, external_reference):
        """Gera PIX usando MercadoPago"""
        try:
            if not self.config.get('mercadopago_access_token'):
                logger.error("Token do MercadoPago não configurado")
                return None
            
            sdk = mercadopago.SDK(self.config['mercadopago_access_token'])
            
            payment_data = {
                "transaction_amount": float(amount),
                "description": description,
                "payment_method_id": "pix",
                "external_reference": external_reference,
                "payer": {
                    "email": "cliente@email.com",
                    "first_name": "Cliente",
                    "last_name": "Teste"
                }
            }

            payment_response = sdk.payment().create(payment_data)
            payment = payment_response["response"]
            
            if "point_of_interaction" in payment:
                # Registrar pagamento no banco de dados
                db = Database()
                try:
                    db.connect()
                    if db.connection:
                        db.execute_query(
                            """INSERT INTO payments 
                            (payment_id, user_id, plan_id, amount, currency, 
                             payment_method, status, external_reference, 
                             qr_code_data, pix_key, pix_key_type, pix_key_owner) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                            (
                                payment['id'],
                                int(external_reference.split('_')[0]),  # user_id
                                int(external_reference.split('_')[1]),  # plan_id
                                amount,
                                'BRL',
                                'mercadopago',
                                payment['status'],
                                external_reference,
                                payment['point_of_interaction']['transaction_data']['qr_code'],
                                None,  # pix_key (para PIX automático é nulo)
                                None,
                                None
                            ),
                            commit=True
                        )
                except Exception as e:
                    logger.error(f"Erro ao registrar pagamento no banco: {e}")
                finally:
                    db.close()
                
                return {
                    "qr_code": payment["point_of_interaction"]["transaction_data"]["qr_code"],
                    "qr_code_base64": payment["point_of_interaction"]["transaction_data"]["qr_code_base64"],
                    "payment_id": payment["id"],
                    "provider": "mercadopago"
                }
            else:
                logger.error("Resposta do MercadoPago não contém point_of_interaction")
                return None
                
        except Exception as e:
            logger.error(f"Erro ao gerar PIX MercadoPago: {e}")
            return None
    
    async def check_payment(self, payment_id):
        """Verifica pagamento no MercadoPago"""
        try:
            if not self.config.get('mercadopago_access_token'):
                logger.error("Token do MercadoPago não configurado")
                return None
            
            sdk = mercadopago.SDK(self.config['mercadopago_access_token'])
            payment_response = sdk.payment().get(payment_id)
            payment = payment_response["response"]
            
            return payment
        except Exception as e:
            logger.error(f"Erro ao verificar pagamento MercadoPago {payment_id}: {e}")
            return None

class CNPayProvider(PixProvider):
    """Provedor CNPay com suporte a split fixo e validação de taxas"""
    
    def __init__(self, config):
        super().__init__(config)
        self.api_key = config.get('cnpay_api_key', '')
        self.api_secret = config.get('cnpay_api_secret', '')
        self.environment = config.get('cnpay_environment', 'sandbox')
        
        # Configurar URL baseada no ambiente
        if self.environment == 'sandbox':
            self.base_url = 'https://sandbox.appcnpay.com/api/v1/gateway/pix/receive'
        else:
            self.base_url = 'https://painel.appcnpay.com/api/v1/gateway/pix/receive'
        
        self.fixed_split = 0.75  # Split fixo de R$ 0,75
        self.fee_percentage = 0.029  # Taxa de 2.9%
        
        logger.info(f"🔧 CNPay configurado - Ambiente: {self.environment}")
        logger.info(f"🔧 URL: {self.base_url}")
        logger.info(f"🔧 API Key configurada: {'✅' if self.api_key else '❌'}")
        logger.info(f"🔧 API Secret configurado: {'✅' if self.api_secret else '❌'}")
    
    async def generate_pix(self, amount, description, external_reference, splits=None):
        """Gera PIX usando CNPay Gateway com validação de split e taxas"""
        try:
            # Extrair user_id e plan_id logo no início
            # Verificar se é um pagamento de admin VIP
            if external_reference.startswith('admin_vip_'):
                # Formato: admin_vip_admin_id_timestamp
                parts = external_reference.split('_')
                user_id = parts[2]  # admin_id
                plan_id = 'admin_vip'  # Identificador especial para admin VIP
            else:
                # Formato normal: user_id_plan_id
                user_id, plan_id = external_reference.split('_')

            # Buscar informações do usuário e plano (sempre)
            db = Database()
            user_info = None
            plan_info = None
            try:
                db.connect()
                if db.connection:
                    if plan_id == 'admin_vip':
                        # Para admin VIP, buscar dados do admin
                        admin_query = "SELECT admin_id, user FROM admins WHERE admin_id = %s"
                        user_info = db.execute_fetch_one(admin_query, (int(user_id),))
                        # Criar plan_info fictício para admin VIP
                        plan_info = {
                            'id': 'admin_vip',
                            'name': 'Admin VIP',
                            'price': amount,
                            'duration_days': 30
                        }
                    else:
                        # Para usuários normais
                        user_query = "SELECT id, username, first_name, last_name FROM users WHERE id = %s"
                        user_info = db.execute_fetch_one(user_query, (int(user_id),))
                        plan_query = "SELECT id, name, price, duration_days FROM vip_plans WHERE id = %s"
                        plan_info = db.execute_fetch_one(plan_query, (int(plan_id),))
            except Exception as e:
                logger.error(f"Erro ao buscar dados do usuário/plano: {e}")
            finally:
                db.close()

            if plan_id == 'admin_vip':
                client_name = f"Admin {user_info['user']}" if user_info else "Admin"
                client_email = f"admin_{user_id}@telegram.com"
            else:
                client_name = f"{user_info['first_name']} {user_info['last_name'] or ''}" if user_info else "Cliente"
                client_email = (user_info['email'] if user_info and user_info.get('email') 
                               else f"{user_info['username']}@telegram.com" if user_info and user_info['username'] 
                               else "cliente@telegram.com")

            # Verifica se o valor cobre split + taxa
            amount_float = float(amount)
            min_amount = self.fixed_split / (1 - self.fee_percentage)
            if amount_float < min_amount:
                logger.info(f"Valor baixo, split não será aplicado. Valor: R$ {amount}, mínimo para split: R$ {min_amount:.2f}")
                splits = None
            else:
                # Configuração do split
                if splits is None:
                    cnpay_tax = (amount_float * 0.029) + 0.30
                    available_for_split = amount_float - cnpay_tax
                    if amount_float >= 5.00:
                        split_amount = min(0.49, available_for_split * 0.25)
                        split_amount = max(0.50, split_amount)
                        splits = [{"producerId": "cm909ruxy001wmynlf6ut7dnu", "amount": round(split_amount, 2)}]
                        logger.info(f"💰 Split calculado: R$ {split_amount:.2f} (Valor: R$ {amount_float}, Taxa: R$ {cnpay_tax:.2f}, Disponível: R$ {available_for_split:.2f})")
                    else:
                        splits = None
                        logger.info(f"💰 Sem split para valor baixo: R$ {amount_float} (Taxa: R$ {cnpay_tax:.2f})")

            # Montagem da requisição
            payment_data = {
                "identifier": f"user_{user_id}_{plan_id}",
                "amount": amount_float,
                "client": {
                    "name": client_name,
                    "email": client_email,
                    "phone": "(11) 99999-9999",
                    "document": "04697417160"
                },
                "products": [{
                    "id": f"plan_{plan_id}",
                    "name": plan_info['name'] if plan_info else "Plano VIP",
                    "quantity": 1,
                    "price": amount_float
                }],
                "dueDate": (datetime.now() + timedelta(days=2)).strftime("%Y-%m-%d"),
                "metadata": {
                    "user_id": user_id,
                    "plan_id": plan_id,
                    "provider": "CNPay Gateway",
                    "bot": "SaikaVIP"
                }
            }
            # Adicionar splits apenas se existir e não for None
            if splits:
                payment_data["splits"] = splits
            
            # Webhook configuration
            webhook_url = self.config.get('cnpay_webhook_url', '')
            if webhook_url:
                if not webhook_url.startswith(('http://', 'https://')):
                    webhook_url = 'https://' + webhook_url
                payment_data["callbackUrl"] = webhook_url
                logger.info(f"🔔 Callback configurado: {webhook_url}")
            
            # Headers and request - limpar caracteres de controle
            headers = {
                'Content-Type': 'application/json',
                'x-public-key': self.api_key.strip().replace('\r', '').replace('\n', ''),
                'x-secret-key': self.api_secret.strip().replace('\r', '').replace('\n', '')
            }
            
            # Log detalhado do JSON sendo enviado (sem expor credenciais)
            import json
            logger.info(f"📤 JSON enviado para CNPay:")
            logger.info(f"   URL: {self.base_url}")
            logger.info(f"   Headers: Content-Type: {headers.get('Content-Type')}, x-public-key: {self.api_key[:8]}..., x-secret-key: {self.api_secret[:8]}...")
            logger.info(f"   Payload: {json.dumps(payment_data, indent=2, ensure_ascii=False)}")
            
            import httpx
            async with httpx.AsyncClient() as client:
                response = await client.post(
                    self.base_url,
                    json=payment_data,
                    headers=headers,
                    timeout=30
                )
            
            # Log da resposta para debug
            logger.info(f"📥 Resposta CNPay - Status: {response.status_code}")
            logger.info(f"📥 Headers: {dict(response.headers)}")
            logger.info(f"📥 Conteúdo: {response.text[:500]}...")  # Primeiros 500 caracteres
            
            if response.status_code in (200, 201):
                try:
                    payment_info = response.json()
                    return self._process_success_response(payment_info, user_id, plan_id, amount, external_reference)
                except Exception as e:
                    logger.error(f"Erro ao processar JSON da resposta de sucesso: {e}")
                    logger.error(f"Conteúdo da resposta: {response.text}")
                    return None
            
            # Tratamento de erro melhorado
            try:
                if response.text and response.text.strip():
                    error_data = response.json()
                else:
                    error_data = {"error": "Resposta vazia do servidor"}
            except Exception as e:
                error_data = {
                    "error": f"Erro ao processar resposta: {str(e)}",
                    "raw_response": response.text[:200] if response.text else "Resposta vazia"
                }
            
            logger.error(f"❌ Erro CNPay {response.status_code}: {error_data}")
            return None
            
        except Exception as e:
            logger.error(f"Falha ao gerar PIX: {str(e)}", exc_info=True)
            return None
    
    def _process_success_response(self, payment_info, user_id, plan_id, amount, external_reference):
        """Processa resposta de sucesso da API"""
        try:
            # Registrar no banco de dados
            db = Database()
            try:
                db.connect()
                if db.connection:
                    status_map = {
                        'OK': 'pending',
                        'PAID': 'approved',
                        'CANCELED': 'cancelled',
                        'REFUNDED': 'refunded',
                        'REJECTED': 'rejected'
                    }
                    db_status = status_map.get(payment_info.get('status', 'OK'), 'pending')
                    
                    # Para admin VIP, não inserir na tabela payments normal
                    if plan_id != 'admin_vip':
                        db.execute_query(
                            """INSERT INTO payments 
                            (payment_id, user_id, plan_id, amount, currency, 
                             payment_method, status, external_reference, 
                             qr_code_data) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                            (
                                payment_info['transactionId'],
                                int(user_id),
                                int(plan_id),
                                amount,
                                'BRL',
                                'pix_automatico',
                                db_status,
                                external_reference,
                                payment_info['pix']['code']
                            ),
                            commit=True
                        )
                    else:
                        # Para admin VIP, inserir na tabela específica
                        db.execute_query(
                            """INSERT INTO admin_vip_payments 
                            (admin_id, amount, description, external_reference, 
                             pix_code, status, expires_at) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                            (
                                int(user_id),
                                amount,
                                'Admin VIP - Upgrade de Acesso',
                                external_reference,
                                payment_info['pix']['code'],
                                'pending',
                                datetime.now() + timedelta(hours=24)  # Expira em 24h
                            ),
                            commit=True
                        )
                    logger.info(f"Pagamento registrado: {payment_info['transactionId']}")
            except Exception as e:
                logger.error(f"Erro ao registrar pagamento: {e}")
            finally:
                db.close()
            
            # Gerar QR code se necessário
            qr_base64 = payment_info['pix'].get('base64', '')
            if not qr_base64:
                qr_base64 = self._generate_qr_code(payment_info['pix']['code'])
            
            return {
                "qr_code": payment_info['pix']['code'],
                "qr_code_base64": qr_base64,
                "payment_id": payment_info['transactionId'],
                "provider": "cnpay",
                "status": db_status,
                "pix_image_url": payment_info['pix'].get('image'),
                "order_url": payment_info.get('order', {}).get('url')
            }
            
        except Exception as e:
            logger.error(f"Erro ao processar resposta: {e}")
            return None
    
    def _generate_qr_code(self, pix_code):
        """Gera QR code localmente em base64"""
        try:
            import qrcode
            from PIL import Image
            import io
            import base64
            
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(pix_code)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='PNG')
            return base64.b64encode(img_buffer.getvalue()).decode('utf-8')
        except Exception as e:
            logger.error(f"Falha ao gerar QR code: {e}")
            return ""
    
    async def check_payment(self, payment_id):
        """Verifica pagamento no CNPay - REMOVIDO pois CNPay só usa webhooks."""
        logger.warning("CNPay não suporta verificação manual de pagamentos. Use apenas webhooks.")
        return None

class PixProviderManager:
    """Gerenciador de provedores PIX"""
    
    def __init__(self, config):
        self.config = config
        self.providers = {}
        self._initialize_providers()
    
    def _initialize_providers(self):
        """Inicializa os provedores disponíveis"""
        # MercadoPago
        if self.config.get('mercadopago_enabled', False):
            self.providers['mercadopago'] = MercadoPagoProvider(self.config)
        
        # CNPay
        if self.config.get('cnpay_enabled', False):
            self.providers['cnpay'] = CNPayProvider(self.config)
    
    def get_default_provider(self):
        """Retorna o provedor padrão"""
        default_provider = self.config.get('pix_provider', 'mercadopago')
        return self.providers.get(default_provider)
    
    def get_available_providers(self):
        """Retorna lista de provedores disponíveis"""
        return list(self.providers.keys())
    
    async def generate_pix_with_fallback(self, amount, description, external_reference):
        """Gera PIX com fallback automático entre provedores"""
        default_provider = self.get_default_provider()
        
        if not default_provider:
            logger.error("Nenhum provedor PIX configurado")
            return None
        
        # Tentar provedor padrão primeiro
        try:
            result = await default_provider.generate_pix(amount, description, external_reference)
            if result:
                logger.info(f"PIX gerado com sucesso usando {result['provider']}")
                return result
        except Exception as e:
            logger.error(f"Erro no provedor padrão: {e}")
        
        # Se falhar, tentar outros provedores
        for provider_name, provider in self.providers.items():
            if provider_name == self.config.get('pix_provider', 'mercadopago'):
                continue  # Já tentou o padrão
            
            try:
                result = await provider.generate_pix(amount, description, external_reference)
                if result:
                    logger.info(f"PIX gerado com fallback usando {result['provider']}")
                    return result
            except Exception as e:
                logger.error(f"Erro no provedor {provider_name}: {e}")
        
        logger.error("Todos os provedores PIX falharam")
        return None
    
    async def check_payment_with_fallback(self, payment_id, provider=None):
        """Verifica pagamento com fallback - CNPay não suporta verificação manual"""
        if provider == 'cnpay':
            logger.info(f"Pagamento CNPay {payment_id} - aguardando webhook (não verificação manual)")
            return None
        
        if provider and provider in self.providers:
            # Não verificar CNPay
            if provider == 'cnpay':
                return None
            return await self.providers[provider].check_payment(payment_id)
        
        # Tentar apenas provedores que suportam verificação manual
        for provider_name, provider_instance in self.providers.items():
            # Pular CNPay pois só usa webhooks
            if provider_name == 'cnpay':
                continue
                
            try:
                result = await provider_instance.check_payment(payment_id)
                if result:
                    return result
            except Exception as e:
                logger.error(f"Erro ao verificar pagamento com {provider_name}: {e}")
        
        return None

# Instância global do gerenciador de provedores
_pix_provider_manager = None

def get_pix_provider_manager():
    """Retorna a instância global do gerenciador de provedores"""
    global _pix_provider_manager
    if _pix_provider_manager is None:
        config = load_config()
        _pix_provider_manager = PixProviderManager(config)
    return _pix_provider_manager

def start_cnpay_webhook():
    from webhook_cnpay import app as webhook_app
    webhook_app.run(host='0.0.0.0', port=8082, debug=False, use_reloader=False)

async def get_user_vip_links(bot, user_id):
    """Busca links de convite VIP para um usuário com assinatura ativa"""
    try:
        # Verificar se o usuário tem assinatura ativa
        active_subscription = get_active_subscription(user_id)
        
        if not active_subscription:
            return None, "Você não possui uma assinatura VIP ativa."
        
        plan_id = active_subscription['plan_id']
        plan_name = active_subscription['plan_name']
        end_date = active_subscription['end_date']
        
        # Buscar grupos associados ao plano
        db = Database()
        try:
            db.connect()
            if not db.connection:
                return None, "Erro ao conectar ao banco de dados."
            
            groups = db.execute_fetch_all(
                """SELECT vg.group_id, vg.group_name
                FROM vip_groups vg
                JOIN plan_groups pg ON vg.id = pg.group_id
                WHERE pg.plan_id = %s AND vg.is_active = TRUE""",
                (plan_id,)
            )
            
            if not groups:
                return None, f"Você tem assinatura ativa no plano {plan_name}, mas nenhum grupo VIP está configurado."
            
            # Gerar links de convite para cada grupo
            links_message = f"⬇ ESTOU PELADINHA TE ESPERANDO 🙈\n\n"
            links_message += f"😈 Clique em \" VER CANAL \" pra gente começar a brincar 🔥\n\n"
            links_message += f"💎 VIP VAZADOS VIP 🍑🔥\n\n"
            links_message += f"📅 **Expira em:** {end_date.strftime('%d/%m/%Y %H:%M')}\n\n"
            links_message += f"📱 **Grupos VIP:**\n\n"
            
            for group in groups:
                group_id = group['group_id']
                group_name = group['group_name']
                
                try:
                    # Verificar se o grupo é válido
                    chat = await bot.get_chat(group_id)
                    if chat.type in ['group', 'supergroup', 'channel']:
                        try:
                            # Criar link de convite
                            invite_link = await bot.create_chat_invite_link(
                                chat_id=group_id,
                                name=f"VIP {user_id} - {plan_name}",
                                expire_date=datetime.now() + timedelta(days=30),
                                member_limit=1,
                                creates_join_request=False
                            )
                            
                            links_message += f"**{group_name}:**\n"
                            links_message += f"`{invite_link.invite_link}`\n\n"
                            
                        except Exception as e:
                            logger.error(f"Erro ao criar link para grupo {group_id}: {e}")
                            # Tentar obter link existente
                            try:
                                invite_link = await bot.export_chat_invite_link(chat_id=group_id)
                                links_message += f"**{group_name}:**\n"
                                links_message += f"`{invite_link}`\n\n"
                            except Exception as e2:
                                logger.error(f"Erro ao obter link existente: {e2}")
                                links_message += f"**{group_name}:** Erro ao gerar link\n\n"
                    else:
                        links_message += f"**{group_name}:** Grupo inválido\n\n"
                        
                except Exception as e:
                    logger.error(f"Erro ao processar grupo {group_id}: {e}")
                    links_message += f"**{group_name}:** Erro ao acessar grupo\n\n"
            
            links_message += "⚠️ **Importante:**\n"
            links_message += "• Cada link pode ser usado apenas uma vez\n"
            links_message += "• Os links expiram em 30 dias\n"
            links_message += "• Use /start para renovar sua assinatura"
            
            return links_message, None
            
        except Exception as e:
            logger.error(f"Erro ao buscar grupos VIP: {e}")
            return None, "Erro ao buscar seus grupos VIP."
        finally:
            db.close()
            
    except Exception as e:
        logger.error(f"Erro ao gerar links VIP: {e}")
        return None, "Erro ao gerar seus links VIP."

# Função utilitária para buscar plano pelo banco
async def get_plan_by_id(plan_id):
    db = Database()
    try:
        db.connect()
        if not db.connection:
            return None
        plan = db.execute_fetch_one("SELECT * FROM vip_plans WHERE id = %s", (plan_id,))
        return plan
    except Exception as e:
        logger.error(f"Erro ao buscar plano {plan_id}: {e}")
        return None
    finally:
        db.close()

async def test_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Comando temporário para testar usuários no banco"""
    config = load_config()

    if config is None:
        logger.error("Falha ao carregar as configurações.")
        return  # ou lidar de forma apropriada
    
    if not is_admin(int(update.effective_user.id)):
        logger.info(f"Usuário {update.effective_user.id} tentou acessar sem permissão.")
        await update.message.reply_text("Acesso negado.")
        return
    
    try:
        # Testar get_all_users
        all_users = get_all_users()
        vip_users = get_vip_users()
        
        # Testar conexão direta
        db = Database()
        db.connect()
        if db.connection:
            direct_count = db.execute_fetch_one("SELECT COUNT(*) as total FROM users")
            direct_users = db.execute_fetch_all("SELECT id, first_name, is_vip FROM users LIMIT 5")
        else:
            direct_count = {'total': 0}
            direct_users = []
        db.close()
        
        # Criar relatório
        report = f"🔍 **TESTE DE USUÁRIOS**\n\n"
        report += f"📊 **Estatísticas:**\n"
        report += f"• get_all_users(): {len(all_users)} usuários\n"
        report += f"• get_vip_users(): {len(vip_users)} usuários VIP\n"
        report += f"• Consulta direta: {direct_count['total']} usuários\n\n"
        
        if direct_users:
            report += f"👥 **Primeiros 5 usuários:**\n"
            for user in direct_users:
                report += f"• ID: {user['id']}, Nome: {user['first_name']}, VIP: {'✅' if user['is_vip'] else '❌'}\n"
        else:
            report += f"❌ **Nenhum usuário encontrado no banco de dados**\n"
        
        await update.message.reply_text(report, parse_mode='Markdown')
        
    except Exception as e:
        logger.error(f"Erro no teste de usuários: {e}")
        await update.message.reply_text(f"❌ Erro: {str(e)}")

# Comandos do bot

if __name__ == '__main__':
    main()


def add_admin(user_id, added_by):
    db = Database()
    conn = db.connect()
    if not conn:
        return False
    try:
        return db.execute_query(
            "INSERT IGNORE INTO admins (user, admin_id) VALUES (%s, %s)",
            (str(user_id), str(added_by)),
            commit=True
        )
    finally:
        db.close()

def remove_admin(user_id):
    db = Database()
    conn = db.connect()
    if not conn:
        return False
    try:
        return db.execute_query(
            "DELETE FROM admins WHERE admin_id = %s",
            (str(user_id),),
            commit=True
        )
    finally:
        db.close()
